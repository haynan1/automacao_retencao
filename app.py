"""Automação de Retenções FMS — camada web (Flask).

Orquestra o fluxo: upload -> análise -> pré-visualização -> mapeamento ->
processamento -> resultado. Toda a regra de negócio vive em services/.
"""

from __future__ import annotations

import io
import os
import re
import tempfile
import traceback
from pathlib import Path
from urllib.parse import urlparse

# Carrega variáveis de um arquivo .env, se existir (opcional).
try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

from flask import (
    Flask,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from services import (
    conferencia,
    historico,
    mapeador,
    molde,
    perfis,
    preenchimento,
    relatorio,
    relatorio_pdf,
)
from services.normalizador import normalizar_texto
from services.parser_listagem import extrair_lancamentos
from services.utils import (
    MODELOS_DIR,
    OUTPUTS_DIR,
    UPLOADS_DIR,
    carregar_sessao,
    configurar_logs,
    criar_pastas,
    formatar_moeda,
    gerar_nome_saida,
    limpar_temporarios,
    novo_id_sessao,
    salvar_sessao,
)

MAX_UPLOAD_MB = 30
# A spec do construtor de molde e texto puro: o molde real da secretaria dá
# ~4 KB. 512 KB é folga de sobra e mantém o corpo JSON longe do teto de upload.
MAX_SPEC_KB = 512
# Teto de um rótulo escolhido no menu. O construtor de molde limita nome de
# setor e de coluna a 200 caracteres e linha de tipo a 60 — acima disto o
# valor não casaria com nada, então nem chega a ser normalizado.
MAX_ROTULO = 300
EXTENSOES_OK = {".xlsx"}
_MIMETYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24).hex())

criar_pastas()
log = configurar_logs()
limpar_temporarios()  # retenção: remove uploads/sessões >24h e saídas >7d


@app.before_request
def _bloqueia_origem_externa():
    """Barra requisições que mudam estado vindas de outra origem (drive-by).

    Como o app não tem auth e roda em localhost, um site aberto no navegador
    poderia POSTar para 127.0.0.1:5000. Só aceitamos POST/etc. cujo Origin
    seja a própria origem do app (ou ausente — curl, form same-origin).
    """
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return None
    origin = request.headers.get("Origin")
    if origin and urlparse(origin).netloc != request.host:
        log.warning("Origem externa bloqueada: %s (host=%s)", origin, request.host)
        abort(403)
    return None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extensao_valida(nome: str) -> bool:
    return os.path.splitext(nome.lower())[1] in EXTENSOES_OK


def _salvar_upload(file_storage, prefixo: str) -> tuple[str, str]:
    """Salva o upload dentro de uploads/ com nome seguro e único.

    Retorna (nome_original, caminho_absoluto). Levanta ValueError se a
    extensão for inválida.
    """
    nome_original = file_storage.filename or ""
    if not _extensao_valida(nome_original):
        raise ValueError(f"Arquivo '{nome_original}' não é .xlsx.")

    base = secure_filename(nome_original) or "arquivo.xlsx"
    destino = UPLOADS_DIR / f"{prefixo}_{base}"
    # Garante que o destino permaneça dentro de uploads/ (defesa em profundidade).
    destino = destino.resolve()
    if UPLOADS_DIR.resolve() not in destino.parents:
        raise ValueError("Caminho de upload inválido.")
    file_storage.save(destino)
    return nome_original, str(destino)


def _sessao_ou_404(sid: str) -> dict:
    dados = carregar_sessao(sid)
    if not dados:
        abort(404)
    return dados


class EscolhaInvalida(ValueError):
    """Escolha que o app não ofereceu no menu. Ver `_escolha_do_menu`."""


def _escolha_do_menu(valor: str, permitidos: list[str], com_ignorar: bool = True) -> str | None:
    """Aceita apenas uma opção que o próprio app colocou no menu.

    Devolve a forma canônica do modelo (não o texto recebido), `None` para
    campo vazio — que significa "manter o automático" — e levanta
    `EscolhaInvalida` para qualquer outra coisa.

    Sem esta guarda, o valor do formulário ia direto para o JSON do perfil.
    Um formulário velho (aba aberta desde antes de o molde mudar) gravava um
    destino inexistente que ninguém escolheu, e ele voltava todo mês como
    pendência sem explicação. E, por ser texto livre, nada impedia que o
    arquivo de vínculos crescesse com lixo do tamanho do corpo do pedido.
    """
    valor = (valor or "").strip()
    if not valor:
        return None
    if com_ignorar and valor == mapeador.IGNORAR:
        return valor
    # Corta pelo comprimento ANTES de normalizar: `normalizar_texto` roda NFKD
    # na string inteira, e um campo de megabytes viraria trabalho de CPU à toa.
    # Nada acima deste tamanho poderia casar — os rótulos do molde são curtos.
    if len(valor) > MAX_ROTULO:
        raise EscolhaInvalida(valor[:80])
    alvo = normalizar_texto(valor)
    for permitido in permitidos:
        if normalizar_texto(permitido) == alvo:
            return permitido
    raise EscolhaInvalida(valor)


def _detectar_modelo(caminho_modelo: str, aba_preferida: str | None):
    """Abre o modelo e detecta setores, colunas, linhas de tipo e abas.

    Retorna (setores, colunas_modelo, tipos_modelo, abas, aba_analisada).
    Uma aba só é oferecida como destino se tiver a MESMA grade da analisada
    — mesmos setores, mesmas colunas e as mesmas linhas de tipo. Trocar de
    aba não pode mudar em silêncio o mapa que acabou de ser confirmado.
    """
    wb = load_workbook(caminho_modelo, data_only=False)
    try:
        abas = list(wb.sheetnames)
        alvo = aba_preferida if aba_preferida in abas else None
        if alvo is None:
            for nome in abas:  # primeira aba que não seja de conferência
                if "CONFER" not in nome.upper():
                    alvo = nome
                    break
        ws = wb[alvo] if alvo else wb.worksheets[0]
        # Uma varredura por aba: ler os três eixos com `listar_*` refaria o
        # mesmo trabalho três vezes, e o modelo pode ter dezenas de abas.
        blocos = preenchimento.localizar_blocos_setores(ws)
        setores = [b["setor"] for b in blocos]
        colunas = preenchimento.colunas_dos_blocos(blocos)
        tipos = preenchimento.tipos_dos_blocos(blocos)
        abas_compativeis = []
        for nome in abas:
            if "CONFER" in nome.upper():
                continue
            blocos_aba = preenchimento.localizar_blocos_setores(wb[nome])
            if (
                [b["setor"] for b in blocos_aba] == setores
                and preenchimento.colunas_dos_blocos(blocos_aba) == colunas
                and preenchimento.tipos_dos_blocos(blocos_aba) == tipos
            ):
                abas_compativeis.append(nome)
        return setores, colunas, tipos, abas_compativeis or ([alvo] if alvo else abas), alvo
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Rotas
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    lista = perfis.listar_perfis()
    moldes = {p["slug"]: perfis.info_molde(p["slug"]) for p in lista}
    return render_template(
        "index.html",
        max_mb=MAX_UPLOAD_MB,
        perfis=lista,
        perfil_padrao=perfis.slug_padrao(),
        moldes=moldes,
    )


@app.route("/secretarias")
def secretarias():
    lista = perfis.listar_perfis()
    return render_template(
        "secretarias.html",
        perfis=lista,
        padrao=perfis.slug_padrao(),
        moldes={p["slug"]: perfis.info_molde(p["slug"]) for p in lista},
    )


@app.route("/secretarias/salvar", methods=["POST"])
def secretarias_salvar():
    nome = (request.form.get("nome") or "").strip()
    slug = (request.form.get("slug") or "").strip()
    deteccao = [
        termo.strip()
        for termo in re.split(r"[,\n;]", request.form.get("deteccao") or "")
        if termo.strip()
    ]
    if not nome:
        return render_template(
            "erro.html",
            titulo="Nome obrigatorio",
            mensagem="Informe o nome da secretaria.",
        ), 400

    slug_final = slug if (slug and perfis.perfil_valido(slug)) else None
    perfis.registrar_perfil(nome, deteccao=deteccao, slug=slug_final)
    return redirect(url_for("secretarias"))


@app.route("/secretarias/<slug>/padrao", methods=["POST"])
def secretarias_padrao(slug):
    perfis.definir_padrao(slug)
    return redirect(url_for("secretarias"))


@app.route("/secretarias/<slug>/remover", methods=["POST"])
def secretarias_remover(slug):
    if not perfis.remover_perfil(slug):
        return render_template(
            "erro.html",
            titulo="Nao foi possivel remover",
            mensagem="Nao da para remover o perfil padrao nem o ultimo.",
        ), 400
    return redirect(url_for("secretarias"))


@app.route("/secretarias/<slug>/molde/limpar", methods=["POST"])
def secretarias_limpar_molde(slug):
    if not perfis.perfil_valido(slug):
        return render_template(
            "erro.html",
            titulo="Secretaria invalida",
            mensagem="Nao foi possivel localizar a secretaria informada.",
        ), 400
    perfis.remover_molde(slug)
    return redirect(url_for("secretarias"))


@app.route("/secretarias/<slug>/molde/baixar")
def secretarias_baixar_molde(slug):
    """Baixa o molde fixo da secretaria (para abrir no Excel ou reaproveitar)."""
    if not perfis.perfil_valido(slug) or not perfis.existe_molde(slug):
        abort(404)
    caminho = perfis.caminho_molde(slug).resolve()
    if MODELOS_DIR.resolve() not in caminho.parents:  # guarda anti path traversal
        abort(403)
    info = perfis.info_molde(slug) or {}
    nome = secure_filename(info.get("nome_original") or "") or f"molde_{slug}.xlsx"
    if not nome.lower().endswith(".xlsx"):
        nome += ".xlsx"
    return send_file(caminho, as_attachment=True, download_name=nome)


# ---------------------------------------------------------------------------
# Construtor de molde — a estrutura da planilha desenhada na interface
# ---------------------------------------------------------------------------

def _perfil_ou_404(slug: str) -> str:
    if not perfis.perfil_valido(slug):
        abort(404)
    return slug


def _spec_do_pedido() -> dict:
    """Le a spec do corpo JSON. Corpo ausente/invalido vira ErroDeMolde.

    O teto de 30 MB do app existe para planilhas; uma spec e texto e nunca
    passa de alguns KB. Recusar pelo Content-Length evita desserializar
    megabytes de JSON so para descobrir depois que era absurdo.
    """
    if (request.content_length or 0) > MAX_SPEC_KB * 1024:
        raise molde.ErroDeMolde(
            [f"Estrutura do molde grande demais (limite de {MAX_SPEC_KB} KB)."]
        )
    corpo = request.get_json(silent=True)
    if not isinstance(corpo, dict) or not isinstance(corpo.get("spec"), dict):
        raise molde.ErroDeMolde(["Estrutura do molde não recebida."])
    return corpo["spec"]


@app.route("/molde/<slug>")
def molde_editor(slug):
    """Tela do construtor: abre o desenho salvo, o molde atual ou uma folha em branco."""
    _perfil_ou_404(slug)
    nome = perfis.nome_perfil(slug)
    spec, origem = molde.spec_inicial(slug, nome)
    return render_template(
        "molde.html",
        slug=slug,
        perfil_nome=nome,
        spec=spec,
        origem=origem,
        spec_branco=molde.spec_em_branco(nome),
        tem_molde=perfis.existe_molde(slug),
        info_molde=perfis.info_molde(slug),
        tipos_sugeridos=list(molde.TIPOS_SUGERIDOS),
        limites={
            "abas": molde.MAX_ABAS,
            "setores": molde.MAX_SETORES,
            "colunas": molde.MAX_COLUNAS,
            "tipos": molde.MAX_TIPOS,
        },
    )


@app.route("/molde/<slug>/importar")
def molde_importar(slug):
    """Devolve a spec extraida do molde fixo atual (para editar em vez de redesenhar)."""
    _perfil_ou_404(slug)
    if not perfis.existe_molde(slug):
        return jsonify({"ok": False, "problemas": ["Esta secretaria ainda não tem molde fixo."]}), 404
    try:
        return jsonify({"ok": True, "spec": molde.extrair_spec(perfis.caminho_molde(slug))})
    except molde.ErroDeMolde as exc:
        return jsonify({"ok": False, "problemas": exc.problemas}), 422
    except Exception as exc:  # noqa: BLE001
        log.error("Falha ao importar molde do perfil '%s': %s", slug, exc)
        return jsonify({"ok": False, "problemas": ["Não foi possível ler o molde atual."]}), 500


@app.route("/molde/<slug>/previa", methods=["POST"])
def molde_previa(slug):
    """Valida o desenho e devolve a prévia — sem gravar nada em disco."""
    _perfil_ou_404(slug)
    try:
        spec, divergencias = molde.conferir(_spec_do_pedido())
    except molde.ErroDeMolde as exc:
        return jsonify({"ok": False, "problemas": exc.problemas, "divergencias": []}), 200
    except Exception as exc:  # noqa: BLE001
        log.error("Falha na prévia do molde '%s': %s\n%s", slug, exc, traceback.format_exc())
        return jsonify({"ok": False, "problemas": ["Erro inesperado ao montar a prévia."],
                        "divergencias": []}), 500

    return jsonify({
        "ok": not divergencias,
        "problemas": [],
        "divergencias": divergencias,
        "spec": spec,
        "resumo": molde.resumo(spec),
        "grade": molde.grade(spec),
    })


@app.route("/molde/<slug>/baixar", methods=["POST"])
def molde_baixar(slug):
    """Gera o .xlsx e devolve para download, sem tocar no molde fixo."""
    _perfil_ou_404(slug)
    try:
        _spec, wb, divergencias = molde.construir(_spec_do_pedido())
    except molde.ErroDeMolde as exc:
        return jsonify({"ok": False, "problemas": exc.problemas}), 422

    if divergencias:
        return jsonify({"ok": False, "problemas": divergencias}), 422

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"molde_{secure_filename(slug) or 'secretaria'}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/molde/<slug>/salvar", methods=["POST"])
def molde_salvar(slug):
    """Fixa o molde desenhado como padrão da secretaria.

    Só grava depois que o motor de preenchimento releu o arquivo gerado e
    confirmou setores, colunas e tipos. `perfis.definir_molde` guarda um
    backup timestampado do molde anterior.
    """
    _perfil_ou_404(slug)
    try:
        spec, wb, divergencias = molde.construir(_spec_do_pedido())
    except molde.ErroDeMolde as exc:
        return jsonify({"ok": False, "problemas": exc.problemas}), 422

    if divergencias:
        log.warning("Molde do perfil '%s' recusado na verificação: %s", slug, divergencias)
        return jsonify({"ok": False, "problemas": divergencias}), 422

    try:
        with tempfile.TemporaryDirectory() as pasta:
            caminho = Path(pasta) / "molde_construido.xlsx"
            wb.save(caminho)
            perfis.definir_molde(slug, caminho, f"molde_construido_{slug}.xlsx")
        molde.salvar_estrutura(slug, spec)
    except OSError as exc:
        log.error("Falha ao gravar molde do perfil '%s': %s", slug, exc)
        return jsonify({"ok": False, "problemas": ["Não foi possível gravar o molde em disco."]}), 500

    log.info("Molde construído para o perfil '%s': %d setor(es), %d coluna(s), %d aba(s).",
             slug, len(spec["setores"]), len(spec["colunas"]), len(spec["abas"]))
    return jsonify({"ok": True, "redirect": url_for("secretarias")})


@app.route("/analisar", methods=["POST"])
def analisar():
    try:
        f_origem = request.files.get("listagem")
        if not f_origem or not f_origem.filename:
            return render_template(
                "erro.html",
                titulo="Listagem faltando",
                mensagem="Envie a Listagem de Eventos (XLSX) exportada do sistema.",
            ), 400

        sid = novo_id_sessao()
        nome_origem, path_origem = _salvar_upload(f_origem, f"{sid}_origem")

        # --- Perfil (secretaria) ---
        perfil = (request.form.get("perfil") or "").strip() or perfis.slug_padrao()
        if not perfis.perfil_valido(perfil):
            perfil = perfis.slug_padrao()

        # --- Origem do molde: fixo (versionado por perfil) ou upload ---
        fonte_modelo = (request.form.get("fonte_modelo") or "").strip()
        f_modelo = request.files.get("modelo")
        usar_fixo = fonte_modelo == "fixo" or (
            not fonte_modelo and perfis.existe_molde(perfil) and (not f_modelo or not f_modelo.filename)
        )

        if usar_fixo:
            if not perfis.existe_molde(perfil):
                return render_template(
                    "erro.html",
                    titulo="Sem molde fixo",
                    mensagem="Este perfil não tem molde fixo definido. Envie uma planilha modelo.",
                ), 400
            info = perfis.info_molde(perfil) or {}
            nome_modelo = info.get("nome_original", "molde_padrao.xlsx")
            path_modelo = str(perfis.caminho_molde(perfil))
        else:
            if not f_modelo or not f_modelo.filename:
                return render_template(
                    "erro.html",
                    titulo="Molde faltando",
                    mensagem="Envie a planilha modelo (RETENÇÃO) ou escolha o molde fixo.",
                ), 400
            nome_modelo, path_modelo = _salvar_upload(f_modelo, f"{sid}_modelo")
            if request.form.get("definir_molde_fixo") == "on":
                perfis.definir_molde(perfil, path_modelo, nome_modelo)
                log.info("Molde fixo do perfil '%s' atualizado: %s", perfil, nome_modelo)

        # --- Parsing do relatório de origem ---
        cfg = mapeador.carregar_config_rubricas()
        regras_rubricas = cfg["regras"]
        fora_de_escopo = cfg["fora_de_escopo"]
        vinculos = mapeador.carregar_vinculos(perfil)
        vinculos_folhas = mapeador.carregar_vinculos_folhas(perfil)

        resultado = extrair_lancamentos(path_origem)
        lancamentos = resultado["lancamentos"]

        # Autodeteccao de secretaria pelo cabecalho — apenas aviso, nao bloqueia.
        detectado = perfis.detectar_perfil(resultado.get("banners", []))
        aviso_perfil = None
        if detectado and detectado != perfil:
            aviso_perfil = (
                f"A Listagem parece ser de '{perfis.nome_perfil(detectado)}', "
                f"mas você selecionou '{perfis.nome_perfil(perfil)}'. Confira o perfil."
            )

        if not resultado.get("faixas"):
            return render_template(
                "erro.html",
                titulo="Estrutura não reconhecida",
                mensagem=(
                    "Não foi possível localizar os cabeçalhos de lançamento "
                    "(Descrição/Valor) no arquivo de origem. Confirme que é a "
                    "Listagem de Eventos exportada em XLSX."
                ),
            ), 422

        if not lancamentos:
            return render_template(
                "erro.html",
                titulo="Nenhum lançamento encontrado",
                mensagem="O relatório foi lido, mas não há linhas de lançamento válidas.",
            ), 422

        if not resultado["lotacoes"]:
            return render_template(
                "erro.html",
                titulo="Nenhuma lotação detectada",
                mensagem="Não foram encontradas lotações no formato '5.0000.0000 - FMS, ...'.",
            ), 422

        # --- Detecta setores/colunas/abas do modelo ---
        aba_sugerida = None
        wb_tmp = load_workbook(path_modelo, data_only=False, read_only=True)
        try:
            aba_sugerida = preenchimento.localizar_aba_destino(wb_tmp, resultado["competencia"])
        finally:
            wb_tmp.close()

        setores_planilha, colunas_modelo, tipos_modelo, abas_modelo, _ = _detectar_modelo(
            path_modelo, aba_sugerida
        )
        if not setores_planilha:
            return render_template(
                "erro.html",
                titulo="Modelo sem blocos de setor",
                mensagem=(
                    "A planilha modelo não apresentou blocos de setor no formato "
                    "esperado (nome do setor seguido de uma linha 'Tipo | rubricas...')."
                ),
            ), 422

        # Sem linhas de tipo não há onde escrever: o fluxo inteiro terminaria
        # numa planilha vazia. Recusar aqui custa um clique; descobrir no fim
        # custa a conferência toda.
        if not tipos_modelo:
            return render_template(
                "erro.html",
                titulo="Modelo sem linhas de tipo de folha",
                mensagem=(
                    "Os blocos da planilha modelo não têm nenhuma linha de tipo "
                    "(Mensal, 13º salário, Férias…) entre o cabeçalho 'Tipo' e a "
                    "linha TOTAL. Sem elas não há onde lançar os valores — use o "
                    "construtor de molde da secretaria para desenhá-las."
                ),
            ), 422

        # --- Mapeamentos: eixo A (lotação→setor), B (evento→coluna), C (folha→linha) ---
        mapa_lotacoes = mapeador.carregar_mapeamento_lotacoes(perfil)
        mapeador.aplicar_mapeamentos(lancamentos, mapa_lotacoes, regras_rubricas)
        mapeador.resolver_colunas(lancamentos, colunas_modelo, fora_de_escopo, vinculos)
        mapeador.resolver_tipos(lancamentos, tipos_modelo, vinculos_folhas)
        pendencias = mapeador.detectar_pendencias(lancamentos)

        lotacoes_info = []
        for lot in resultado["lotacoes"]:
            setor = mapeador.mapear_lotacao(lot, mapa_lotacoes)
            sugestao = setor or mapeador.sugerir_setor(lot, setores_planilha)
            lotacoes_info.append(
                {"lotacao": lot, "setor_atual": setor, "sugestao": sugestao, "mapeada": bool(setor)}
            )

        grupos = mapeador.construir_grupos_rubricas(
            lancamentos, colunas_modelo, fora_de_escopo, vinculos
        )
        grupos_folhas = mapeador.construir_grupos_folhas(
            lancamentos, tipos_modelo, vinculos_folhas
        )
        rubricas_identificadas = sorted(
            {g["coluna"] for g in grupos if g["coluna"]}
        )

        dados = {
            "id": sid,
            "perfil": perfil,
            "perfil_nome": perfis.nome_perfil(perfil),
            "aviso_perfil": aviso_perfil,
            "competencia": resultado["competencia"],
            "aba_origem": resultado["aba"],
            "arquivo_origem_nome": nome_origem,
            "arquivo_origem_path": path_origem,
            "arquivo_modelo_nome": nome_modelo,
            "arquivo_modelo_path": path_modelo,
            "lancamentos": lancamentos,
            "lotacoes": resultado["lotacoes"],
            "lotacoes_info": lotacoes_info,
            "folhas": resultado["folhas"],
            "setores_planilha": setores_planilha,
            "colunas_modelo": colunas_modelo,
            "tipos_modelo": tipos_modelo,
            "fora_de_escopo": fora_de_escopo,
            "grupos": grupos,
            "grupos_folhas": grupos_folhas,
            "abas_modelo": abas_modelo,
            "aba_sugerida": aba_sugerida,
            "rubricas_identificadas": rubricas_identificadas,
            "pendencias": pendencias,
        }
        salvar_sessao(sid, dados)
        log.info("Análise concluída sid=%s lançamentos=%d lotações=%d eventos=%d folhas=%d",
                 sid, len(lancamentos), len(resultado["lotacoes"]), len(grupos),
                 len(grupos_folhas))
        return redirect(url_for("preview", sid=sid))

    except ValueError as exc:
        log.warning("Validação falhou: %s", exc)
        return render_template("erro.html", titulo="Arquivo inválido", mensagem=str(exc)), 400
    except Exception as exc:  # noqa: BLE001 — nunca deixar a app cair
        log.error("Falha na análise: %s\n%s", exc, traceback.format_exc())
        return render_template(
            "erro.html",
            titulo="Erro ao analisar os arquivos",
            mensagem="Ocorreu um erro inesperado. Detalhes técnicos foram registrados em logs/app.log.",
        ), 500


@app.route("/preview/<sid>")
def preview(sid):
    dados = _sessao_ou_404(sid)
    p = dados["pendencias"]
    return render_template(
        "preview.html",
        sid=sid,
        perfil_nome=dados.get("perfil_nome"),
        aviso_perfil=dados.get("aviso_perfil"),
        competencia=dados["competencia"],
        qtd_lancamentos=len(dados["lancamentos"]),
        qtd_lotacoes=len(dados["lotacoes"]),
        rubricas=dados["rubricas_identificadas"],
        lotacoes_info=dados["lotacoes_info"],
        grupos_folhas=dados.get("grupos_folhas", []),
        lotacoes_nao_mapeadas=p["lotacoes_nao_mapeadas"],
        rubricas_sem_vinculo=p["rubricas_sem_vinculo"],
        rubricas_fora_escopo=p["rubricas_fora_escopo"],
        folhas_sem_vinculo=p["folhas_sem_vinculo"],
        folhas_sugeridas=p["folhas_sugeridas"],
    )


@app.route("/mapeamento/<sid>")
def mapeamento(sid):
    dados = _sessao_ou_404(sid)
    return render_template(
        "mapeamento.html",
        sid=sid,
        lotacoes_info=dados["lotacoes_info"],
        setores=dados["setores_planilha"],
        colunas=dados["colunas_modelo"],
        tipos=dados.get("tipos_modelo", []),
        grupos=dados["grupos"],
        grupos_folhas=dados.get("grupos_folhas", []),
        abas=dados["abas_modelo"],
        aba_sugerida=dados["aba_sugerida"],
        competencia=dados["competencia"],
        IGNORAR=mapeador.IGNORAR,
    )


@app.route("/processar/<sid>", methods=["POST"])
def processar(sid):
    dados = _sessao_ou_404(sid)
    try:
        lancamentos = dados["lancamentos"]
        lotacoes = dados["lotacoes"]
        grupos = dados["grupos"]
        grupos_folhas = dados.get("grupos_folhas", [])
        fora_de_escopo = dados["fora_de_escopo"]
        perfil = dados.get("perfil") or perfis.slug_padrao()

        # Toda escolha é conferida contra o que o app ofereceu. As listas da
        # sessão valem para qualquer aba de destino porque só entram em
        # `abas_modelo` abas com a MESMA grade (ver `_detectar_modelo`).
        vinculos = mapeador.carregar_vinculos(perfil)
        vinculos_folhas = mapeador.carregar_vinculos_folhas(perfil)
        try:
            # --- Eixo A: lotação → setor (por índice, robusto a acentos) ---
            overrides_lot: dict[str, str] = {}
            for i, lot in enumerate(lotacoes):
                escolha = _escolha_do_menu(
                    request.form.get(f"setor_{i}"), dados["setores_planilha"],
                    com_ignorar=False,
                )
                if escolha:
                    overrides_lot[lot] = escolha

            # --- Eixo B: evento → coluna (menu suspenso), aprendido ---
            for i, g in enumerate(grupos):
                escolha = _escolha_do_menu(
                    request.form.get(f"coluna_{i}"), dados["colunas_modelo"]
                )
                if escolha:  # vazio = manter automático
                    vinculos[g["chave"]] = escolha

            # --- Eixo C: folha → linha de tipo (menu suspenso), aprendido ---
            for i, g in enumerate(grupos_folhas):
                escolha = _escolha_do_menu(
                    request.form.get(f"tipo_{i}"), dados.get("tipos_modelo", [])
                )
                if escolha:
                    vinculos_folhas[g["chave"]] = escolha
        except EscolhaInvalida as exc:
            log.warning("Escolha fora do menu em sid=%s: %.80s", sid, exc)
            return render_template(
                "erro.html", titulo="Escolha inválida",
                mensagem=(
                    "Uma das opções enviadas não existe nesta planilha modelo. "
                    "Isso costuma acontecer quando a página ficou aberta e o molde "
                    "mudou desde então. Recarregue o mapeamento e escolha de novo."
                ),
            ), 400

        for reg in lancamentos:
            if reg["lotacao_original"] in overrides_lot:
                reg["setor_destino"] = overrides_lot[reg["lotacao_original"]]

        # --- Aba de destino ---
        aba_destino = (request.form.get("aba_destino") or "").strip()
        if aba_destino not in dados["abas_modelo"]:
            return render_template(
                "erro.html", titulo="Aba inválida",
                mensagem="Selecione uma aba de destino existente na planilha modelo.",
            ), 400

        # --- Abre o modelo e usa as colunas REAIS da aba escolhida ---
        wb = load_workbook(dados["arquivo_modelo_path"], data_only=False)
        ws = wb[aba_destino]
        blocos = preenchimento.localizar_blocos_setores(ws)
        colunas_modelo = preenchimento.listar_colunas_modelo(ws)
        tipos_modelo = preenchimento.listar_tipos_modelo(ws)

        # Re-resolve os dois eixos da grade com as colunas e linhas REAIS da
        # aba escolhida e os vínculos que acabaram de ser confirmados.
        mapeador.resolver_colunas(lancamentos, colunas_modelo, fora_de_escopo, vinculos)
        mapeador.resolver_tipos(lancamentos, tipos_modelo, vinculos_folhas)

        # Persistir mapeamentos aprendidos no perfil, se solicitado.
        if request.form.get("salvar_mapeamento") == "on":
            if overrides_lot:
                mapa = mapeador.carregar_mapeamento_lotacoes(perfil)
                mapa.update(overrides_lot)
                mapeador.salvar_mapeamento_lotacoes(perfil, mapa)
            mapeador.salvar_vinculos(perfil, vinculos)
            mapeador.salvar_vinculos_folhas(perfil, vinculos_folhas)
            log.info("Perfil '%s': %d lotações, %d eventos e %d folhas salvos.",
                     perfil, len(overrides_lot), len(vinculos), len(vinculos_folhas))

        # Grupos reconstruídos DEPOIS da decisão: é este o retrato que vai
        # para a conferência e para a tela final — o que de fato aconteceu,
        # não o que havia sido sugerido antes de você escolher.
        decisoes_rubricas = mapeador.construir_grupos_rubricas(
            lancamentos, colunas_modelo, fora_de_escopo, vinculos
        )
        decisoes_folhas = mapeador.construir_grupos_folhas(
            lancamentos, tipos_modelo, vinculos_folhas
        )

        # --- Agregação e preenchimento ---
        agregados, baldes = conferencia.agregar_lancamentos(lancamentos)
        preenchimento.limpar_area_lancamento(ws, blocos)
        relatorio = preenchimento.preencher_valores(ws, agregados, blocos)

        # --- Totais, reconciliação e conferência ---
        total_lido = conferencia.calcular_totais_lidos(lancamentos)
        agregados_tot = conferencia.calcular_totais_agregados(agregados)
        # Por evento sai por fora: 'INSS' e 'INSS do 13º' somam separados
        # aqui e juntos na coluna. É esta série que alimenta o detalhamento
        # por evento no relatório consolidado do histórico.
        agregados_tot["por_evento"] = conferencia.totais_por_evento(lancamentos)
        rec = conferencia.reconciliar(total_lido, baldes, relatorio["pendencias_estrutura"])
        pendencias_map = mapeador.detectar_pendencias(lancamentos)

        dados_conf = {
            "competencia": dados["competencia"],
            "arquivo_origem": dados["arquivo_origem_nome"],
            "arquivo_modelo": dados["arquivo_modelo_nome"],
            "aba_destino": aba_destino,
            "reconciliacao": rec,
            "por_setor": agregados_tot["por_setor"],
            "por_coluna": agregados_tot["por_coluna"],
            "por_tipo": agregados_tot["por_tipo"],
            "decisoes_rubricas": decisoes_rubricas,
            "decisoes_folhas": decisoes_folhas,
            **{chave: pendencias_map[chave] for chave in (
                "lotacoes_nao_mapeadas", "rubricas_sem_vinculo", "rubricas_fora_escopo",
                "folhas_sem_vinculo", "folhas_sugeridas", "folhas_fora_escopo",
            )},
            "pendencias_estrutura": relatorio["pendencias_estrutura"],
        }
        conferencia.criar_aba_conferencia(wb, dados_conf)
        preenchimento.preservar_formulas(wb)

        # --- Salvar saída ---
        nome_saida = gerar_nome_saida()
        caminho_saida = (OUTPUTS_DIR / nome_saida).resolve()
        if OUTPUTS_DIR.resolve() not in caminho_saida.parents:
            raise ValueError("Caminho de saída inválido.")
        wb.save(caminho_saida)
        wb.close()

        dados["output_nome"] = nome_saida
        dados["output_path"] = str(caminho_saida)
        dados["resumo"] = {
            "aba_destino": aba_destino,
            "reconciliacao": rec,
            "por_setor": agregados_tot["por_setor"],
            "por_coluna": agregados_tot["por_coluna"],
            "por_tipo": agregados_tot["por_tipo"],
            "por_evento": agregados_tot["por_evento"],
            "qtd_celulas": len(relatorio["preenchidos"]),
            "pendencias_estrutura": relatorio["pendencias_estrutura"],
            "decisoes_rubricas": decisoes_rubricas,
            "decisoes_folhas": decisoes_folhas,
            **{chave: pendencias_map[chave] for chave in (
                "lotacoes_nao_mapeadas", "rubricas_sem_vinculo", "rubricas_fora_escopo",
                "rubricas_por_regra", "folhas_sem_vinculo", "folhas_sugeridas",
                "folhas_fora_escopo",
            )},
        }
        salvar_sessao(sid, dados)

        # --- Histórico local (nunca vai para o git) ---
        # O histórico precisa fechar sozinho, sem a sessão:
        #   lido = preenchido + estrutura + fora de escopo + pendente
        # Guardar só o fora de escopo das rubricas deixava de fora o das folhas,
        # e omitir 'estrutura' escondia o que não achou lugar na planilha — o
        # relatório consolidado acusaria uma diferença que não existe.
        pendente = (rec["total_sem_vinculo"] + rec["total_folha_sem_vinculo"]
                    + rec["total_setor_nao_mapeado"])
        fora_escopo = rec["total_fora_escopo"] + rec["total_folha_fora_escopo"]
        graficos_hist = {}
        # "Evento" entra aqui para que o relatório consolidado consiga
        # detalhar por evento do relatório de origem, e não só pela
        # coluna de destino em que vários eventos se somam.
        for gnome, gchave in (("Setor", "por_setor"), ("Rubrica", "por_coluna"),
                              ("Tipo", "por_tipo"), ("Evento", "por_evento")):
            gserie = _serie_grafico(agregados_tot.get(gchave))
            if gserie:
                graficos_hist[gnome] = gserie
        historico.registrar_operacao({
            "sid": sid,
            "graficos": graficos_hist,
            "perfil": perfil,
            "perfil_nome": dados.get("perfil_nome"),
            "competencia": dados["competencia"],
            "aba_destino": aba_destino,
            "arquivo_origem": dados["arquivo_origem_nome"],
            "arquivo_modelo": dados["arquivo_modelo_nome"],
            "output_nome": nome_saida,
            "qtd_lancamentos": len(lancamentos),
            "qtd_celulas": len(relatorio["preenchidos"]),
            "total_lido": rec["total_lido"],
            "total_preenchido": rec["total_preenchido"],
            "total_fora_escopo": fora_escopo,
            "total_estrutura": rec["total_estrutura"],
            "total_pendente": pendente,
            "confere": rec["confere"],
            "pendencias": {
                "lotacoes_nao_mapeadas": len(pendencias_map["lotacoes_nao_mapeadas"]),
                "rubricas_sem_vinculo": len(pendencias_map["rubricas_sem_vinculo"]),
                "folhas_sem_vinculo": len(pendencias_map["folhas_sem_vinculo"]),
            },
        })

        log.info("Processamento concluído sid=%s células=%d confere=%s saída=%s",
                 sid, len(relatorio["preenchidos"]), rec["confere"], nome_saida)
        return redirect(url_for("resultado", sid=sid))

    except KeyError as exc:
        log.error("Aba/estrutura ausente: %s", exc)
        return render_template(
            "erro.html", titulo="Aba não encontrada",
            mensagem=f"Não foi possível abrir a aba selecionada: {exc}.",
        ), 422
    except Exception as exc:  # noqa: BLE001
        log.error("Falha no processamento: %s\n%s", exc, traceback.format_exc())
        return render_template(
            "erro.html",
            titulo="Erro ao preencher a planilha",
            mensagem="Ocorreu um erro inesperado. Detalhes foram registrados em logs/app.log.",
        ), 500


def _serie_grafico(d: dict | None) -> list[dict]:
    """Transforma {rótulo: Decimal} em série [{label, valor(float)}], ordenada
    por valor (desc). float para serializar em JSON no gráfico SVG."""
    if not d:
        return []
    return [
        {"label": k, "valor": float(v)}
        for k, v in sorted(d.items(), key=lambda kv: kv[1], reverse=True)
    ]


@app.route("/resultado/<sid>")
def resultado(sid):
    dados = _sessao_ou_404(sid)
    if "resumo" not in dados:
        return redirect(url_for("mapeamento", sid=sid))
    resumo = dados["resumo"]
    graficos = {}
    for nome, chave in (("Setor", "por_setor"), ("Rubrica", "por_coluna"), ("Tipo", "por_tipo")):
        serie = _serie_grafico(resumo.get(chave))
        if serie:
            graficos[nome] = serie
    return render_template(
        "resultado.html",
        sid=sid,
        competencia=dados["competencia"],
        output_nome=dados["output_nome"],
        resumo=resumo,
        graficos=graficos,
    )


@app.route("/download/<sid>")
def download(sid):
    dados = _sessao_ou_404(sid)
    caminho = dados.get("output_path")
    if not caminho or not os.path.exists(caminho):
        abort(404)
    # Garante que o arquivo servido está dentro de outputs/ (anti path traversal).
    resolvido = Path(caminho).resolve()
    if OUTPUTS_DIR.resolve() not in resolvido.parents:
        abort(403)
    return send_file(resolvido, as_attachment=True, download_name=dados["output_nome"])


@app.route("/historico")
def historico_view():
    operacoes = historico.listar_operacoes()
    # Marca quais saídas ainda existem em disco (podem ter sido limpas após 7 dias).
    for op in operacoes:
        nome = op.get("output_nome") or ""
        op["disponivel"] = bool(nome) and (OUTPUTS_DIR / nome).exists()
    return render_template("historico.html", operacoes=operacoes)


@app.route("/guia")
def guia():
    return render_template("guia.html")


@app.route("/historico/<op_id>/remover", methods=["POST"])
def historico_remover(op_id):
    historico.remover_operacao(op_id)
    return redirect(url_for("historico_view"))


@app.route("/historico/<op_id>/graficos")
def historico_graficos(op_id):
    op = historico.buscar_operacao(op_id)
    if not op:
        abort(404)
    return jsonify(op.get("graficos") or {})


@app.route("/historico/exportar", methods=["POST"])
def historico_exportar():
    """Gera o relatório consolidado das operações marcadas na tela.

    O arquivo é montado em memória e entregue direto: nada é gravado em
    `outputs/`, que é varrido a cada 7 dias e existe para as planilhas
    preenchidas — um relatório é derivado, refazê-lo custa um clique.
    """
    ids = request.form.getlist("ids")
    if len(ids) > relatorio.MAX_OPERACOES:
        ids = ids[:relatorio.MAX_OPERACOES]

    operacoes = historico.buscar_operacoes(ids)
    if not operacoes:
        return render_template(
            "erro.html", titulo="Nenhuma operação selecionada",
            mensagem=(
                "Marque ao menos uma operação no histórico para gerar o relatório. "
                "Se você marcou e mesmo assim chegou aqui, a operação pode ter sido "
                "apagada nesta ou em outra aba."
            ),
        ), 400

    formato = "pdf" if (request.form.get("formato") or "").strip() == "pdf" else "xlsx"
    try:
        if formato == "pdf":
            buffer = io.BytesIO(relatorio_pdf.montar_pdf(operacoes))
        else:
            buffer = io.BytesIO()
            relatorio.montar_relatorio(operacoes).save(buffer)
    except relatorio.SemOperacoes as exc:
        return render_template(
            "erro.html", titulo="Nada a exportar", mensagem=str(exc),
        ), 400
    except Exception as exc:  # noqa: BLE001 — nunca deixar a app cair
        # O histórico é append-only e atravessa versões: um registro antigo
        # pode ter uma forma que a montagem não previu. Isso não pode virar
        # um 500 cru — o resto do app continua utilizável.
        log.error("Falha ao gerar o relatório consolidado (%s): %s\n%s",
                  formato, exc, traceback.format_exc())
        return render_template(
            "erro.html", titulo="Erro ao gerar o relatório",
            mensagem=("Não foi possível montar o relatório com as operações selecionadas. "
                      "Detalhes técnicos foram registrados em logs/app.log."),
        ), 500

    buffer.seek(0)
    log.info("Relatório consolidado gerado em %s: %d operação(ões).", formato, len(operacoes))
    return send_file(
        buffer,
        as_attachment=True,
        download_name=relatorio.nome_arquivo(formato),
        mimetype=_MIMETYPES[formato],
    )


@app.route("/historico/limpar", methods=["POST"])
def historico_limpar():
    n = historico.limpar_historico()
    log.info("Histórico limpo: %d operação(ões) removida(s).", n)
    return redirect(url_for("historico_view"))


@app.route("/historico/baixar/<nome>")
def historico_baixar(nome):
    seguro = secure_filename(nome)
    if not seguro:
        abort(404)
    caminho = (OUTPUTS_DIR / seguro).resolve()
    if OUTPUTS_DIR.resolve() not in caminho.parents or not caminho.exists():
        abort(404)
    return send_file(caminho, as_attachment=True, download_name=seguro)


@app.errorhandler(413)
def arquivo_grande(_):
    return render_template(
        "erro.html",
        titulo="Arquivo muito grande",
        mensagem=f"O limite de upload é {MAX_UPLOAD_MB} MB por arquivo.",
    ), 413


@app.errorhandler(404)
def nao_encontrado(_):
    return render_template(
        "erro.html", titulo="Não encontrado",
        mensagem="A página ou sessão de trabalho não foi encontrada. Recomece o envio.",
    ), 404


@app.template_filter("moeda")
def _moeda(valor) -> str:
    """Filtro das telas. A formatação vive em services.utils, uma só, para
    que o número da tela e o do PDF nunca divirjam."""
    return formatar_moeda(valor)


if __name__ == "__main__":
    # Debug desligado por padrão (o debugger do Werkzeug expõe tracebacks e
    # execução de código). Para desenvolvimento: FLASK_DEBUG=1 python app.py
    debug = os.environ.get("FLASK_DEBUG", "").lower() in ("1", "true", "yes")
    app.run(host="127.0.0.1", port=5000, debug=debug)
