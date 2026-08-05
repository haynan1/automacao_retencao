"""Vocabulario visual das planilhas que o app GERA (nao das que preenche).

Duas superficies usam este vocabulario: a aba `CONFERÊNCIA_AUTOMAÇÃO` que
acompanha cada arquivo preenchido (services/conferencia.py) e o relatorio
consolidado do historico (services/relatorio.py). Elas saem da mesma mao —
mesma paleta, mesma hierarquia, mesmo formato de moeda — porque quem recebe
as duas nao deveria conseguir dizer que foram escritas por codigos
diferentes.

Aqui ficam SO os atomos (cor, fonte, borda) e as tres faixas de hierarquia
(titulo, secao, cabecalho de tabela). O corpo de cada tabela e diferente em
cada documento e continua no modulo de quem a escreve.

Nada aqui toca no molde de Retencao: aquele imita o arquivo que o
Departamento Pessoal ja imprime, e sua paleta vive em services/molde.py.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Paleta — slate escuro, a mesma do tema da interface.
COR_TITULO = "0F172A"
COR_SECAO = "1E293B"
COR_CABECALHO = "334155"
COR_ALERTA = "7F1D1D"
COR_OK = "14532D"
COR_BORDA = "CBD5E1"
COR_APAGADO = "94A3B8"
COR_NEGATIVO = "B91C1C"

FONTE_TITULO = Font(bold=True, size=14, color="FFFFFF")
FONTE_SECAO = Font(bold=True, size=11, color="FFFFFF")
FONTE_CABECALHO = Font(bold=True, color="FFFFFF")
FONTE_FORTE = Font(bold=True)
FONTE_NOTA = Font(italic=True, color=COR_APAGADO)

FILL_TITULO = PatternFill("solid", fgColor=COR_TITULO)
FILL_SECAO = PatternFill("solid", fgColor=COR_SECAO)
FILL_CABECALHO = PatternFill("solid", fgColor=COR_CABECALHO)
FILL_ALERTA = PatternFill("solid", fgColor=COR_ALERTA)
FILL_OK = PatternFill("solid", fgColor=COR_OK)

BORDA = Border(*(Side(style="thin", color=COR_BORDA),) * 4)

# Formato de moeda das celulas de valor. O Excel formata; o numero gravado e
# sempre o valor cheio, para que somas e graficos leiam o dado, nao o texto.
MOEDA = "R$ #,##0.00"


# Prefixos que o Excel interpreta como formula/comando ao abrir o arquivo.
# `=cmd|'/c calc'!A1` num nome de setor executaria na maquina de quem abre;
# `=HYPERLINK` exfiltra o conteudo da planilha para uma URL.
_PREFIXOS_FORMULA = ("=", "+", "-", "@")


def texto_seguro(valor) -> str:
    """Neutraliza texto que o Excel leria como formula, preservando o que se le.

    O apostrofo inicial e o marcador de "isto e texto" do proprio Excel: nao
    aparece na celula, mas impede a interpretacao. Vale para TODO texto que
    veio de fora — nome de secretaria digitado na tela, rotulo de setor lido
    de um .xlsx enviado, descricao de evento do relatorio de origem.

    O construtor de molde ja recusa esses prefixos no que ELE gera; isto
    cobre o que chega por outros caminhos, que ele nao tem como validar.
    Nao se aplica as formulas que o proprio app escreve — aquelas sao
    atribuidas direto na celula, de proposito.
    """
    texto = "" if valor is None else str(valor)
    if texto[:1] in _PREFIXOS_FORMULA:
        return "'" + texto
    return texto


def titulo(ws, linha: int, texto: str, largura: int) -> int:
    """Faixa de titulo mesclada. Devolve a proxima linha livre."""
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura)
    celula = ws.cell(row=linha, column=1, value=texto)
    celula.font, celula.fill = FONTE_TITULO, FILL_TITULO
    celula.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[linha].height = 26
    return linha + 1


def secao(ws, linha: int, texto: str, largura: int, alerta: bool = False) -> int:
    """Faixa de secao mesclada. Devolve a proxima linha livre."""
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura)
    celula = ws.cell(row=linha, column=1, value=texto)
    celula.font = FONTE_SECAO
    celula.fill = FILL_ALERTA if alerta else FILL_SECAO
    celula.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[linha].height = 20
    return linha + 1


def cabecalho(ws, linha: int, rotulos) -> int:
    """Linha de cabecalho de tabela. Devolve a proxima linha livre."""
    for coluna, rotulo in enumerate(rotulos, start=1):
        celula = ws.cell(row=linha, column=coluna, value=rotulo)
        celula.font, celula.fill = FONTE_CABECALHO, FILL_CABECALHO
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return linha + 1


def moeda(ws, linha: int, coluna: int, valor, negrito: bool = False):
    """Escreve um valor monetario ja formatado. Devolve a celula."""
    celula = ws.cell(row=linha, column=coluna, value=float(valor))
    celula.number_format = MOEDA
    celula.border = BORDA
    if negrito:
        celula.font = FONTE_FORTE
    return celula
