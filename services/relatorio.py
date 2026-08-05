"""Relatorio consolidado do historico: o compilado de varias operacoes.

Nao e uma lista do que foi processado — e a SOMA. Quem pede um consolidado
quer saber quanto foi retido por evento, por secretaria e por setor no
periodo inteiro, nao reler mes a mes e somar na mao.

A separacao que sustenta o modulo:

    compilar(operacoes) -> dict        # os numeros, sem formatacao
    montar_relatorio(operacoes) -> Workbook   # o mesmo compilado em .xlsx
    services/relatorio_pdf.montar_pdf(...)    # o mesmo compilado em .pdf

Os dois formatos leem a MESMA compilacao. Se o PDF e a planilha pudessem
divergir num centavo, um dos dois estaria mentindo — e ninguem saberia qual.

Tres decisoes que valem o comentario:

**Numero, nunca texto.** Toda celula de valor recebe o numero cheio com
formato de moeda. Gravar "R$ 1.234,56" como texto deixaria a planilha bonita
e inutil: nao somaria, nao ordenaria e nao viraria grafico.

**Total em formula, nao em constante.** Os totais sao `=SUM(...)` sobre o
intervalo. Quem receber o arquivo pode apagar uma linha que nao interessa e
o total se corrige sozinho — um numero fixo mentiria em silencio.

**A tabela mostra tudo; so o grafico corta.** O compilado nao trunca lista
nenhuma: se ha 300 eventos, os 300 aparecem. O corte por legibilidade
acontece apenas no desenho, onde 300 barras nao comunicam nada.

O historico nao guarda PII (so metadados e totais agregados), entao este
relatorio tambem nao tem — nenhum nome, matricula ou CPF passa por aqui.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from . import estilo_xlsx as estilo
from .utils import recortar_serie

_ZERO = Decimal("0.00")
_CENTAVO = Decimal("0.01")

ABA_COMPILADO = "Compilado"
ABA_GRAFICO = "Gráfico"

# Teto de operacoes num relatorio. O historico inteiro cabe em memoria sem
# esforco; o limite protege o documento, que acima disto vira ilegivel.
MAX_OPERACOES = 200

# Quantas fatias os GRAFICOS mostram. Nao afeta as tabelas: o compilado
# lista tudo. Acima disto as legendas se sobrepoem e o desenho para de
# comunicar, entao o excedente vira uma barra "outras" — e a soma continua
# igual a da tabela.
TOP_GRAFICO = 12

# As dimensoes do compilado, na ordem em que se lê: do maior agrupamento
# para o detalhe. 'Secretaria' sai das proprias operacoes; as demais saem
# das series que cada operacao guardou.
_DIMENSOES = (
    ("Secretaria", "POR SECRETARIA", "Secretaria"),
    ("Setor", "POR SETOR", "Setor"),
    ("Evento", "POR EVENTO DO RELATÓRIO", "Evento"),
    ("Rubrica", "POR RUBRICA (COLUNA DA PLANILHA)", "Rubrica"),
    ("Tipo", "POR LINHA DE TIPO DE FOLHA", "Linha"),
)

# Linhas do resumo geral: (rotulo, chave, e_destaque). Publico porque e
# CONTRATO entre a compilacao e os dois renderizadores — o .xlsx e o .pdf
# precisam mostrar as mesmas linhas, na mesma ordem.
LINHAS_RESUMO = (
    ("Total lido (bruto)", "lido", True),
    ("Preenchido na planilha", "preenchido", True),
    ("Fora de escopo (ignorado de propósito)", "fora", False),
    ("Sem lugar na planilha", "estrutura", False),
    ("Pendente (sem vínculo)", "pendente", False),
)


class SemOperacoes(ValueError):
    """Nenhuma operação válida para exportar."""


# ---------------------------------------------------------------------------
# Leitura tolerante do histórico
# ---------------------------------------------------------------------------
#
# O histórico é append-only e atravessa versões do app: uma operação gravada
# meses atrás não tem os campos criados depois. Ler com tolerância é o que
# permite compilar o registro antigo junto com o de ontem.

def _decimal(valor) -> Decimal:
    """Valor monetário do registro. Ausente, nulo ou ilegível -> zero."""
    if valor is None or valor == "":
        return _ZERO
    try:
        return Decimal(str(valor)).quantize(_CENTAVO)
    except (InvalidOperation, ValueError, ArithmeticError):
        return _ZERO


def _inteiro(valor) -> int:
    try:
        return int(valor or 0)
    except (TypeError, ValueError):
        return 0


def _texto(valor, padrao: str = "—") -> str:
    """Texto já neutralizado contra injeção de fórmula (ver estilo_xlsx)."""
    texto = "" if valor is None else str(valor).strip()
    return estilo.texto_seguro(texto) if texto else padrao


def _serie(op: dict, dimensao: str) -> list[dict]:
    """Série salva de uma dimensão ('Setor' | 'Rubrica' | 'Tipo' | 'Evento')."""
    graficos = op.get("graficos")
    if not isinstance(graficos, dict):
        return []
    serie = graficos.get(dimensao)
    return serie if isinstance(serie, list) else []


# ---------------------------------------------------------------------------
# Compilação — os números, sem uma linha de formatação
# ---------------------------------------------------------------------------

def _operacao(op: dict) -> dict:
    """Normaliza uma operação do histórico para o que o compilado usa."""
    return {
        "datahora": _texto(op.get("datahora")),
        "secretaria": _texto(op.get("perfil_nome") or op.get("perfil"), "Secretaria"),
        "competencia": _texto(op.get("competencia"), ""),
        "aba": _texto(op.get("aba_destino")),
        "lancamentos": _inteiro(op.get("qtd_lancamentos")),
        "celulas": _inteiro(op.get("qtd_celulas")),
        "lido": _decimal(op.get("total_lido")),
        "preenchido": _decimal(op.get("total_preenchido")),
        "fora": _decimal(op.get("total_fora_escopo")),
        "estrutura": _decimal(op.get("total_estrutura")),
        "pendente": _decimal(op.get("total_pendente")),
        "confere": bool(op.get("confere")),
        "bruto": op,
    }


def _dimensao_das_series(operacoes: list[dict], chave: str) -> dict:
    """Soma uma dimensão guardada nas séries de cada operação.

    Devolve também o que NÃO está representado: operações registradas antes
    de a dimensão existir não têm a série, e omitir isso em silêncio faria o
    compilado parecer completo quando falta dinheiro nele.
    """
    somas: dict[str, Decimal] = {}
    sem_serie, valor_sem_serie = 0, _ZERO

    for op in operacoes:
        itens = _serie(op["bruto"], chave)
        if not itens:
            sem_serie += 1
            valor_sem_serie += op["preenchido"]
            continue
        for item in itens:
            if not isinstance(item, dict):
                continue
            rotulo = _texto(item.get("label"), "")
            if not rotulo:
                continue
            somas[rotulo] = somas.get(rotulo, _ZERO) + _decimal(item.get("valor"))

    return {"somas": somas, "sem_detalhe": sem_serie, "valor_sem_detalhe": valor_sem_serie}


def _dimensao_por_secretaria(operacoes: list[dict]) -> dict:
    """Sai das próprias operações — sempre completa, em qualquer versão."""
    somas: dict[str, Decimal] = {}
    for op in operacoes:
        somas[op["secretaria"]] = somas.get(op["secretaria"], _ZERO) + op["preenchido"]
    return {"somas": somas, "sem_detalhe": 0, "valor_sem_detalhe": _ZERO}


def compilar(operacoes_brutas: list[dict]) -> dict:
    """Reduz as operações do histórico a um compilado — só números.

    É a fonte única do .xlsx e do .pdf: os dois renderizam este dicionário e
    por isso não têm como divergir.
    """
    if not operacoes_brutas:
        raise SemOperacoes("Selecione ao menos uma operação para gerar o relatório.")
    operacoes = [_operacao(op) for op in operacoes_brutas[:MAX_OPERACOES]]

    resumo = {chave: sum((op[chave] for op in operacoes), _ZERO)
              for chave in ("lido", "preenchido", "fora", "estrutura", "pendente")}
    resumo["diferenca"] = resumo["lido"] - (
        resumo["preenchido"] + resumo["fora"] + resumo["estrutura"] + resumo["pendente"]
    )
    resumo["lancamentos"] = sum(op["lancamentos"] for op in operacoes)
    resumo["celulas"] = sum(op["celulas"] for op in operacoes)
    resumo["operacoes"] = len(operacoes)

    dimensoes = []
    for chave, titulo, rotulo_coluna in _DIMENSOES:
        bruto = (_dimensao_por_secretaria(operacoes) if chave == "Secretaria"
                 else _dimensao_das_series(operacoes, chave))
        itens = sorted(bruto["somas"].items(), key=lambda kv: kv[1], reverse=True)
        if not itens and bruto["sem_detalhe"] == len(operacoes):
            continue  # dimensão que nenhuma operação selecionada registrou
        dimensoes.append({
            "chave": chave,
            "titulo": titulo,
            "rotulo_coluna": rotulo_coluna,
            "itens": itens,
            "total": sum((v for _k, v in itens), _ZERO),
            "sem_detalhe": bruto["sem_detalhe"],
            "valor_sem_detalhe": bruto["valor_sem_detalhe"],
        })

    return {
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "operacoes": operacoes,
        "competencias": sorted({op["competencia"] for op in operacoes} - {""}),
        "secretarias": sorted({op["secretaria"] for op in operacoes}),
        "resumo": resumo,
        "dimensoes": dimensoes,
    }


def para_grafico(dimensao: dict) -> list[tuple[str, Decimal]]:
    """Recorte legível de uma dimensão para desenhar.

    A regra vive em `utils.recortar_serie`, uma só para os dois formatos: o
    gráfico do .xlsx (aqui) e os do PDF cortam no mesmo ponto, senão o mesmo
    compilado sairia com desenhos diferentes em cada arquivo.
    """
    return recortar_serie(dimensao["itens"], TOP_GRAFICO)


def nome_arquivo(extensao: str = "xlsx") -> str:
    return f"RELATORIO_CONSOLIDADO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{extensao}"


# ---------------------------------------------------------------------------
# Aba 1 — Compilado
# ---------------------------------------------------------------------------

_LARGURA = 6  # A: rótulo | B..F: números


def _intervalo(coluna: int, inicio: int, fim: int) -> str:
    letra = get_column_letter(coluna)
    return f"{letra}{inicio}:{letra}{fim}"


def _cabecalho_do_documento(ws, dados: dict) -> int:
    linha = estilo.titulo(ws, 1, "RELATÓRIO CONSOLIDADO DE RETENÇÕES", _LARGURA)
    linha += 1
    for rotulo, valor in (
        ("Gerado em", dados["gerado_em"]),
        ("Operações compiladas", str(dados["resumo"]["operacoes"])),
        ("Competências", ", ".join(dados["competencias"]) or "—"),
        ("Secretarias", ", ".join(dados["secretarias"]) or "—"),
    ):
        ws.cell(row=linha, column=1, value=rotulo).font = estilo.FONTE_FORTE
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=_LARGURA)
        ws.cell(row=linha, column=2, value=estilo.texto_seguro(valor))
        linha += 1
    return linha + 1


def _secao_resumo(ws, linha: int, resumo: dict) -> int:
    linha = estilo.secao(ws, linha, "RESUMO GERAL DO PERÍODO", _LARGURA)
    for rotulo, chave, destaque in LINHAS_RESUMO:
        celula = ws.cell(row=linha, column=1, value=rotulo)
        celula.font = estilo.FONTE_FORTE if destaque else Font()
        celula.border = estilo.BORDA
        estilo.moeda(ws, linha, 2, resumo[chave], negrito=destaque)
        linha += 1

    ws.cell(row=linha, column=1, value="Diferença (deve ser zero)").font = estilo.FONTE_FORTE
    ws.cell(row=linha, column=1).border = estilo.BORDA
    celula = estilo.moeda(ws, linha, 2, resumo["diferenca"], negrito=True)
    if abs(resumo["diferenca"]) >= _CENTAVO:
        celula.font = Font(bold=True, color=estilo.COR_NEGATIVO)
    linha += 1

    ws.cell(row=linha, column=1, value="Lançamentos lidos").border = estilo.BORDA
    ws.cell(row=linha, column=2, value=resumo["lancamentos"]).border = estilo.BORDA
    linha += 1
    ws.cell(row=linha, column=1, value="Células preenchidas").border = estilo.BORDA
    ws.cell(row=linha, column=2, value=resumo["celulas"]).border = estilo.BORDA
    return linha + 2


def _secao_dimensao(ws, linha: int, dimensao: dict) -> int:
    linha = estilo.secao(ws, linha, dimensao["titulo"], _LARGURA)
    linha = estilo.cabecalho(ws, linha, [dimensao["rotulo_coluna"], "Total preenchido"])
    primeira = linha

    for rotulo, valor in dimensao["itens"]:
        ws.cell(row=linha, column=1, value=rotulo).border = estilo.BORDA
        estilo.moeda(ws, linha, 2, valor)
        linha += 1

    if dimensao["itens"]:
        ws.cell(row=linha, column=1, value="TOTAL").font = estilo.FONTE_FORTE
        ws.cell(row=linha, column=1).border = estilo.BORDA
        celula = estilo.moeda(ws, linha, 2, 0, negrito=True)
        celula.value = f"=SUM({_intervalo(2, primeira, linha - 1)})"
        linha += 1

    if dimensao["sem_detalhe"]:
        aviso = (
            f"{dimensao['sem_detalhe']} operação(ões) não têm este detalhamento "
            f"(registradas antes de a dimensão existir): "
            f"R$ {dimensao['valor_sem_detalhe']:.2f} preenchido não aparece acima."
        )
        ws.cell(row=linha, column=1, value=aviso).font = estilo.FONTE_NOTA
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=_LARGURA)
        linha += 1

    return linha + 1


# Idem: as colunas da secao de procedencia, iguais nos dois formatos.
COLUNAS_OPERACOES = ("Data / hora", "Secretaria", "Competência", "Aba",
                     "Total lido", "Preenchido")


def _secao_operacoes(ws, linha: int, operacoes: list[dict]) -> int:
    """Procedência: de onde saiu o compilado. Um relatório sem origem
    rastreável é um número que ninguém consegue conferir."""
    linha = estilo.secao(ws, linha, "OPERAÇÕES COMPILADAS", _LARGURA)
    linha = estilo.cabecalho(ws, linha, COLUNAS_OPERACOES)
    primeira = linha

    for op in operacoes:
        for coluna, valor in ((1, op["datahora"]), (2, op["secretaria"]),
                              (3, op["competencia"] or "—"), (4, op["aba"])):
            ws.cell(row=linha, column=coluna, value=valor).border = estilo.BORDA
        estilo.moeda(ws, linha, 5, op["lido"])
        estilo.moeda(ws, linha, 6, op["preenchido"])
        linha += 1

    ws.cell(row=linha, column=1, value="TOTAL").font = estilo.FONTE_FORTE
    for coluna in (5, 6):
        celula = estilo.moeda(ws, linha, coluna, 0, negrito=True)
        celula.value = f"=SUM({_intervalo(coluna, primeira, linha - 1)})"
    return linha + 2


def _aba_compilado(ws, dados: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 48
    for indice in range(2, _LARGURA + 1):
        ws.column_dimensions[get_column_letter(indice)].width = 18

    linha = _cabecalho_do_documento(ws, dados)
    linha = _secao_resumo(ws, linha, dados["resumo"])
    for dimensao in dados["dimensoes"]:
        linha = _secao_dimensao(ws, linha, dimensao)
    linha = _secao_operacoes(ws, linha, dados["operacoes"])

    ws.cell(row=linha, column=1, value=(
        "Os totais por dimensão somam o que foi destinado à planilha. "
        "Este relatório vem do histórico local, que não guarda dados pessoais: "
        "nenhum nome, matrícula ou CPF."
    )).font = estilo.FONTE_NOTA
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=_LARGURA)

    ws.page_setup.orientation = "portrait"
    ws.print_options.horizontalCentered = True


# ---------------------------------------------------------------------------
# Aba 2 — Gráfico
# ---------------------------------------------------------------------------

def _bloco_series(ws, linha: int, titulo: str, cabecalhos: list[str],
                  linhas: list[tuple]) -> tuple[int, int, int]:
    """Escreve um bloco de série. Devolve (linha_cabeçalho, primeira, última)."""
    estilo.secao(ws, linha, titulo, len(cabecalhos))
    cab = linha + 1
    estilo.cabecalho(ws, cab, cabecalhos)
    atual = cab + 1
    for rotulo, *valores in linhas:
        ws.cell(row=atual, column=1, value=rotulo).border = estilo.BORDA
        for deslocamento, valor in enumerate(valores, start=2):
            estilo.moeda(ws, atual, deslocamento, valor)
        atual += 1
    return cab, cab + 1, atual - 1


def _barras(ws, titulo: str, cab: int, primeira: int, ultima: int,
            col_ini: int, col_fim: int, ancora: str, empilhado: bool = False) -> None:
    grafico = BarChart()
    grafico.type = "col"
    grafico.grouping = "stacked" if empilhado else "clustered"
    if empilhado:
        grafico.overlap = 100
    grafico.title = titulo
    grafico.style = 10
    grafico.y_axis.numFmt = estilo.MOEDA
    grafico.height, grafico.width = 9.5, 20

    dados = Reference(ws, min_col=col_ini, max_col=col_fim, min_row=cab, max_row=ultima)
    grafico.add_data(dados, titles_from_data=True)
    grafico.set_categories(Reference(ws, min_col=1, min_row=primeira, max_row=ultima))
    ws.add_chart(grafico, ancora)


def _pizza(ws, titulo: str, cab: int, primeira: int, ultima: int, ancora: str) -> None:
    grafico = PieChart()
    grafico.title = titulo
    grafico.style = 10
    grafico.height, grafico.width = 9.5, 20
    grafico.add_data(Reference(ws, min_col=2, min_row=cab, max_row=ultima),
                     titles_from_data=True)
    grafico.set_categories(Reference(ws, min_col=1, min_row=primeira, max_row=ultima))
    grafico.dataLabels = DataLabelList()
    grafico.dataLabels.showPercent = True
    ws.add_chart(grafico, ancora)


def _aba_grafico(ws, dados: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    for letra in ("B", "C", "D"):
        ws.column_dimensions[letra].width = 16

    linha = estilo.titulo(ws, 1, "VISÃO GRÁFICA DO COMPILADO", 4)
    linha += 1

    # Composição do total lido por operação: a reconciliação virando desenho.
    linhas = [(f"{op['secretaria']} · {op['competencia']}" if op["competencia"]
               else op["secretaria"],
               op["preenchido"], op["fora"], op["estrutura"] + op["pendente"])
              for op in dados["operacoes"]]
    cab, primeira, ultima = _bloco_series(
        ws, linha, "COMPOSIÇÃO POR OPERAÇÃO",
        ["Operação", "Preenchido", "Fora de escopo", "Não preenchido"], linhas,
    )
    _barras(ws, "Composição do total lido, por operação", cab, primeira, ultima,
            2, 4, "F3", empilhado=True)
    linha = max(ultima, 22) + 3

    for dimensao in dados["dimensoes"]:
        itens = para_grafico(dimensao)
        if not itens:
            continue
        cab, primeira, ultima = _bloco_series(
            ws, linha, dimensao["titulo"],
            [dimensao["rotulo_coluna"], "Total preenchido"], list(itens),
        )
        ancora = f"F{linha + 1}"
        if dimensao["chave"] == "Tipo":
            _pizza(ws, f"Distribuição — {dimensao['rotulo_coluna'].lower()}",
                   cab, primeira, ultima, ancora)
        else:
            _barras(ws, f"Total preenchido por {dimensao['rotulo_coluna'].lower()}",
                    cab, primeira, ultima, 2, 2, ancora)
        linha = max(ultima, linha + 20) + 3

    ws.page_setup.orientation = "landscape"


# ---------------------------------------------------------------------------
# Montagem
# ---------------------------------------------------------------------------

def montar_relatorio(operacoes: list[dict]) -> Workbook:
    """Monta o .xlsx do compilado. Levanta `SemOperacoes` se não houver nada."""
    dados = compilar(operacoes)
    wb = Workbook()
    wb.remove(wb.active)
    _aba_compilado(wb.create_sheet(title=ABA_COMPILADO), dados)
    _aba_grafico(wb.create_sheet(title=ABA_GRAFICO), dados)
    # Sem isto o Excel abre a planilha com as fórmulas de TOTAL em branco.
    wb.calculation.fullCalcOnLoad = True
    return wb
