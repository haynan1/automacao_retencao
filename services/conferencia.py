"""Agregacao e conferencia com reconciliacao exata.

Invariante central (garantia "sem erros"):

    total_lido = total_preenchido
               + total_fora_escopo         (rubrica ignorada de proposito)
               + total_folha_fora_escopo   (folha ignorada de proposito)
               + total_sem_vinculo         (evento sem coluna — precisa decisao)
               + total_folha_sem_vinculo   (folha sem linha — precisa decisao)
               + total_setor_nao_mapeado   (lotacao sem setor)

Se essa soma nao bater com o total lido, algo escapou — e o sistema mostra.
Todos os totais sao calculados na aplicacao com Decimal, sem depender do
Excel recalcular.

Um lancamento precisa das TRES coordenadas para ser preenchido: setor,
coluna e linha. Faltando qualquer uma, ele cai num balde nomeado — nunca
num destino aproximado. Cada balde e um motivo diferente para o dinheiro
nao ter entrado na planilha, e a tela mostra todos separados: 'ignorei de
proposito' e 'nao soube onde por' nao podem parecer a mesma coisa.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal

from openpyxl.styles import Font

from . import estilo_xlsx as estilo

_ZERO = Decimal("0.00")


# Status de rubrica/folha que autorizam preenchimento. 'regra' e 'sugerido'
# preenchem porque preencher e o comportamento certo — mas sao deduzidos
# pelo sistema, e por isso viajam marcados ate a tela e a conferencia.
_RUBRICA_PREENCHE = {"ok", "regra"}
_FOLHA_PREENCHE = {"ok", "sugerido"}


def _preenchivel(reg: dict) -> bool:
    """True se o lancamento tem as tres coordenadas e pode ser escrito.

    Existe para que `agregar_lancamentos` e `totais_por_evento` nao possam
    discordar sobre o que conta: dois lugares aplicando a mesma regra de
    cabeca e como o mesmo dinheiro aparece em dois totais diferentes.
    """
    return (
        bool(reg.get("setor_destino"))
        and reg.get("rubrica_status") in _RUBRICA_PREENCHE
        and reg.get("folha_status") in _FOLHA_PREENCHE
        and bool(reg.get("coluna_destino"))
        and bool(reg.get("tipo_destino"))
    )


def totais_por_evento(lancamentos: list[dict]) -> dict:
    """Total lancado por EVENTO do relatorio de origem.

    O evento nao e coordenada da grade — ele escolhe a coluna, e dois
    eventos podem cair na mesma. Por isso este total sai por fora da
    agregacao: 'INSS' e 'INSS do 13º' somam separados aqui e juntos na
    coluna, e as duas visoes sao verdadeiras ao mesmo tempo.
    """
    por_evento: dict[str, Decimal] = {}
    for reg in lancamentos:
        if not _preenchivel(reg):
            continue
        rotulo = (reg.get("evento") or reg.get("descricao_original") or "").strip()
        if not rotulo:
            continue
        por_evento[rotulo] = por_evento.get(rotulo, _ZERO) + reg["valor"]
    return dict(sorted(por_evento.items()))


def agregar_lancamentos(lancamentos: list[dict]) -> tuple[dict, dict]:
    """Agrupa lancamentos preenchiveis por (setor, tipo, coluna) somando valor.

    So entram lancamentos com as tres coordenadas resolvidas. Retorna
    (agregados, baldes), em que os baldes classificam — por MOTIVO — tudo
    o que nao foi agregado.

    A ordem dos testes e deliberada: primeiro o que falta (setor), depois o
    que foi ignorado de proposito (fora de escopo), so entao o que ficou sem
    decisao. Um lancamento com dois problemas aparece uma vez so, sempre no
    balde de cima — a soma nunca conta o mesmo centavo duas vezes.
    """
    agregados: dict[tuple, Decimal] = {}
    baldes = {
        "preenchivel": _ZERO,
        "fora_escopo": _ZERO,
        "folha_fora_escopo": _ZERO,
        "sem_vinculo": _ZERO,
        "folha_sem_vinculo": _ZERO,
        "setor_nao_mapeado": _ZERO,
    }

    for reg in lancamentos:
        valor = reg["valor"]
        setor = reg.get("setor_destino")
        tipo = reg.get("tipo_destino")
        coluna = reg.get("coluna_destino")
        status_rubrica = reg.get("rubrica_status")
        status_folha = reg.get("folha_status")

        if not setor:
            baldes["setor_nao_mapeado"] += valor
            continue
        if status_rubrica == "fora_escopo":
            baldes["fora_escopo"] += valor
            continue
        if status_folha == "fora_escopo":
            baldes["folha_fora_escopo"] += valor
            continue
        if status_rubrica not in _RUBRICA_PREENCHE or not coluna:
            baldes["sem_vinculo"] += valor
            continue
        if status_folha not in _FOLHA_PREENCHE or not tipo:
            baldes["folha_sem_vinculo"] += valor
            continue

        chave = (setor, tipo, coluna)
        agregados[chave] = agregados.get(chave, _ZERO) + valor
        baldes["preenchivel"] += valor

    return agregados, baldes


def calcular_totais_lidos(lancamentos: list[dict]) -> Decimal:
    total = _ZERO
    for reg in lancamentos:
        total += reg["valor"]
    return total


def calcular_totais_agregados(agregados: dict) -> dict:
    """Totais por setor, por coluna (rubrica) e por tipo."""
    por_setor: dict[str, Decimal] = {}
    por_coluna: dict[str, Decimal] = {}
    por_tipo: dict[str, Decimal] = {}
    total = _ZERO
    for (setor, tipo, coluna), valor in agregados.items():
        por_setor[setor] = por_setor.get(setor, _ZERO) + valor
        por_coluna[coluna] = por_coluna.get(coluna, _ZERO) + valor
        por_tipo[tipo] = por_tipo.get(tipo, _ZERO) + valor
        total += valor
    return {
        "por_setor": dict(sorted(por_setor.items())),
        "por_coluna": dict(sorted(por_coluna.items())),
        "por_tipo": dict(sorted(por_tipo.items())),
        "total": total,
    }


def reconciliar(total_lido: Decimal, baldes: dict, pendencias_estrutura: list[dict]) -> dict:
    """Verifica a invariante de reconciliacao. Retorna diagnostico completo."""
    total_estrutura = sum((i.get("valor", _ZERO) for i in pendencias_estrutura), _ZERO)
    # Soma TODOS os baldes, sem listar um por um: eles sao mutuamente
    # exclusivos por construcao (`agregar_lancamentos` classifica cada
    # lancamento uma unica vez), e assim um balde novo entra na invariante
    # sem que ninguem precise lembrar de soma-lo aqui.
    soma_baldes = sum(baldes.values(), _ZERO)
    # Itens que estavam preenchiveis mas nao acharam lugar na planilha migram
    # de 'preenchido' para 'estrutura'.
    total_preenchido = baldes["preenchivel"] - total_estrutura
    diferenca = total_lido - soma_baldes
    return {
        "total_lido": total_lido,
        "total_preenchido": total_preenchido,
        "total_fora_escopo": baldes["fora_escopo"],
        "total_folha_fora_escopo": baldes["folha_fora_escopo"],
        "total_sem_vinculo": baldes["sem_vinculo"],
        "total_folha_sem_vinculo": baldes["folha_sem_vinculo"],
        "total_setor_nao_mapeado": baldes["setor_nao_mapeado"],
        "total_estrutura": total_estrutura,
        "diferenca": diferenca,
        "confere": abs(diferenca) < Decimal("0.01"),
    }


# ---------------------------------------------------------------------------
# Aba de conferencia
# ---------------------------------------------------------------------------

# O vocabulario visual e compartilhado com o relatorio consolidado do
# historico: os dois documentos saem da mesma mao (ver services/estilo_xlsx).
_FONTE_CAB = estilo.FONTE_CABECALHO
_FILL_CAB = estilo.FILL_CABECALHO
_FILL_ALERTA = estilo.FILL_ALERTA
_FILL_OK = estilo.FILL_OK
_BORDA = estilo.BORDA
_MOEDA = estilo.MOEDA
_LARGURA = 4  # colunas da aba: item | destino | valor | motivo


def criar_aba_conferencia(wb, dados: dict) -> None:
    nome = "CONFERÊNCIA_AUTOMAÇÃO"
    if nome in wb.sheetnames:
        del wb[nome]
    ws = wb.create_sheet(title=nome)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 52  # o motivo de cada decisão, por extenso

    linha = _titulo(ws, 1, "CONFERÊNCIA DA AUTOMAÇÃO DE RETENÇÕES")
    linha += 1

    for rotulo, valor in (
        ("Processado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Competência", dados.get("competencia") or "—"),
        ("Arquivo de origem", dados.get("arquivo_origem") or "—"),
        ("Planilha modelo", dados.get("arquivo_modelo") or "—"),
        ("Aba preenchida", dados.get("aba_destino") or "—"),
    ):
        ws.cell(row=linha, column=1, value=rotulo).font = Font(bold=True)
        ws.cell(row=linha, column=2, value=estilo.texto_seguro(valor))
        linha += 1
    linha += 1

    # Reconciliacao ------------------------------------------------------
    rec = dados["reconciliacao"]
    linha = _secao(ws, linha, "RECONCILIAÇÃO", alerta=not rec["confere"])
    status = "✔ CONFERE (bate ao centavo)" if rec["confere"] else f"✘ DIVERGÊNCIA de {rec['diferenca']}"
    c = ws.cell(row=linha, column=1, value=status)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = _FILL_OK if rec["confere"] else _FILL_ALERTA
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    linha += 1
    for rotulo, chave in (
        ("Total lido (bruto)", "total_lido"),
        ("Total preenchido na planilha", "total_preenchido"),
        ("Fora de escopo — rubrica (ignorado)", "total_fora_escopo"),
        ("Fora de escopo — folha (ignorado)", "total_folha_fora_escopo"),
        ("Sem vínculo (evento sem coluna)", "total_sem_vinculo"),
        ("Sem vínculo (folha sem linha)", "total_folha_sem_vinculo"),
        ("Setor não mapeado", "total_setor_nao_mapeado"),
    ):
        ws.cell(row=linha, column=1, value=rotulo).font = Font(bold=(chave in ("total_lido", "total_preenchido")))
        cc = ws.cell(row=linha, column=2, value=float(rec[chave]))
        cc.number_format = _MOEDA
        linha += 1
    linha += 1

    linha = _tabela_valor(ws, linha, "TOTAL POR SETOR", dados.get("por_setor", {}))
    linha = _tabela_valor(ws, linha, "TOTAL POR RUBRICA (COLUNA)", dados.get("por_coluna", {}))
    linha = _tabela_valor(ws, linha, "TOTAL POR TIPO DE FOLHA (LINHA)", dados.get("por_tipo", {}))

    # Rastro de auditoria: cada vinculo com o motivo por extenso. Quem abrir
    # a planilha daqui a seis meses consegue reconstruir POR QUE cada valor
    # foi parar naquela celula, sem ter o app na frente.
    linha = _secao(ws, linha, "COMO O SISTEMA DECIDIU")
    linha = _tabela_decisao(
        ws, linha, "Folha → linha da planilha", dados.get("decisoes_folhas", []),
        "Folha (relatório)", "Linha", "tipo",
    )
    linha = _tabela_decisao(
        ws, linha, "Evento → coluna da planilha", dados.get("decisoes_rubricas", []),
        "Evento (relatório)", "Coluna", "coluna",
    )

    linha = _secao(ws, linha, "PENDÊNCIAS E OBSERVAÇÕES", alerta=True)
    linha = _tabela_contagem(ws, linha, "Lotações não mapeadas", dados.get("lotacoes_nao_mapeadas", {}))
    linha = _tabela_contagem(ws, linha, "Eventos sem vínculo (precisam de decisão)", dados.get("rubricas_sem_vinculo", {}))
    linha = _tabela_contagem(ws, linha, "Folhas sem linha de destino (precisam de decisão)", dados.get("folhas_sem_vinculo", {}))
    linha = _tabela_contagem(ws, linha, "Folhas somadas em outra linha (sugestão do sistema — conferir)", dados.get("folhas_sugeridas", {}))
    linha = _tabela_contagem(ws, linha, "Eventos fora de escopo (ignorados)", dados.get("rubricas_fora_escopo", {}))
    linha = _tabela_contagem(ws, linha, "Folhas fora de escopo (ignoradas)", dados.get("folhas_fora_escopo", {}))

    estrutura = dados.get("pendencias_estrutura", [])
    if estrutura:
        ws.cell(row=linha, column=1, value="Itens sem lugar na planilha").font = Font(bold=True, italic=True)
        linha += 1
        for cab, col in (("Motivo", 1), ("Setor / Rubrica", 2), ("Valor", 3)):
            cel = ws.cell(row=linha, column=col, value=cab)
            cel.font, cel.fill = _FONTE_CAB, _FILL_CAB
        linha += 1
        for item in estrutura:
            ws.cell(row=linha, column=1, value=estilo.texto_seguro(item.get("motivo")))
            ws.cell(row=linha, column=2,
                    value=estilo.texto_seguro(f"{item.get('setor')} / {item.get('rubrica')}"))
            cc = ws.cell(row=linha, column=3, value=float(item.get("valor", _ZERO)))
            cc.number_format = _MOEDA
            linha += 1


def _titulo(ws, linha, texto):
    return estilo.titulo(ws, linha, texto, _LARGURA)


def _secao(ws, linha, texto, alerta=False):
    return estilo.secao(ws, linha, texto, _LARGURA, alerta)


def _tabela_valor(ws, linha, titulo, dados: dict):
    linha = _secao(ws, linha, titulo)
    if not dados:
        ws.cell(row=linha, column=1, value="— nada a exibir —").font = Font(italic=True, color="94A3B8")
        return linha + 2
    for cab, col in (("Descrição", 1), ("Valor", 2)):
        cel = ws.cell(row=linha, column=col, value=cab)
        cel.font, cel.fill = _FONTE_CAB, _FILL_CAB
    linha += 1
    total = _ZERO
    for nome, valor in dados.items():
        ws.cell(row=linha, column=1, value=estilo.texto_seguro(nome)).border = _BORDA
        cc = ws.cell(row=linha, column=2, value=float(valor))
        cc.number_format, cc.border = _MOEDA, _BORDA
        total += valor
        linha += 1
    ws.cell(row=linha, column=1, value="TOTAL").font = Font(bold=True)
    cc = ws.cell(row=linha, column=2, value=float(total))
    cc.number_format, cc.font = _MOEDA, Font(bold=True)
    return linha + 2


_ROTULO_STATUS = {
    "ok": "definido",
    "regra": "deduzido por regra",
    "sugerido": "sugerido pelo sistema",
    "fora_escopo": "ignorado de propósito",
    "sem_vinculo": "SEM DESTINO",
}


def _tabela_decisao(ws, linha, titulo, grupos: list, cab_item: str, cab_destino: str, campo: str):
    """Tabela item -> destino -> motivo, com o valor de cada grupo.

    `campo` diz onde ler o destino no grupo ('tipo' para folhas, 'coluna'
    para eventos): as duas listas tem o mesmo formato em tudo o mais.
    """
    ws.cell(row=linha, column=1, value=titulo).font = Font(bold=True, italic=True)
    linha += 1
    if not grupos:
        ws.cell(row=linha, column=1, value="— nada a exibir —").font = Font(italic=True, color="94A3B8")
        return linha + 2

    for cab, col in ((cab_item, 1), (cab_destino, 2), ("Valor", 3), ("Por quê", 4)):
        cel = ws.cell(row=linha, column=col, value=cab)
        cel.font, cel.fill = _FONTE_CAB, _FILL_CAB
    linha += 1

    for grupo in grupos:
        ws.cell(row=linha, column=1, value=estilo.texto_seguro(grupo.get("rotulo"))).border = _BORDA
        destino = grupo.get(campo) or "— não preenchido —"
        ws.cell(row=linha, column=2, value=estilo.texto_seguro(destino)).border = _BORDA
        cc = ws.cell(row=linha, column=3, value=float(grupo.get("total", _ZERO)))
        cc.number_format, cc.border = _MOEDA, _BORDA
        status = _ROTULO_STATUS.get(grupo.get("status"), grupo.get("status") or "")
        motivo = grupo.get("motivo") or ""
        cel = ws.cell(row=linha, column=4,
                      value=estilo.texto_seguro(f"{status} — {motivo}" if motivo else status))
        cel.border = _BORDA
        if grupo.get("status") == "sem_vinculo":
            cel.font = Font(bold=True, color="B91C1C")
        linha += 1
    return linha + 1


def _tabela_contagem(ws, linha, titulo, dados: dict):
    ws.cell(row=linha, column=1, value=titulo).font = Font(bold=True, italic=True)
    linha += 1
    if not dados:
        ws.cell(row=linha, column=1, value="Nenhuma").font = Font(italic=True, color="16A34A")
        return linha + 2
    for cab, col in (("Item", 1), ("Ocorrências", 2)):
        cel = ws.cell(row=linha, column=col, value=cab)
        cel.font, cel.fill = _FONTE_CAB, _FILL_CAB
    linha += 1
    for nome, qtd in dados.items():
        ws.cell(row=linha, column=1, value=estilo.texto_seguro(nome)).border = _BORDA
        ws.cell(row=linha, column=2, value=qtd).border = _BORDA
        linha += 1
    return linha + 2
