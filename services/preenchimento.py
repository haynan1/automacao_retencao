"""Preenchimento da planilha modelo de Retencao.

Nada aqui e feito por coordenada fixa. Os blocos de setor, as colunas de
rubrica e as linhas de tipo sao localizados dinamicamente pela estrutura da
planilha, preservando formulas, estilos, mesclagens e dashboard.

As linhas de tipo sao ABERTAS: o que a planilha desenhar — Mensal, 13o
salario, Ferias, Rescisao, Complementar, o que for — vira um destino
possivel. O motor nao carrega uma lista fechada de tipos; ele le os rotulos
que existem. Quem decide qual folha cai em qual linha e o vinculo do eixo 3
(services/mapeador.py), na tela, e nao uma regra escondida aqui.

Como um bloco e delimitado, ja que o rotulo de uma linha de tipo pode ser
qualquer texto:

    NOME DO SETOR      <- abre o bloco (a linha seguinte comeca com 'Tipo')
    Tipo | ARSEM | ...  <- cabecalho
    <qualquer rotulo>   <- linha de tipo (destino de lancamento)
    ...
    TOTAL               <- fecha o bloco (nunca e tocada — mantem formula)

A varredura para no rotulo iniciado/contendo TOTAL, na proxima linha que
abre um bloco, ou num novo cabecalho 'Tipo'. E estrutura, nao vocabulario.
"""

from __future__ import annotations

from decimal import Decimal

from openpyxl.worksheet.worksheet import Worksheet

from .normalizador import normalizar_texto

# Rotulo da primeira coluna do cabecalho do bloco, ja normalizado.
ROTULO_TIPO_NORM = "TIPO"

# Teto de linhas varridas dentro de um bloco. Nao e regra de negocio: e o
# freio que impede um molde malformado (sem TOTAL e sem bloco seguinte) de
# arrastar a varredura pela planilha inteira. Fica com folha sobre o teto de
# tipos do construtor (`molde.MAX_TIPOS`) para que todo molde desenhado na
# tela caiba aqui — um teste guarda essa relacao.
MAX_LINHAS_TIPO = 48

# Meses por numero para casar a competencia com o nome da aba (FMS JUNHO).
_MESES_PT = {
    "01": "JANEIRO", "02": "FEVEREIRO", "03": "MARÇO", "04": "ABRIL",
    "05": "MAIO", "06": "JUNHO", "07": "JULHO", "08": "AGOSTO",
    "09": "SETEMBRO", "10": "OUTUBRO", "11": "NOVEMBRO", "12": "DEZEMBRO",
}


def localizar_aba_destino(wb, competencia: str | None) -> str | None:
    """Sugere a aba de destino a partir da competencia (MM/AAAA).

    06/2026 -> procura uma aba cujo nome contenha 'JUNHO'. Retorna o nome
    exato da aba encontrada, ou None se nao houver correspondencia.
    """
    if not competencia or "/" not in competencia:
        return None
    mes = competencia.split("/")[0]
    nome_mes = _MESES_PT.get(mes)
    if not nome_mes:
        return None

    alvo = normalizar_texto(nome_mes)
    for nome in wb.sheetnames:
        if alvo in normalizar_texto(nome):
            return nome
    return None


def chave_tipo(texto: str) -> str:
    """Chave canonica de uma linha de tipo: o proprio rotulo normalizado.

    Publico de proposito. Toda comparacao entre 'o tipo que o vinculo
    escolheu' e 'a linha que existe na planilha' passa por aqui — assim
    “13º salário”, “13o SALARIO” e “13º  Salário” sao a mesma linha, e
    nenhum lado precisa conhecer uma lista fechada de tipos.
    """
    return normalizar_texto(texto)


def localizar_colunas_rubricas(ws: Worksheet, linha_cabecalho: int) -> dict:
    """Mapeia rubrica_normalizada -> {'coluna': idx, 'rotulo': texto}.

    Lê a linha de cabecalho do bloco ('Tipo | ARSEM | CEF | ...') e indexa
    cada coluna pela rubrica normalizada, ignorando a coluna 'Tipo'.
    """
    colunas: dict[str, dict] = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=linha_cabecalho, column=col).value
        rotulo = "" if valor is None else str(valor).strip()
        norm = normalizar_texto(rotulo)
        # Ignora a coluna 'Tipo' e colunas de total (ex.: 'TOTAL DO EVENTO'),
        # que nao sao rubricas e nunca devem receber lancamento.
        if not norm or norm == "TIPO" or norm.startswith("TOTAL"):
            continue
        # Se a mesma rubrica aparecer duas vezes, mantem a primeira ocorrencia.
        colunas.setdefault(norm, {"coluna": col, "rotulo": rotulo})
    return colunas


def colunas_dos_blocos(blocos: list[dict]) -> list[str]:
    """Rotulos de rubrica, na ordem das colunas, a partir de blocos ja lidos.

    As colunas sao identicas em todos os blocos; usamos o bloco com mais
    colunas como referencia. Esta lista e a FONTE DA VERDADE para o vinculo
    evento -> coluna (nunca se escreve numa coluna que nao exista).
    """
    if not blocos:
        return []
    ref = max(blocos, key=lambda b: len(b["colunas"]))
    itens = sorted(ref["colunas"].values(), key=lambda x: x["coluna"])
    return [i["rotulo"] for i in itens]


def tipos_dos_blocos(blocos: list[dict]) -> list[str]:
    """Rotulos de tipo, na ordem das linhas, a partir de blocos ja lidos.

    Contraparte de `colunas_dos_blocos` para o outro eixo da grade. E a
    FONTE DA VERDADE do vinculo folha -> linha: nunca se escreve numa linha
    que nao exista. Usa o bloco com mais linhas como referencia, pelo mesmo
    motivo das colunas (um bloco incompleto nao empobrece o menu da tela).
    """
    if not blocos:
        return []
    ref = max(blocos, key=lambda b: len(b["tipos"]))
    return list(ref["tipos"])


# As duas formas convenientes, para quem tem a aba e nao os blocos. Quem le
# os dois eixos da MESMA aba deve varrer uma vez e usar as funcoes acima:
# cada chamada destas refaz a varredura inteira do bloco.

def listar_colunas_modelo(ws: Worksheet) -> list[str]:
    return colunas_dos_blocos(localizar_blocos_setores(ws))


def listar_tipos_modelo(ws: Worksheet) -> list[str]:
    return tipos_dos_blocos(localizar_blocos_setores(ws))


def _rotulo_da_linha(ws: Worksheet, linha: int) -> str:
    valor = ws.cell(row=linha, column=1).value
    return "" if valor is None else str(valor).strip()


def _abre_bloco(ws: Worksheet, linha: int) -> bool:
    """True se `linha` e o nome de um setor — a linha seguinte e o 'Tipo'.

    Este e o unico sinal que separa 'nome do proximo setor' de 'mais uma
    linha de tipo deste bloco'. Enquanto os tipos eram dois nomes fixos,
    dava para adivinhar pelo texto; com tipos abertos, so a estrutura
    responde — e ela responde sempre, inclusive em molde sem linha TOTAL.
    """
    seguinte = ws.cell(row=linha + 1, column=1).value
    return seguinte is not None and normalizar_texto(str(seguinte)) == ROTULO_TIPO_NORM


def localizar_blocos_setores(ws: Worksheet) -> list[dict]:
    """Localiza dinamicamente todos os blocos de setor da aba.

    Um bloco e identificado quando a coluna A tem um nome (setor) e a
    linha imediatamente abaixo comeca com 'Tipo'. A partir dai, as linhas
    de tipo e a linha TOTAL sao localizadas ate o fim do bloco.

    Cada bloco devolve:
        setor, setor_norm, linha_setor, linha_cabecalho,
        colunas     -> {rubrica_norm: {coluna, rotulo}}
        linhas_tipo -> {tipo_norm: linha}      (chave via `chave_tipo`)
        tipos       -> [rotulo, ...] na ordem da planilha
        linha_total -> int | None
    """
    blocos: list[dict] = []
    max_row = ws.max_row

    linha = 1
    while linha <= max_row:
        if _rotulo_da_linha(ws, linha) and _abre_bloco(ws, linha):
            setor = _rotulo_da_linha(ws, linha)
            linha_cabecalho = linha + 1
            colunas = localizar_colunas_rubricas(ws, linha_cabecalho)
            linhas_tipo, tipos, linha_total, proxima = _localizar_linhas_tipo(
                ws, linha_cabecalho, max_row
            )
            blocos.append(
                {
                    "setor": setor,
                    "setor_norm": normalizar_texto(setor),
                    "linha_setor": linha,
                    "linha_cabecalho": linha_cabecalho,
                    "colunas": colunas,
                    "linhas_tipo": linhas_tipo,
                    "tipos": tipos,
                    "linha_total": linha_total,
                }
            )
            # `proxima` ja e a primeira linha FORA do bloco: quando a
            # varredura parou no nome do proximo setor, retomar em
            # 'proxima + 1' engoliria aquele bloco inteiro.
            linha = max(proxima, linha_cabecalho + 1)
            continue
        linha += 1

    return blocos


def _localizar_linhas_tipo(ws, linha_cabecalho: int, max_row: int):
    """A partir do cabecalho, le as linhas de tipo e localiza o TOTAL.

    Retorna (linhas_tipo, tipos, linha_total, proxima_linha), em que
    `proxima_linha` e a primeira linha que ja NAO pertence a este bloco — e
    onde a varredura de blocos deve retomar.

    Qualquer rotulo e um tipo valido. A varredura so termina por estrutura:

      * TOTAL   — pertence ao bloco e o fecha (a linha seguinte ja e de fora);
      * um nome de setor (a linha abaixo dele e o cabecalho 'Tipo');
      * um novo cabecalho 'Tipo' — que rotula outra secao (o TOTAL GERAL),
        e portanto nunca e uma linha de valor;
      * o teto de `MAX_LINHAS_TIPO`, freio contra molde malformado.

    Rotulo repetido mantem a PRIMEIRA ocorrencia, pela mesma razao das
    colunas: e onde o preenchimento vai escrever, e escrever em dois
    lugares seria pior que escrever em um.
    """
    linhas_tipo: dict[str, int] = {}
    tipos: list[str] = []
    linha = linha_cabecalho + 1
    limite = min(max_row, linha_cabecalho + MAX_LINHAS_TIPO + 1)

    while linha <= limite:
        texto = _rotulo_da_linha(ws, linha)
        norm = normalizar_texto(texto)

        if not texto:
            linha += 1
            continue
        if "TOTAL" in norm:
            # TOTAL pertence ao bloco: o proximo comeca na linha seguinte.
            return linhas_tipo, tipos, linha, linha + 1
        if norm == ROTULO_TIPO_NORM or _abre_bloco(ws, linha):
            # Ja e outra secao. Devolve a propria linha para reavaliacao.
            return linhas_tipo, tipos, None, linha

        chave = chave_tipo(texto)
        if chave not in linhas_tipo:
            linhas_tipo[chave] = linha
            tipos.append(texto)
        linha += 1

    return linhas_tipo, tipos, None, linha


def limpar_area_lancamento(ws: Worksheet, blocos: list[dict]) -> None:
    """Zera as celulas de valor (Mensal / 13º) antes de preencher, para
    garantir idempotencia. Nao toca em TOTAL nem em nenhuma formula."""
    for bloco in blocos:
        for tipo, linha in bloco["linhas_tipo"].items():
            for info in bloco["colunas"].values():
                celula = ws.cell(row=linha, column=info["coluna"])
                # So limpa celulas que NAO contem formula.
                if not (isinstance(celula.value, str) and celula.value.startswith("=")):
                    celula.value = None


def preencher_valores(ws: Worksheet, agregados: dict, blocos: list[dict]) -> dict:
    """Escreve os valores agregados nas celulas corretas.

    'agregados' e um dict {(setor, tipo, rubrica): Decimal}.

    Retorna um relatorio de preenchimento:
        {
          "preenchidos": [ {setor, tipo, rubrica, valor, celula} ],
          "total_preenchido": Decimal,
          "pendencias_estrutura": [ {motivo, setor, tipo, rubrica, valor} ],
        }
    """
    indice_setor = {b["setor_norm"]: b for b in blocos}

    preenchidos: list[dict] = []
    pendencias: list[dict] = []
    total = Decimal("0.00")

    for (setor, tipo, rubrica), valor in agregados.items():
        bloco = indice_setor.get(normalizar_texto(setor))
        if bloco is None:
            pendencias.append(_pend("setor não encontrado na planilha", setor, tipo, rubrica, valor))
            continue

        linha = bloco["linhas_tipo"].get(chave_tipo(tipo))
        if linha is None:
            pendencias.append(_pend(f"linha '{tipo}' não encontrada no setor", setor, tipo, rubrica, valor))
            continue

        info_col = bloco["colunas"].get(normalizar_texto(rubrica))
        if info_col is None:
            pendencias.append(_pend("rubrica sem coluna na planilha", setor, tipo, rubrica, valor))
            continue

        celula = ws.cell(row=linha, column=info_col["coluna"])
        # Seguranca extra: jamais sobrescrever uma formula.
        if isinstance(celula.value, str) and celula.value.startswith("="):
            pendencias.append(_pend("célula contém fórmula (protegida)", setor, tipo, rubrica, valor))
            continue

        celula.value = float(valor)
        total += valor
        preenchidos.append(
            {
                "setor": setor,
                "tipo": tipo,
                "rubrica": rubrica,
                "valor": valor,
                "celula": f"{celula.coordinate}",
            }
        )

    return {
        "preenchidos": preenchidos,
        "total_preenchido": total,
        "pendencias_estrutura": pendencias,
    }


def _pend(motivo, setor, tipo, rubrica, valor) -> dict:
    return {"motivo": motivo, "setor": setor, "tipo": tipo, "rubrica": rubrica, "valor": valor}


def preservar_formulas(wb) -> None:
    """Força o Excel a recalcular todas as formulas ao abrir o arquivo.

    openpyxl nao recalcula formulas; garantimos que o Excel o faca na
    abertura, mantendo dashboard e totais coerentes.
    """
    wb.calculation.fullCalcOnLoad = True
    try:
        wb.calculation.forceFullCalc = True
    except Exception:  # atributo pode variar entre versoes do openpyxl
        pass
    wb.calculation.calcMode = "auto"
