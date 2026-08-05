# -*- coding: utf-8 -*-
"""Fixtures sintéticos que reproduzem as armadilhas do arquivo real:
- cabeçalhos/banners mesclados na largura toda;
- coluna Descrição deslocada em relação ao cabeçalho (mesclagem);
- seção de resumo com números soltos.
"""
import pytest
from openpyxl import Workbook


def _merge(ws, row, c1, c2, value):
    ws.cell(row=row, column=c1, value=value)
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)


def _linha_dados(ws, row, mat, nome, cpf, folha, descricao, base, refer, valor):
    _merge(ws, row, 1, 4, mat)        # Mat.
    _merge(ws, row, 5, 10, nome)      # Funcionário
    _merge(ws, row, 11, 13, cpf)      # CPF
    _merge(ws, row, 14, 16, folha)    # Folha
    _merge(ws, row, 18, 23, descricao)  # Descrição — DESLOCADA (R:W) de propósito
    ws.cell(row=row, column=24, value=base)   # Base de calc.
    ws.cell(row=row, column=25, value=refer)  # Refer.
    _merge(ws, row, 26, 29, valor)    # Valor (Z:AC)


@pytest.fixture
def listagem(tmp_path):
    """Cria uma ListagemEventos.xlsx sintética fiel à estrutura real."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Page1"
    # Banners (largura total) — não podem virar lançamento.
    _merge(ws, 2, 1, 29, "MUNICÍPIO DE SÃO LUÍS DE MONTES BELOS")
    _merge(ws, 3, 1, 29, "FMS FUNDO MUNICIPAL DE SAUDE SLMB - CNPJ: 10.581.764/0001-71")
    _merge(ws, 4, 1, 29, "LISTAGEM DE EVENTOS")
    _merge(ws, 5, 1, 29, "06/2026")
    _merge(ws, 6, 1, 29, "5.0043.0000 - FMS, AGENTE COMUNITÁRIO DE SAÚDE - ACS")
    # Cabeçalho — Descrição mesclada em Q:V (17:22).
    _merge(ws, 7, 1, 4, "Mat.")
    _merge(ws, 7, 5, 10, "Funcionário")
    _merge(ws, 7, 11, 13, "CPF")
    _merge(ws, 7, 14, 16, "Folha")
    _merge(ws, 7, 17, 22, "Descrição")
    ws.cell(row=7, column=24, value="Base de calc.")
    ws.cell(row=7, column=25, value="Refer.")
    _merge(ws, 7, 26, 29, "Valor")
    # Lançamentos ACS
    _linha_dados(ws, 8, "1809", "ADEMAR F", "424.270.101-20", "MENSAL", "PREVIBELOS", "4988.55", "14.0", "698.39")
    _linha_dados(ws, 9, "1809", "ADEMAR F", "424.270.101-20", "MENSAL", "12/120 - EMPRESTIMO CEF 1", "0.0", "0.0", "88.0")
    _linha_dados(ws, 10, "1809", "ADEMAR F", "424.270.101-20", "MENSAL", "IRRF", "7506.82", "27.5", "963.58")
    _linha_dados(ws, 11, "1732", "ANA L", "788.676.571-00", "FÉRIAS", "PREVIBELOS", "4957.58", "14.0", "694.06")
    _linha_dados(ws, 12, "1732", "ANA L", "788.676.571-00", "MENSAL", "INSS", "0.0", "0.0", "120.00")  # fora de escopo
    # Nova lotação (não mapeada)
    _merge(ws, 13, 1, 29, "5.9999.0000 - FMS, LOTAÇÃO NOVA SEM SETOR")
    _linha_dados(ws, 14, "2000", "BIA N", "111.111.111-11", "MENSAL", "PREVIBELOS", "1000.0", "14.0", "100.00")
    # Seção de resumo — números soltos não podem ser somados como lançamento.
    _merge(ws, 16, 1, 29, "Resumo")
    ws.cell(row=17, column=18, value="99999.99")
    caminho = tmp_path / "ListagemEventos.xlsx"
    wb.save(caminho)
    return str(caminho)


@pytest.fixture
def modelo(tmp_path):
    """Cria um RETENÇÃO.xlsx sintético com blocos de setor e coluna TOTAL DO EVENTO."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MOLDE"
    ws.cell(row=1, column=1, value="MODELO DE RETENÇÕES")
    colunas = ["ARSEM", "CEF", "IR", "PREVIBELOS", "IPASGO", "TOTAL DO EVENTO"]

    def bloco(start, setor):
        ws.cell(row=start, column=1, value=setor)
        ws.cell(row=start + 1, column=1, value="Tipo")
        for i, c in enumerate(colunas, start=2):
            ws.cell(row=start + 1, column=i, value=c)
        ws.cell(row=start + 2, column=1, value="Mensal")
        ws.cell(row=start + 3, column=1, value="13º salário")
        ws.cell(row=start + 4, column=1, value="TOTAL")
        for i in range(2, 2 + len(colunas)):
            col = ws.cell(row=start + 2, column=i).column_letter
            ws.cell(row=start + 4, column=i, value=f"=SUM({col}{start+2}:{col}{start+3})")
        return start + 6

    n = bloco(3, "ADMINISTRATIVO")
    bloco(n, "ACS")
    caminho = tmp_path / "RETENCAO.xlsx"
    wb.save(caminho)
    return str(caminho)


# ---------------------------------------------------------------------------
# Fábricas — para cenários que precisam de outra grade (linhas de tipo
# próprias) ou de outros eventos (INSS x INSS do 13º).
# ---------------------------------------------------------------------------

@pytest.fixture
def perfis_tmp(tmp_path, monkeypatch):
    """Redireciona os dados de perfil para uma pasta descartável.

    Sem isto, um teste que aprende um vínculo escreveria no config/ real —
    e o mês seguinte começaria com uma decisão que ninguém tomou.
    """
    from services import perfis
    monkeypatch.setattr(perfis, "PERFIS_CONFIG_DIR", tmp_path / "config" / "perfis")
    monkeypatch.setattr(perfis, "PERFIS_MODELOS_DIR", tmp_path / "modelos" / "perfis")
    monkeypatch.setattr(perfis, "REGISTRO", tmp_path / "config" / "perfis.json")
    (tmp_path / "config").mkdir(parents=True, exist_ok=True)
    return tmp_path


@pytest.fixture
def historico_tmp(tmp_path, monkeypatch):
    """Histórico numa pasta descartável — um teste não pode sujar o registro real."""
    from services import historico
    monkeypatch.setattr(historico, "_ARQ", tmp_path / "historico" / "operacoes.jsonl")
    (tmp_path / "historico").mkdir(parents=True, exist_ok=True)
    return historico


@pytest.fixture
def fazer_listagem(tmp_path):
    """Monta uma Listagem sintética a partir de (lotação, folha, evento, valor).

    Mantém as armadilhas do arquivo real — banners mesclados na largura toda
    e a coluna Descrição deslocada — porque é contra elas que o parser foi
    escrito, e um fixture "limpo" testaria um arquivo que não existe.
    """
    def _fazer(itens, nome="ListagemEventos.xlsx", competencia="06/2026"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Page1"
        _merge(ws, 1, 1, 29, "MUNICÍPIO DE SÃO LUÍS DE MONTES BELOS")
        _merge(ws, 2, 1, 29, "LISTAGEM DE EVENTOS")
        _merge(ws, 3, 1, 29, competencia)
        _merge(ws, 4, 1, 4, "Mat.")
        _merge(ws, 4, 5, 10, "Funcionário")
        _merge(ws, 4, 11, 13, "CPF")
        _merge(ws, 4, 14, 16, "Folha")
        _merge(ws, 4, 17, 22, "Descrição")
        ws.cell(row=4, column=24, value="Base de calc.")
        ws.cell(row=4, column=25, value="Refer.")
        _merge(ws, 4, 26, 29, "Valor")

        linha = 5
        lotacao_atual = None
        for lotacao, folha, evento, valor in itens:
            if lotacao != lotacao_atual:
                _merge(ws, linha, 1, 29, lotacao)
                lotacao_atual = lotacao
                linha += 1
            _linha_dados(ws, linha, "1", "SERVIDOR", "000.000.000-00",
                         folha, evento, "0.0", "0.0", valor)
            linha += 1

        caminho = tmp_path / nome
        wb.save(caminho)
        return str(caminho)
    return _fazer


@pytest.fixture
def fazer_modelo(tmp_path):
    """Monta um modelo com os setores, colunas e LINHAS DE TIPO informados."""
    def _fazer(setores, colunas, tipos, nome="RETENCAO.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "MOLDE"
        ws.cell(row=1, column=1, value="MODELO DE RETENÇÕES")

        linha = 3
        for setor in setores:
            ws.cell(row=linha, column=1, value=setor)
            ws.cell(row=linha + 1, column=1, value="Tipo")
            for i, coluna in enumerate(colunas, start=2):
                ws.cell(row=linha + 1, column=i, value=coluna)
            primeira = linha + 2
            for desloc, tipo in enumerate(tipos):
                ws.cell(row=primeira + desloc, column=1, value=tipo)
            total = primeira + len(tipos)
            ws.cell(row=total, column=1, value="TOTAL")
            for i in range(2, 2 + len(colunas)):
                letra = ws.cell(row=primeira, column=i).column_letter
                ws.cell(row=total, column=i,
                        value=f"=SUM({letra}{primeira}:{letra}{total - 1})")
            linha = total + 2

        caminho = tmp_path / nome
        wb.save(caminho)
        return str(caminho)
    return _fazer
