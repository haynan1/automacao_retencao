# -*- coding: utf-8 -*-
"""A exportação do histórico pelo HTTP: seleção na tela → arquivo na mão."""
import io

import pytest
from openpyxl import load_workbook

import app as app_mod
from services import relatorio


@pytest.fixture
def cliente(historico_tmp):
    for i in range(3):
        historico_tmp.registrar_operacao({
            "id": f"op{i}", "perfil_nome": f"Secretaria {i}", "competencia": f"0{i + 1}/2026",
            "aba_destino": f"ABA{i}", "qtd_lancamentos": 10 * i, "qtd_celulas": i,
            "total_lido": "100.00", "total_preenchido": "90.00",
            "total_fora_escopo": "5.00", "total_estrutura": "0.00",
            "total_pendente": "5.00", "confere": True,
            "graficos": {"Tipo": [{"label": "Mensal", "valor": 90.0}]},
        })
    return app_mod.app.test_client()


def _exportar(cliente, ids, formato=None):
    dados = {"ids": ids}
    if formato:
        dados["formato"] = formato
    return cliente.post("/historico/exportar", data=dados)


# ---------------------------------------------------------------------------
# A tela
# ---------------------------------------------------------------------------

def test_tela_oferece_selecao_e_os_dois_formatos(cliente):
    html = cliente.get("/historico").get_data(as_text=True)
    assert 'id="sel-todas"' in html
    assert html.count('name="ids"') == 3
    assert 'name="formato" value="xlsx"' in html
    assert 'name="formato" value="pdf"' in html


def test_formularios_nao_ficam_aninhados(cliente):
    """<form> dentro de <form> é inválido: o navegador descarta o interno.

    Se a remoção fosse aninhada no formulário de exportação, clicar na
    lixeira exportaria o relatório em vez de apagar a operação.
    """
    html = cliente.get("/historico").get_data(as_text=True)
    assert html.count("<form") == html.count("</form>")

    profundidade = maxima = 0
    for marca in [t for t in html.split("<") if t.startswith("form") or t.startswith("/form")]:
        profundidade += -1 if marca.startswith("/form") else 1
        maxima = max(maxima, profundidade)
    assert maxima == 1


# ---------------------------------------------------------------------------
# A exportação
# ---------------------------------------------------------------------------

def test_exporta_as_operacoes_marcadas_em_excel(cliente):
    resposta = _exportar(cliente, ["op0", "op2"], "xlsx")
    assert resposta.status_code == 200
    assert "spreadsheetml" in resposta.headers["Content-Type"]
    assert "RELATORIO_CONSOLIDADO_" in resposta.headers["Content-Disposition"]
    assert resposta.headers["Content-Disposition"].endswith(".xlsx")

    wb = load_workbook(io.BytesIO(resposta.data))
    assert wb.sheetnames == [relatorio.ABA_COMPILADO, relatorio.ABA_GRAFICO]

    ws = wb[relatorio.ABA_COMPILADO]
    textos = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
    assert "Secretaria 0" in textos and "Secretaria 2" in textos
    assert "Secretaria 1" not in textos   # não foi marcada


def test_exporta_em_pdf(cliente):
    resposta = _exportar(cliente, ["op0", "op2"], "pdf")
    assert resposta.status_code == 200
    assert resposta.headers["Content-Type"] == "application/pdf"
    assert resposta.headers["Content-Disposition"].endswith(".pdf")
    assert resposta.data.startswith(b"%PDF-")


def test_formato_ausente_ou_desconhecido_cai_para_excel(cliente):
    """Um formato inventado não pode gerar arquivo com extensão mentirosa."""
    for dados in ({"ids": ["op0"]}, {"ids": ["op0"], "formato": "exe"}):
        resposta = cliente.post("/historico/exportar", data=dados)
        assert resposta.status_code == 200
        assert "spreadsheetml" in resposta.headers["Content-Type"]
        assert resposta.headers["Content-Disposition"].endswith(".xlsx")


def test_os_dois_formatos_compilam_as_mesmas_operacoes(cliente):
    from services import historico as hist

    esperado = relatorio.compilar(hist.buscar_operacoes(["op0", "op1", "op2"]))
    xlsx = load_workbook(io.BytesIO(_exportar(cliente, ["op0", "op1", "op2"], "xlsx").data))
    ws = xlsx[relatorio.ABA_COMPILADO]
    linha = next(r for r in range(1, ws.max_row + 1)
                 if ws.cell(row=r, column=1).value == "Total lido (bruto)")
    assert ws.cell(row=linha, column=2).value == float(esperado["resumo"]["lido"])

    import pypdfium2 as pdfium
    doc = pdfium.PdfDocument(io.BytesIO(_exportar(cliente, ["op0", "op1", "op2"], "pdf").data))
    texto = "\n".join(p.get_textpage().get_text_range() for p in doc)
    from services.utils import formatar_moeda
    assert formatar_moeda(esperado["resumo"]["lido"]) in texto


def test_sem_selecao_recusa_com_mensagem_util(cliente):
    resposta = cliente.post("/historico/exportar", data={})
    assert resposta.status_code == 400
    assert "Marque ao menos uma operação" in resposta.get_data(as_text=True)


def test_todos_os_ids_desconhecidos_recusam(cliente):
    assert _exportar(cliente, ["fantasma", "outro"]).status_code == 400


def test_id_desconhecido_no_meio_nao_invalida_a_seleção(cliente):
    """Uma aba aberta há tempo pode citar operação já apagada em outra aba.

    Isso não é motivo para negar o resto — o relatório sai com o que existe.
    """
    resposta = _exportar(cliente, ["op1", "ja-apagada"])
    assert resposta.status_code == 200
    ws = load_workbook(io.BytesIO(resposta.data))[relatorio.ABA_COMPILADO]
    textos = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
    assert "Secretaria 1" in textos


def test_excesso_de_ids_e_cortado_antes_de_ler_o_historico(cliente):
    resposta = _exportar(cliente, ["op0"] * (relatorio.MAX_OPERACOES + 500))
    assert resposta.status_code == 200


def test_exportacao_de_outra_origem_e_bloqueada(cliente):
    """Muda estado? Não. Mas vaza o histórico inteiro para um site aberto."""
    resposta = cliente.post(
        "/historico/exportar", data={"ids": ["op0"]},
        headers={"Origin": "http://site-malicioso.example"},
    )
    assert resposta.status_code == 403


# ---------------------------------------------------------------------------
# O que o histórico grava — o relatório só fecha se o registro fechar
# ---------------------------------------------------------------------------

def test_registro_do_historico_reconcilia_sozinho(
        tmp_path, monkeypatch, perfis_tmp, historico_tmp, fazer_listagem, fazer_modelo):
    """`lido = preenchido + estrutura + fora de escopo + pendente`, sem a sessão.

    Guardar só o fora de escopo das RUBRICAS deixava de fora o das FOLHAS, e
    omitir 'estrutura' escondia o que não achou lugar na planilha — o
    relatório consolidado acusaria uma diferença que não existe.
    """
    from decimal import Decimal

    from services import mapeador, utils

    for nome in ("uploads", "outputs", "sessions"):
        (tmp_path / nome).mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(app_mod, "UPLOADS_DIR", tmp_path / "uploads")
    monkeypatch.setattr(app_mod, "OUTPUTS_DIR", tmp_path / "outputs")
    monkeypatch.setattr(utils, "OUTPUTS_DIR", tmp_path / "outputs")
    monkeypatch.setattr(utils, "SESSIONS_DIR", tmp_path / "sessions")

    lot = "5.0043.0000 - FMS, AGENTE COMUNITÁRIO DE SAÚDE - ACS"
    listagem = fazer_listagem([
        (lot, "MENSAL", "PREVIBELOS", "100.00"),
        (lot, "MENSAL", "INSS", "20.00"),            # rubrica fora de escopo
        (lot, "COMPLEMENTAR", "PREVIBELOS", "30.00"),  # folha que vamos ignorar
    ])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal"])

    # A folha COMPLEMENTAR é marcada como "não preencher" antes do processamento.
    mapeador.salvar_vinculos_folhas("saude", {"COMPLEMENTAR": mapeador.IGNORAR})

    cliente = app_mod.app.test_client()
    dados = {
        "listagem": (io.BytesIO(open(listagem, "rb").read()), "L.xlsx"),
        "modelo": (io.BytesIO(open(modelo, "rb").read()), "M.xlsx"),
        "fonte_modelo": "upload",
    }
    resposta = cliente.post("/analisar", data=dados, content_type="multipart/form-data")
    sid = resposta.headers["Location"].rsplit("/", 1)[-1]
    assert cliente.post(f"/processar/{sid}",
                        data={"aba_destino": "MOLDE", "setor_0": "ACS"}).status_code == 302

    registro = historico_tmp.listar_operacoes()[0]
    partes = sum(Decimal(registro[c]) for c in
                 ("total_preenchido", "total_estrutura", "total_fora_escopo", "total_pendente"))
    assert Decimal(registro["total_lido"]) == partes == Decimal("150.00")
    # Os 20 do INSS (rubrica) e os 30 da folha ignorada, juntos.
    assert Decimal(registro["total_fora_escopo"]) == Decimal("50.00")


def test_falha_inesperada_na_montagem_nao_derruba_o_app(cliente, monkeypatch):
    """O histórico atravessa versões: um registro pode ter forma imprevista.

    Isso não pode virar um 500 cru — o resto do app continua utilizável.
    """
    def explodir(_operacoes):
        raise RuntimeError("forma inesperada de registro")

    monkeypatch.setattr(relatorio, "montar_relatorio", explodir)
    resposta = _exportar(cliente, ["op0"])
    assert resposta.status_code == 500
    assert "Erro ao gerar o relatório" in resposta.get_data(as_text=True)


def test_historico_vazio_nao_mostra_a_barra_de_exportacao(historico_tmp):
    """Estado vazio é desenhado, não é a tela cheia sem linhas."""
    html = app_mod.app.test_client().get("/historico").get_data(as_text=True)
    assert 'name="formato"' not in html          # nenhum botão de exportar
    assert 'name="ids"' not in html              # nenhuma caixa de seleção
    assert "Relatório consolidado" not in html   # nem a barra que os contém
    assert "Nenhuma operação ainda" in html


def test_botoes_de_exportar_nascem_desabilitados(cliente):
    """Com nada marcado, exportar não faz sentido — e o botão precisa dizer isso.

    Um botão que parece vivo e não responde é pior que um botão apagado.
    """
    html = cliente.get("/historico").get_data(as_text=True)
    for formato in ("xlsx", "pdf"):
        marcador = f'name="formato" value="{formato}"'
        botao = html[html.index("<button", html.index(marcador) - 400):]
        assert "disabled" in botao[:botao.index(">")], f"botão {formato} nasce habilitado"


# ---------------------------------------------------------------------------
# A conferência de uma operação, reemitida pelo histórico
# ---------------------------------------------------------------------------

_RETRATO = {
    "competencia": "07/2026", "perfil_nome": "Secretaria de Cultura",
    "aba_destino": "SEC CULTURA", "output_nome": "RETENCAO_PREENCHIDA_20260805_141815.xlsx",
    "qtd_celulas": 7,
    "reconciliacao": {
        "total_lido": "1500.00", "total_preenchido": "1500.00", "total_estrutura": "0.00",
        "total_fora_escopo": "0.00", "total_folha_fora_escopo": "0.00",
        "total_sem_vinculo": "0.00", "total_folha_sem_vinculo": "0.00",
        "total_setor_nao_mapeado": "0.00", "diferenca": "0.00", "confere": True,
    },
    "por_setor": {"ACS": "1500.00"}, "por_coluna": {"CEF": "1500.00"},
    "por_tipo": {"Mensal": "1500.00"},
    "decisoes_folhas": [{"rotulo": "MENSAL", "tipo": "Mensal", "total": "1500.00",
                         "status": "ok", "motivo": "nome igual ao da linha"}],
    "decisoes_rubricas": [], "pendencias_estrutura": [],
}


@pytest.fixture
def cliente_com_retrato(historico_tmp):
    historico_tmp.registrar_operacao({
        "id": "comretrato", "perfil_nome": "Secretaria de Cultura",
        "competencia": "07/2026", "aba_destino": "SEC CULTURA",
        "output_nome": "RETENCAO_PREENCHIDA_20260805_141815.xlsx",
        "total_lido": "1500.00", "total_preenchido": "1500.00", "confere": True,
        "retrato": _RETRATO,
    })
    historico_tmp.registrar_operacao({
        "id": "semretrato", "perfil_nome": "Antiga", "competencia": "01/2026",
        "aba_destino": "X", "total_lido": "10.00", "confere": True,
    })
    return app_mod.app.test_client()


def test_tela_oferece_o_pdf_so_de_quem_tem_retrato(cliente_com_retrato):
    html = cliente_com_retrato.get("/historico").get_data(as_text=True)
    assert "/historico/comretrato/pdf" in html
    assert "/historico/semretrato/pdf" not in html


def test_baixa_a_conferencia_de_uma_operacao_do_historico(cliente_com_retrato):
    """Meses depois, sem a sessão de trabalho (que expira em 24h)."""
    import pypdfium2 as pdfium

    resposta = cliente_com_retrato.get("/historico/comretrato/pdf")
    assert resposta.status_code == 200
    assert resposta.headers["Content-Type"] == "application/pdf"
    # O carimbo é o da planilha que ele confere: os dois ficam lado a lado.
    assert "CONFERENCIA_20260805_141815.pdf" in resposta.headers["Content-Disposition"]

    doc = pdfium.PdfDocument(io.BytesIO(resposta.data))
    texto = "\n".join(p.get_textpage().get_text_range() for p in doc)
    assert "CONFERÊNCIA DA AUTOMAÇÃO DE RETENÇÕES" in texto
    assert "Secretaria de Cultura" in texto and "SEC CULTURA" in texto
    assert "R$ 1.500,00" in texto and "CONFERE" in texto


def test_operacao_anterior_ao_recurso_responde_404(cliente_com_retrato):
    """Sem retrato não há conferência — melhor 404 que um PDF de zeros."""
    assert cliente_com_retrato.get("/historico/semretrato/pdf").status_code == 404


def test_operacao_inexistente_responde_404(cliente_com_retrato):
    assert cliente_com_retrato.get("/historico/naoexiste/pdf").status_code == 404


@pytest.mark.parametrize("op_id", ["../../app", "..%2f..%2fapp", "....//app"])
def test_op_id_com_travessia_nao_serve_arquivo(cliente_com_retrato, op_id):
    resposta = cliente_com_retrato.get(f"/historico/{op_id}/pdf")
    assert resposta.status_code in (404, 308)
    assert not resposta.data.startswith(b"%PDF-")


def test_apagar_a_operacao_tira_o_pdf_do_ar(cliente_com_retrato):
    """O dado sai do disco junto com o registro — nada fica para trás."""
    assert cliente_com_retrato.get("/historico/comretrato/pdf").status_code == 200
    cliente_com_retrato.post("/historico/comretrato/remover")
    assert cliente_com_retrato.get("/historico/comretrato/pdf").status_code == 404


def test_registro_de_outra_versao_nao_derruba_a_tela(historico_tmp):
    """Append-only atravessa versões — e agora o registro fica para sempre.

    Uma operação gravada antes de um campo existir não pode levar a tela
    inteira junto: 200 operações sumiriam por causa de uma.
    """
    historico_tmp.registrar_operacao({"id": "arcaica", "competencia": "01/2024"})
    historico_tmp.registrar_operacao({
        "id": "atual", "perfil_nome": "Saúde", "competencia": "07/2026",
        "aba_destino": "ABA", "total_lido": "10.00", "total_preenchido": "10.00",
        "total_pendente": "0.00", "confere": True, "qtd_celulas": 1,
    })
    resposta = app_mod.app.test_client().get("/historico")
    assert resposta.status_code == 200
    html = resposta.get_data(as_text=True)
    assert "01/2024" in html and "07/2026" in html


def test_operacao_sem_arquivo_nem_retrato_mostra_traco(historico_tmp):
    """Estado vazio desenhado: a célula não pode ficar em branco sem explicação."""
    historico_tmp.registrar_operacao({
        "id": "nada", "perfil_nome": "Antiga", "competencia": "01/2024",
        "aba_destino": "X", "total_lido": "10.00", "total_pendente": "0.00",
        "output_nome": "sumiu_do_disco.xlsx", "confere": True,
    })
    html = app_mod.app.test_client().get("/historico").get_data(as_text=True)
    assert "/historico/nada/pdf" not in html
    assert "Arquivo removido do disco" in html
