# -*- coding: utf-8 -*-
"""Os dois eixos que faltavam: folha → linha e evento → coluna por evento.

Cada teste aqui nasceu de uma perda concreta de controle:

  * toda folha que não fosse 13º somava na linha "Mensal", sem onde discordar;
  * "INSS" e "INSS do 13º salário" viravam um grupo só, sem como separar.

São os dois casos que este arquivo prende no lugar.
"""
from decimal import Decimal

from openpyxl import load_workbook

from services import conferencia, mapeador, preenchimento

LOT = "5.0043.0000 - FMS, AGENTE COMUNITÁRIO DE SAÚDE - ACS"
MAPA = {LOT: "ACS"}
COLUNAS = ["INSS", "INSS 13", "PREVIBELOS"]


def _preparar(caminho_listagem, caminho_modelo, vinculos=None, vinculos_folhas=None,
              fora_de_escopo=()):
    """Roda os três eixos como o app roda, e devolve o estado para inspeção."""
    from services.parser_listagem import extrair_lancamentos

    lanc = extrair_lancamentos(caminho_listagem)["lancamentos"]
    wb = load_workbook(caminho_modelo)
    ws = wb["MOLDE"]
    colunas = preenchimento.listar_colunas_modelo(ws)
    tipos = preenchimento.listar_tipos_modelo(ws)
    regras = mapeador.carregar_regras_rubricas()

    mapeador.aplicar_mapeamentos(lanc, MAPA, regras)
    mapeador.resolver_colunas(lanc, colunas, list(fora_de_escopo), vinculos or {})
    mapeador.resolver_tipos(lanc, tipos, vinculos_folhas or {})
    return lanc, ws, colunas, tipos


def _valor(ws, bloco, tipo, coluna):
    linha = bloco["linhas_tipo"][preenchimento.chave_tipo(tipo)]
    return ws.cell(row=linha, column=bloco["colunas"][coluna]["coluna"]).value


# ===========================================================================
# Eixo 3 — Folha → Linha
# ===========================================================================

def test_ferias_vao_para_a_linha_de_ferias_quando_ela_existe(fazer_listagem, fazer_modelo):
    """A demanda: parar de somar tudo na mensal.

    Com uma linha de FÉRIAS no molde, o lançamento de férias vai para ela —
    e a linha Mensal fica só com o que é mensal.
    """
    listagem = fazer_listagem([
        (LOT, "MENSAL", "PREVIBELOS", "100.00"),
        (LOT, "FÉRIAS", "PREVIBELOS", "30.00"),
        (LOT, "13º SALÁRIO", "PREVIBELOS", "10.00"),
    ])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal", "13º salário", "Férias"])

    lanc, ws, _colunas, tipos = _preparar(listagem, modelo)
    assert tipos == ["Mensal", "13º salário", "Férias"]

    destinos = {r["folha"]: r["tipo_destino"] for r in lanc}
    assert destinos == {"MENSAL": "Mensal", "FÉRIAS": "Férias", "13º SALÁRIO": "13º salário"}
    assert all(r["folha_status"] == "ok" for r in lanc)

    agregados, _ = conferencia.agregar_lancamentos(lanc)
    blocos = preenchimento.localizar_blocos_setores(ws)
    preenchimento.preencher_valores(ws, agregados, blocos)
    bloco = blocos[0]
    assert _valor(ws, bloco, "Mensal", "PREVIBELOS") == 100.00
    assert _valor(ws, bloco, "Férias", "PREVIBELOS") == 30.00
    assert _valor(ws, bloco, "13º salário", "PREVIBELOS") == 10.00


def test_sem_linha_propria_ferias_somam_na_mensal_mas_marcadas(fazer_listagem, fazer_modelo):
    """Nada regride: o comportamento antigo continua, agora VISÍVEL.

    Sem linha de férias, o valor ainda soma na mensal — mas o status é
    'sugerido' e o motivo diz por quê, em vez de acontecer no escuro.
    """
    listagem = fazer_listagem([
        (LOT, "MENSAL", "PREVIBELOS", "100.00"),
        (LOT, "FÉRIAS", "PREVIBELOS", "30.00"),
    ])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal", "13º salário"])

    lanc, ws, _c, _t = _preparar(listagem, modelo)
    ferias = next(r for r in lanc if r["folha"] == "FÉRIAS")
    assert ferias["tipo_destino"] == "Mensal"
    assert ferias["folha_status"] == "sugerido"
    assert "não tem linha própria" in ferias["folha_motivo"]

    agregados, _ = conferencia.agregar_lancamentos(lanc)
    blocos = preenchimento.localizar_blocos_setores(ws)
    preenchimento.preencher_valores(ws, agregados, blocos)
    assert _valor(ws, blocos[0], "Mensal", "PREVIBELOS") == 130.00


def test_folha_desconhecida_nao_recebe_destino_automatico(fazer_listagem, fazer_modelo):
    """Preencher no escuro é pior que segurar o valor e perguntar."""
    listagem = fazer_listagem([(LOT, "ADIANTAMENTO XYZ", "PREVIBELOS", "50.00")])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal", "13º salário"])

    lanc, _ws, _c, _t = _preparar(listagem, modelo)
    assert lanc[0]["tipo_destino"] is None
    assert lanc[0]["folha_status"] == "sem_vinculo"

    _agregados, baldes = conferencia.agregar_lancamentos(lanc)
    assert baldes["folha_sem_vinculo"] == Decimal("50.00")
    assert baldes["preenchivel"] == Decimal("0.00")


def test_vinculo_de_folha_aprendido_vence_a_sugestao(fazer_listagem, fazer_modelo):
    """A decisão de quem opera é a palavra final — e sobrevive ao mês seguinte."""
    listagem = fazer_listagem([(LOT, "RESCISÃO", "PREVIBELOS", "70.00")])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal", "13º salário", "Rescisão"])

    # Sem vínculo, a linha de mesmo nome já é encontrada sozinha.
    lanc, _ws, _c, _t = _preparar(listagem, modelo)
    assert lanc[0]["tipo_destino"] == "Rescisão"

    # Com vínculo salvo, manda o vínculo — mesmo contra o nome igual.
    lanc, _ws, _c, _t = _preparar(
        listagem, modelo, vinculos_folhas={"RESCISAO": "13º salário"}
    )
    assert lanc[0]["tipo_destino"] == "13º salário"
    assert lanc[0]["folha_status"] == "ok"


def test_folha_marcada_para_ignorar_nao_preenche_e_reconcilia(fazer_listagem, fazer_modelo):
    listagem = fazer_listagem([
        (LOT, "MENSAL", "PREVIBELOS", "100.00"),
        (LOT, "COMPLEMENTAR", "PREVIBELOS", "25.00"),
    ])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal"])

    lanc, _ws, _c, _t = _preparar(
        listagem, modelo, vinculos_folhas={"COMPLEMENTAR": mapeador.IGNORAR}
    )
    compl = next(r for r in lanc if r["folha"] == "COMPLEMENTAR")
    assert compl["folha_status"] == "fora_escopo" and compl["tipo_destino"] is None

    _ag, baldes = conferencia.agregar_lancamentos(lanc)
    rec = conferencia.reconciliar(conferencia.calcular_totais_lidos(lanc), baldes, [])
    assert rec["confere"]
    assert rec["total_folha_fora_escopo"] == Decimal("25.00")


def test_vinculo_de_folha_para_linha_inexistente_vira_pendencia(fazer_listagem, fazer_modelo):
    """Molde trocado não pode fazer o vínculo antigo escrever no lugar errado."""
    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "100.00")])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal"])

    lanc, _ws, _c, _t = _preparar(
        listagem, modelo, vinculos_folhas={"MENSAL": "Linha Que Sumiu"}
    )
    assert lanc[0]["folha_status"] == "sem_vinculo"
    assert "não é linha desta planilha" in lanc[0]["folha_motivo"]


# ===========================================================================
# Eixo 2 — um grupo por EVENTO, não por rubrica
# ===========================================================================

def test_inss_e_inss_do_13_sao_grupos_separados(fazer_listagem, fazer_modelo):
    """A segunda demanda: os dois batem na regra 'contém INSS' e viravam um só."""
    listagem = fazer_listagem([
        (LOT, "MENSAL", "INSS", "100.00"),
        (LOT, "13º SALÁRIO", "INSS DO 13º SALÁRIO", "40.00"),
    ])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal", "13º salário"])

    lanc, _ws, colunas, _t = _preparar(listagem, modelo)
    grupos = mapeador.construir_grupos_rubricas(lanc, colunas, [], {})
    rotulos = {g["rotulo"] for g in grupos}
    assert rotulos == {"INSS", "INSS DO 13º SALÁRIO"}

    # Cada grupo carrega a folha de onde veio — a visão que faltava.
    por_rotulo = {g["rotulo"]: g for g in grupos}
    assert por_rotulo["INSS"]["folhas"] == ["MENSAL"]
    assert por_rotulo["INSS DO 13º SALÁRIO"]["folhas"] == ["13º SALÁRIO"]


def test_evento_especifico_pode_ir_para_outra_coluna_que_a_rubrica(
        fazer_listagem, fazer_modelo):
    """O ponto todo: separar um evento sem arrastar o outro junto."""
    listagem = fazer_listagem([
        (LOT, "MENSAL", "INSS", "100.00"),
        (LOT, "13º SALÁRIO", "INSS DO 13º SALÁRIO", "40.00"),
    ])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal", "13º salário"])

    vinculos = {"INSS DO 13O SALARIO": "INSS 13"}
    lanc, ws, _c, _t = _preparar(listagem, modelo, vinculos=vinculos)

    destinos = {r["descricao_original"]: r["coluna_destino"] for r in lanc}
    assert destinos == {"INSS": "INSS", "INSS DO 13º SALÁRIO": "INSS 13"}

    agregados, _ = conferencia.agregar_lancamentos(lanc)
    blocos = preenchimento.localizar_blocos_setores(ws)
    preenchimento.preencher_valores(ws, agregados, blocos)
    assert _valor(ws, blocos[0], "Mensal", "INSS") == 100.00
    assert _valor(ws, blocos[0], "13º salário", "INSS 13") == 40.00


def test_vinculo_da_rubrica_serve_de_padrao_para_os_eventos(fazer_listagem, fazer_modelo):
    """Granularidade não pode virar trabalho manual repetido.

    Um vínculo no nome da RUBRICA vale para todos os eventos dela — é o
    padrão herdado. O vínculo do evento específico sobrepõe só aquele.
    """
    listagem = fazer_listagem([
        (LOT, "MENSAL", "12/120 - EMPRESTIMO CEF 1", "100.00"),
        (LOT, "MENSAL", "36/60 - EMPRESTIMO CEF 2", "60.00"),
    ])
    modelo = fazer_modelo(["ACS"], ["CEF", "PREVIBELOS", "OUTRA"],
                          ["Mensal", "13º salário"])

    lanc, _ws, colunas, _t = _preparar(listagem, modelo, vinculos={"CEF": "CEF"})
    assert {r["coluna_destino"] for r in lanc} == {"CEF"}
    grupos = mapeador.construir_grupos_rubricas(lanc, colunas, [], {"CEF": "CEF"})
    assert len(grupos) == 2  # dois eventos distintos, mesmo destino
    assert all("rubrica “CEF”" in g["motivo"] for g in grupos)

    # Um deles vai para outra coluna sem mexer no outro.
    vinculos = {"CEF": "CEF", "EMPRESTIMO CEF 2": "OUTRA"}
    lanc, _ws, _c, _t = _preparar(listagem, modelo, vinculos=vinculos)
    destinos = {r["descricao_original"]: r["coluna_destino"] for r in lanc}
    assert destinos["12/120 - EMPRESTIMO CEF 1"] == "CEF"
    assert destinos["36/60 - EMPRESTIMO CEF 2"] == "OUTRA"


def test_motivo_explica_a_decisao_de_cada_evento(fazer_listagem, fazer_modelo):
    """Poder discordar do sistema começa por conseguir ler o que ele fez."""
    listagem = fazer_listagem([
        (LOT, "MENSAL", "IRRF", "10.00"),
        (LOT, "MENSAL", "PREVIBELOS", "20.00"),
        (LOT, "MENSAL", "EVENTO SEM REGRA NENHUMA", "30.00"),
    ])
    modelo = fazer_modelo(["ACS"], ["IR", "PREVIBELOS"], ["Mensal"])

    lanc, _ws, colunas, _t = _preparar(listagem, modelo)
    motivos = {r["descricao_original"]: (r["rubrica_status"], r["rubrica_motivo"])
               for r in lanc}

    status, motivo = motivos["IRRF"]
    assert status == "regra" and "IRRF" in motivo  # mostra o termo que bateu

    status, motivo = motivos["PREVIBELOS"]
    assert status == "ok" and "mesmo nome" in motivo

    status, motivo = motivos["EVENTO SEM REGRA NENHUMA"]
    assert status == "sem_vinculo"

    grupos = mapeador.construir_grupos_rubricas(lanc, colunas, [], {})
    assert grupos[0]["status"] == "sem_vinculo"  # pendentes primeiro
    assert grupos[1]["status"] == "regra"        # deduzidos antes dos decididos


def test_evento_fora_de_escopo_pode_ser_resgatado_individualmente(
        fazer_listagem, fazer_modelo):
    """INSS é fora de escopo por configuração — mas a decisão é reversível.

    Marcar a rubrica inteira como fora de escopo não pode impedir que UM
    evento dela seja mandado para uma coluna quando isso for o certo.
    """
    listagem = fazer_listagem([
        (LOT, "MENSAL", "INSS", "100.00"),
        (LOT, "13º SALÁRIO", "INSS DO 13º SALÁRIO", "40.00"),
    ])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal", "13º salário"])

    lanc, _ws, _c, _t = _preparar(
        listagem, modelo,
        vinculos={"INSS DO 13O SALARIO": "INSS 13"},
        fora_de_escopo=["INSS"],
    )
    por_evento = {r["descricao_original"]: r for r in lanc}
    assert por_evento["INSS"]["rubrica_status"] == "fora_escopo"
    assert por_evento["INSS DO 13º SALÁRIO"]["coluna_destino"] == "INSS 13"


# ===========================================================================
# Reconciliação — a invariante continua fechando com os eixos novos
# ===========================================================================

def test_reconciliacao_fecha_com_todos_os_motivos(fazer_listagem, fazer_modelo):
    listagem = fazer_listagem([
        (LOT, "MENSAL", "PREVIBELOS", "100.00"),          # preenche
        (LOT, "MENSAL", "INSS", "20.00"),                  # rubrica fora de escopo
        (LOT, "MENSAL", "EVENTO DESCONHECIDO", "30.00"),   # evento sem coluna
        (LOT, "PAGAMENTO AVULSO", "PREVIBELOS", "40.00"),  # folha sem linha
        ("5.9999.0000 - FMS, SEM SETOR", "MENSAL", "PREVIBELOS", "50.00"),  # sem setor
    ])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal"])

    lanc, _ws, _c, _t = _preparar(listagem, modelo, fora_de_escopo=["INSS"])
    _agregados, baldes = conferencia.agregar_lancamentos(lanc)
    total_lido = conferencia.calcular_totais_lidos(lanc)
    rec = conferencia.reconciliar(total_lido, baldes, [])

    assert rec["confere"], f"diferença {rec['diferenca']}"
    assert rec["total_lido"] == Decimal("240.00")
    assert rec["total_preenchido"] == Decimal("100.00")
    assert rec["total_fora_escopo"] == Decimal("20.00")
    assert rec["total_sem_vinculo"] == Decimal("30.00")
    assert rec["total_folha_sem_vinculo"] == Decimal("40.00")
    assert rec["total_setor_nao_mapeado"] == Decimal("50.00")


def test_setor_sem_a_linha_escolhida_vira_pendencia_e_nao_some(fazer_listagem, tmp_path):
    """Blocos heterogêneos: o menu oferece o que nem todo setor tem.

    `listar_tipos_modelo` usa como referência o bloco com MAIS linhas — um
    menu pobre esconderia destinos válidos. O preço é que a linha oferecida
    pode não existir em outro bloco. O que não pode acontecer é o valor
    sumir: ele vira pendência de estrutura, sai do total preenchido e a
    reconciliação continua fechando ao centavo.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "MOLDE"
    linha = 1
    for setor, tipos in (("ACS", ["Mensal", "Férias"]), ("CAPS", ["Mensal"])):
        ws.cell(row=linha, column=1, value=setor)
        ws.cell(row=linha + 1, column=1, value="Tipo")
        ws.cell(row=linha + 1, column=2, value="PREVIBELOS")
        for i, tipo in enumerate(tipos):
            ws.cell(row=linha + 2 + i, column=1, value=tipo)
        ws.cell(row=linha + 2 + len(tipos), column=1, value="TOTAL")
        linha += len(tipos) + 4
    modelo = tmp_path / "HETEROGENEO.xlsx"
    wb.save(modelo)

    listagem = fazer_listagem([
        ("5.0043.0000 - FMS, AGENTE COMUNITÁRIO DE SAÚDE - ACS", "FÉRIAS", "PREVIBELOS", "30.00"),
        ("5.0044.0000 - FMS, CAPS", "FÉRIAS", "PREVIBELOS", "20.00"),
    ])
    lanc, ws2, _c, tipos = _preparar(
        listagem, str(modelo),
        vinculos_folhas={},
    )
    # O menu oferece Férias, vinda do bloco mais completo.
    assert tipos == ["Mensal", "Férias"]
    for reg in lanc:  # o segundo setor precisa ser mapeado à mão aqui
        if "CAPS" in reg["lotacao_original"]:
            reg["setor_destino"] = "CAPS"

    agregados, baldes = conferencia.agregar_lancamentos(lanc)
    blocos = preenchimento.localizar_blocos_setores(ws2)
    relatorio = preenchimento.preencher_valores(ws2, agregados, blocos)

    # ACS tem a linha; CAPS não — e isso aparece por escrito.
    assert _valor(ws2, blocos[0], "Férias", "PREVIBELOS") == 30.00
    motivos = [p["motivo"] for p in relatorio["pendencias_estrutura"]]
    assert any("Férias" in m and "não encontrada" in m for m in motivos)

    rec = conferencia.reconciliar(
        conferencia.calcular_totais_lidos(lanc), baldes, relatorio["pendencias_estrutura"]
    )
    assert rec["confere"], f"diferença {rec['diferenca']}"
    assert rec["total_preenchido"] == Decimal("30.00")   # só o que coube
    assert rec["total_estrutura"] == Decimal("20.00")    # o que não coube, visível


def test_celula_com_formula_nunca_e_sobrescrita(fazer_listagem, fazer_modelo):
    """Guarda de dinheiro: fórmula é resultado, não destino de lançamento."""
    from openpyxl import load_workbook as abrir

    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "100.00")])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal"])

    wb = abrir(modelo)
    ws = wb["MOLDE"]
    blocos = preenchimento.localizar_blocos_setores(ws)
    alvo = blocos[0]["linhas_tipo"][preenchimento.chave_tipo("Mensal")]
    coluna = blocos[0]["colunas"]["PREVIBELOS"]["coluna"]
    ws.cell(row=alvo, column=coluna).value = "=1+1"   # alguém pôs fórmula aqui
    wb.save(modelo)

    lanc, ws2, _c, _t = _preparar(listagem, modelo)
    agregados, baldes = conferencia.agregar_lancamentos(lanc)
    blocos = preenchimento.localizar_blocos_setores(ws2)
    preenchimento.limpar_area_lancamento(ws2, blocos)
    relatorio = preenchimento.preencher_valores(ws2, agregados, blocos)

    assert ws2.cell(row=alvo, column=coluna).value == "=1+1"
    assert any("fórmula" in p["motivo"] for p in relatorio["pendencias_estrutura"])
    rec = conferencia.reconciliar(
        conferencia.calcular_totais_lidos(lanc), baldes, relatorio["pendencias_estrutura"]
    )
    assert rec["confere"] and rec["total_preenchido"] == Decimal("0.00")


def test_pendencias_separam_decisao_de_deducao(fazer_listagem, fazer_modelo):
    listagem = fazer_listagem([
        (LOT, "FÉRIAS", "PREVIBELOS", "30.00"),
        (LOT, "COISA ESTRANHA", "PREVIBELOS", "10.00"),
    ])
    modelo = fazer_modelo(["ACS"], COLUNAS, ["Mensal"])

    lanc, _ws, _c, _t = _preparar(listagem, modelo)
    pend = mapeador.detectar_pendencias(lanc)

    assert "COISA ESTRANHA" in pend["folhas_sem_vinculo"]
    assert any("FÉRIAS" in k for k in pend["folhas_sugeridas"])
    assert pend["lotacoes_nao_mapeadas"] == {}
