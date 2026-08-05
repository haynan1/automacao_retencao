# -*- coding: utf-8 -*-
"""O fluxo inteiro pelo HTTP: upload → prévia → mapeamento → arquivo.

Os testes de serviço provam a regra; este prova que ela chega ao usuário.
É o único lugar onde um erro de Jinja, um campo de formulário com nome
trocado ou um índice de menu fora de ordem aparecem — e qualquer um dos
três faria a escolha da tela ser silenciosamente ignorada.
"""
import io
from pathlib import Path

import pytest
from openpyxl import load_workbook

import app as app_mod
from services import preenchimento, utils

LOT = "5.0043.0000 - FMS, AGENTE COMUNITÁRIO DE SAÚDE - ACS"


@pytest.fixture
def cliente(tmp_path, monkeypatch, perfis_tmp, historico_tmp):
    """Cliente HTTP com uploads, saídas, sessões e histórico descartáveis.

    `historico_tmp` não é decoração: `/processar` registra a operação, e sem
    ele cada rodada de teste escrevia no histórico real da máquina.
    """
    for nome in ("uploads", "outputs", "sessions"):
        (tmp_path / nome).mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(app_mod, "UPLOADS_DIR", tmp_path / "uploads")
    monkeypatch.setattr(app_mod, "OUTPUTS_DIR", tmp_path / "outputs")
    monkeypatch.setattr(utils, "UPLOADS_DIR", tmp_path / "uploads")
    monkeypatch.setattr(utils, "OUTPUTS_DIR", tmp_path / "outputs")
    monkeypatch.setattr(utils, "SESSIONS_DIR", tmp_path / "sessions")
    return app_mod.app.test_client()


def _enviar(cliente, listagem, modelo):
    """Faz o upload dos dois arquivos e devolve o id da sessão."""
    dados = {
        "listagem": (io.BytesIO(Path(listagem).read_bytes()), "ListagemEventos.xlsx"),
        "modelo": (io.BytesIO(Path(modelo).read_bytes()), "RETENCAO.xlsx"),
        "fonte_modelo": "upload",
    }
    resposta = cliente.post("/analisar", data=dados, content_type="multipart/form-data")
    assert resposta.status_code == 302, resposta.data[:400]
    return resposta.headers["Location"].rsplit("/", 1)[-1]


def _indice(grupos, rotulo):
    """Posição do grupo na tela — é ela que nomeia o campo do formulário."""
    return next(i for i, g in enumerate(grupos) if g["rotulo"] == rotulo)


def test_fluxo_completo_com_os_tres_eixos(cliente, fazer_listagem, fazer_modelo):
    """Escolhas feitas na tela têm de aparecer no arquivo baixado.

    Cobre as duas demandas de uma vez: FÉRIAS numa linha própria (eixo 3) e
    o INSS do 13º numa coluna diferente do INSS (eixo 2).
    """
    listagem = fazer_listagem([
        (LOT, "MENSAL", "INSS", "100.00"),
        (LOT, "13º SALÁRIO", "INSS DO 13º SALÁRIO", "40.00"),
        (LOT, "FÉRIAS", "PREVIBELOS", "30.00"),
        (LOT, "MENSAL", "PREVIBELOS", "70.00"),
    ])
    modelo = fazer_modelo(
        ["ACS"], ["INSS", "INSS 13", "PREVIBELOS"], ["Mensal", "13º salário", "Férias"]
    )

    sid = _enviar(cliente, listagem, modelo)

    # --- as telas renderizam e mostram o terceiro eixo -----------------
    previa = cliente.get(f"/preview/{sid}")
    assert previa.status_code == 200
    assert "Folhas → linha da planilha" in previa.get_data(as_text=True)

    tela = cliente.get(f"/mapeamento/{sid}")
    assert tela.status_code == 200
    html = tela.get_data(as_text=True)
    assert "Eixo 3" in html and "Folha → Linha da planilha" in html
    # Os eventos aparecem separados, cada um com seu menu.
    assert "INSS DO 13º SALÁRIO" in html

    # --- decide na tela ------------------------------------------------
    sessao = utils.carregar_sessao(sid)
    grupos, folhas = sessao["grupos"], sessao["grupos_folhas"]

    formulario = {
        "aba_destino": "MOLDE",
        "setor_0": "ACS",
        f"coluna_{_indice(grupos, 'INSS DO 13º SALÁRIO')}": "INSS 13",
        f"tipo_{_indice(folhas, 'FÉRIAS')}": "Férias",
    }
    processado = cliente.post(f"/processar/{sid}", data=formulario)
    assert processado.status_code == 302, processado.data[:400]

    resultado = cliente.get(f"/resultado/{sid}")
    assert resultado.status_code == 200
    assert "Como cada folha foi destinada" in resultado.get_data(as_text=True)

    # --- o arquivo entregue reflete as escolhas ------------------------
    baixado = cliente.get(f"/download/{sid}")
    assert baixado.status_code == 200
    wb = load_workbook(io.BytesIO(baixado.data))
    ws = wb["MOLDE"]
    bloco = preenchimento.localizar_blocos_setores(ws)[0]

    def valor(tipo, coluna):
        linha = bloco["linhas_tipo"][preenchimento.chave_tipo(tipo)]
        return ws.cell(row=linha, column=bloco["colunas"][coluna]["coluna"]).value

    # Eixo 3: férias na linha de férias, e a mensal só com o que é mensal.
    assert valor("Férias", "PREVIBELOS") == 30.00
    assert valor("Mensal", "PREVIBELOS") == 70.00

    # Eixo 2: "INSS" é fora de escopo por configuração (não é retenção para
    # credor) e continua sem preencher. Mas o "INSS DO 13º SALÁRIO" foi
    # resgatado no menu e foi para a própria coluna — que é exatamente o que
    # não dava para fazer enquanto os dois eram um grupo só.
    assert valor("Mensal", "INSS") is None
    assert valor("13º salário", "INSS 13") == 40.00

    # A aba de conferência registra o porquê de cada destino.
    conf = wb["CONFERÊNCIA_AUTOMAÇÃO"]
    texto = " ".join(str(c.value) for linha in conf.iter_rows() for c in linha if c.value)
    assert "COMO O SISTEMA DECIDIU" in texto
    assert "Folha → linha da planilha" in texto


def test_aba_com_grade_diferente_nao_e_oferecida(cliente, fazer_listagem, tmp_path):
    """Trocar de aba não pode trocar o mapa por baixo da decisão."""
    from openpyxl import Workbook

    wb = Workbook()
    for nome, tipos in (("IGUAL", ["Mensal", "Férias"]), ("DIFERENTE", ["Mensal"])):
        ws = wb.create_sheet(title=nome)
        ws.cell(row=1, column=1, value="ACS")
        ws.cell(row=2, column=1, value="Tipo")
        ws.cell(row=2, column=2, value="PREVIBELOS")
        for i, tipo in enumerate(tipos):
            ws.cell(row=3 + i, column=1, value=tipo)
        ws.cell(row=3 + len(tipos), column=1, value="TOTAL")
    wb.remove(wb["Sheet"])
    modelo = tmp_path / "MODELO.xlsx"
    wb.save(modelo)

    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "10.00")])
    sid = _enviar(cliente, listagem, str(modelo))

    abas = utils.carregar_sessao(sid)["abas_modelo"]
    assert abas == ["IGUAL"]


def test_folha_sem_linha_aparece_como_pendencia_na_tela(cliente, fazer_listagem, fazer_modelo):
    """O que não tem destino precisa gritar antes de o arquivo ser baixado."""
    listagem = fazer_listagem([(LOT, "ADIANTAMENTO XYZ", "PREVIBELOS", "50.00")])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal"])

    sid = _enviar(cliente, listagem, modelo)
    html = cliente.get(f"/preview/{sid}").get_data(as_text=True)
    assert "Folhas sem linha" in html
    assert "ADIANTAMENTO XYZ" in html


# ---------------------------------------------------------------------------
# Guardas — o que o app NÃO pode aceitar
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("campo,lixo", [
    ("setor_0", "SETOR QUE NÃO EXISTE"),
    ("coluna_0", "COLUNA INVENTADA"),
    ("tipo_0", "LINHA INVENTADA"),
])
def test_escolha_fora_do_menu_e_recusada_e_nao_e_gravada(
        cliente, fazer_listagem, fazer_modelo, campo, lixo):
    """Formulário velho não pode gravar um destino que ninguém escolheu.

    O valor ia direto para o JSON do perfil e voltava como pendência todo
    mês, sem explicação — e, por ser texto livre, sem teto de tamanho.
    """
    from services import mapeador

    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "10.00")])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal"])
    sid = _enviar(cliente, listagem, modelo)

    resposta = cliente.post(f"/processar/{sid}", data={
        "aba_destino": "MOLDE", campo: lixo, "salvar_mapeamento": "on",
    })
    assert resposta.status_code == 400
    assert "não existe nesta planilha" in resposta.get_data(as_text=True)

    # E nada foi aprendido a partir do pedido recusado.
    assert mapeador.carregar_vinculos("saude") == {}
    assert mapeador.carregar_vinculos_folhas("saude") == {}
    assert mapeador.carregar_mapeamento_lotacoes("saude") == {}


def test_escolha_gigante_e_cortada_antes_de_normalizar(cliente, fazer_listagem, fazer_modelo):
    """Campo enorme é recusado pelo tamanho, sem passar pelo NFKD.

    Acima de ~500 KB o próprio Werkzeug corta o formulário (413). A faixa
    entre um rótulo plausível e esse teto é que fica por nossa conta: sem a
    guarda, cada campo desses viraria uma normalização Unicode inteira.
    """
    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "10.00")])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal"])
    sid = _enviar(cliente, listagem, modelo)

    resposta = cliente.post(f"/processar/{sid}", data={
        "aba_destino": "MOLDE", "tipo_0": "M" * 100_000,
    })
    assert resposta.status_code == 400
    assert "não existe nesta planilha" in resposta.get_data(as_text=True)


def test_escolha_valida_com_caixa_diferente_e_aceita_na_forma_do_modelo(
        cliente, fazer_listagem, fazer_modelo):
    """A guarda não pode ser rígida a ponto de recusar o que é a mesma coisa."""
    from services import mapeador

    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "10.00")])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal", "13º salário"])
    sid = _enviar(cliente, listagem, modelo)

    resposta = cliente.post(f"/processar/{sid}", data={
        "aba_destino": "MOLDE", "setor_0": "acs", "tipo_0": "13o SALARIO",
        "salvar_mapeamento": "on",
    })
    assert resposta.status_code == 302

    # Grava a forma canônica do modelo, não o texto digitado.
    assert mapeador.carregar_vinculos_folhas("saude") == {"MENSAL": "13º salário"}
    assert mapeador.carregar_mapeamento_lotacoes("saude") == {LOT: "ACS"}


def test_modelo_sem_linha_de_tipo_e_recusado_no_inicio(cliente, fazer_listagem, tmp_path):
    """Recusar aqui custa um clique; descobrir no fim custa a conferência."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "MOLDE"
    ws.cell(row=1, column=1, value="ACS")
    ws.cell(row=2, column=1, value="Tipo")
    ws.cell(row=2, column=2, value="PREVIBELOS")
    ws.cell(row=3, column=1, value="TOTAL")  # bloco sem nenhuma linha de valor
    modelo = tmp_path / "SEM_TIPOS.xlsx"
    wb.save(modelo)

    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "10.00")])
    dados = {
        "listagem": (io.BytesIO(Path(listagem).read_bytes()), "L.xlsx"),
        "modelo": (io.BytesIO(Path(modelo).read_bytes()), "M.xlsx"),
        "fonte_modelo": "upload",
    }
    resposta = cliente.post("/analisar", data=dados, content_type="multipart/form-data")
    assert resposta.status_code == 422
    assert "linhas de tipo" in resposta.get_data(as_text=True)


# ---------------------------------------------------------------------------
# A conferência em PDF, na tela de resultado
# ---------------------------------------------------------------------------

def _processar_ate_o_resultado(cliente, fazer_listagem, fazer_modelo):
    """Roda o fluxo inteiro e devolve o id da sessão já processada."""
    listagem = fazer_listagem([
        (LOT, "MENSAL", "PREVIBELOS", "70.00"),
        (LOT, "FÉRIAS", "PREVIBELOS", "30.00"),
    ])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal", "Férias"])
    sid = _enviar(cliente, listagem, modelo)
    resposta = cliente.post(f"/processar/{sid}", data={
        "aba_destino": "MOLDE", "setor_0": "ACS",
    })
    assert resposta.status_code == 302, resposta.data[:400]
    return sid


def test_tela_de_resultado_oferece_os_dois_formatos(cliente, fazer_listagem, fazer_modelo):
    sid = _processar_ate_o_resultado(cliente, fazer_listagem, fazer_modelo)
    html = cliente.get(f"/resultado/{sid}").get_data(as_text=True)
    assert f"/download/{sid}" in html
    assert f"/resultado/{sid}/pdf" in html


def test_baixa_a_conferencia_em_pdf(cliente, fazer_listagem, fazer_modelo):
    """O PDF sai do MESMO retrato que a tela e a aba da planilha mostram."""
    import pypdfium2 as pdfium

    from services.utils import formatar_moeda

    sid = _processar_ate_o_resultado(cliente, fazer_listagem, fazer_modelo)
    resposta = cliente.get(f"/resultado/{sid}/pdf")

    assert resposta.status_code == 200
    assert resposta.headers["Content-Type"] == "application/pdf"
    disposicao = resposta.headers["Content-Disposition"]
    assert "CONFERENCIA_" in disposicao and disposicao.endswith(".pdf")
    assert resposta.data.startswith(b"%PDF-")

    doc = pdfium.PdfDocument(io.BytesIO(resposta.data))
    texto = "\n".join(p.get_textpage().get_text_range() for p in doc)
    resumo = utils.carregar_sessao(sid)["resumo"]
    assert formatar_moeda(resumo["reconciliacao"]["total_lido"]) in texto
    assert "CONFERE" in texto
    assert "PREVIBELOS" in texto and "Férias" in texto


def test_pdf_de_sessao_nao_processada_volta_ao_mapeamento(cliente, fazer_listagem, fazer_modelo):
    """Sem processamento não existe conferência para emitir."""
    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "10.00")])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal"])
    sid = _enviar(cliente, listagem, modelo)

    resposta = cliente.get(f"/resultado/{sid}/pdf")
    assert resposta.status_code == 302
    assert resposta.headers["Location"].endswith(f"/mapeamento/{sid}")


def test_pdf_de_sessao_inexistente_e_404(cliente):
    assert cliente.get("/resultado/naoexiste/pdf").status_code == 404


def test_sid_com_travessia_nao_le_arquivo_de_fora_da_pasta_de_sessoes(cliente):
    """O sid entra num caminho de arquivo: não pode virar `../`."""
    for sid in ("../../config/perfis", "..%2f..%2fconfig%2fperfis",
                "....//....//config/perfis", "%2e%2e%2f%2e%2e%2fapp"):
        resposta = cliente.get(f"/resultado/{sid}/pdf")
        assert resposta.status_code in (404, 308), f"{sid} -> {resposta.status_code}"
        assert not resposta.data.startswith(b"%PDF-")


def test_retrato_corrompido_na_sessao_vira_erro_tratado(cliente, fazer_listagem,
                                                        fazer_modelo, monkeypatch):
    """Sessão de outra versão não pode devolver traceback nem derrubar a app."""
    sid = _processar_ate_o_resultado(cliente, fazer_listagem, fazer_modelo)
    sessao = utils.carregar_sessao(sid)
    sessao["resumo"] = "isto não é um retrato"
    utils.salvar_sessao(sid, sessao)

    resposta = cliente.get(f"/resultado/{sid}/pdf")
    assert resposta.status_code == 500
    corpo = resposta.get_data(as_text=True)
    assert "Traceback" not in corpo and "AttributeError" not in corpo
    assert "planilha preenchida continua disponível" in corpo


def test_nome_de_saida_adulterado_nao_injeta_cabecalho_http(cliente, fazer_listagem,
                                                            fazer_modelo):
    """`Content-Disposition` é montado com o nome do arquivo: CRLF ali dentro
    partiria a resposta em dois."""
    sid = _processar_ate_o_resultado(cliente, fazer_listagem, fazer_modelo)
    sessao = utils.carregar_sessao(sid)
    sessao["output_nome"] = "planilha\r\nSet-Cookie: roubado=1\r\n\r\n.xlsx"
    utils.salvar_sessao(sid, sessao)

    resposta = cliente.get(f"/resultado/{sid}/pdf")
    assert resposta.status_code == 200
    assert "roubado" not in str(resposta.headers)
    assert "Set-Cookie" not in resposta.headers
    assert resposta.headers["Content-Disposition"].endswith(".pdf")


def test_o_que_fica_para_sempre_nao_tem_dado_pessoal(cliente, fazer_listagem, fazer_modelo):
    """A contrapartida da permanência.

    A Listagem enviada tem nome, matrícula e CPF de cada servidor e é apagada
    em 24h. O que FICA — a planilha preenchida, a aba de conferência, o
    retrato no histórico e o PDF — só pode conter agregados. Se um nome
    vazasse para cá, ele ficaria na máquina para sempre.
    """
    import re

    from services import historico as hist

    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "10.00")])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal"])
    sid = _enviar(cliente, listagem, modelo)
    cliente.post(f"/processar/{sid}", data={"aba_destino": "MOLDE", "setor_0": "ACS"})

    cpf = re.compile(r"\d{3}\.\d{3}\.\d{3}-\d{2}")
    pessoais = ("SERVIDOR", "funcionario", "matricula")

    permanentes = {
        "planilha": cliente.get(f"/download/{sid}").data.decode("latin-1"),
        "pdf": cliente.get(f"/resultado/{sid}/pdf").data.decode("latin-1"),
    }
    op = hist.listar_operacoes()[0]
    permanentes["registro"] = str(op)
    permanentes["retrato"] = str(hist.carregar_retrato(op["id"]))

    for onde, conteudo in permanentes.items():
        assert not cpf.search(conteudo), f"CPF encontrado em: {onde}"
        for termo in pessoais:
            assert termo.lower() not in conteudo.lower(), f"'{termo}' encontrado em: {onde}"


def test_o_arquivo_enviado_com_pii_e_o_unico_que_expira(cliente, fazer_listagem, fazer_modelo):
    """Prova a assimetria de ponta a ponta: a saída fica, o upload sai."""
    listagem = fazer_listagem([(LOT, "MENSAL", "PREVIBELOS", "10.00")])
    modelo = fazer_modelo(["ACS"], ["PREVIBELOS"], ["Mensal"])
    sid = _enviar(cliente, listagem, modelo)
    cliente.post(f"/processar/{sid}", data={"aba_destino": "MOLDE", "setor_0": "ACS"})

    saida = Path(utils.carregar_sessao(sid)["output_path"])
    enviados = list(utils.UPLOADS_DIR.iterdir())
    assert saida.exists() and enviados

    import os
    import time
    velho = time.time() - 30 * 24 * 3600
    for arq in [saida] + enviados:
        os.utime(arq, (velho, velho))

    utils.limpar_temporarios()

    assert saida.exists(), "a planilha entregue não pode expirar"
    assert not any(a.exists() for a in enviados), "o arquivo com PII tem de sair"
