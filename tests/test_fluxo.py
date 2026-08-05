# -*- coding: utf-8 -*-
"""Fluxo de negócio: mapeamento → agregação → reconciliação → preenchimento."""
from decimal import Decimal

from openpyxl import load_workbook

from services import conferencia, mapeador, preenchimento
from services.parser_listagem import extrair_lancamentos


def _preparar(listagem, modelo):
    r = extrair_lancamentos(listagem)
    lanc = r["lancamentos"]
    cfg = mapeador.carregar_config_rubricas()
    wb = load_workbook(modelo)
    ws = wb["MOLDE"]
    colunas = preenchimento.listar_colunas_modelo(ws)
    tipos = preenchimento.listar_tipos_modelo(ws)
    setores = [b["setor"] for b in preenchimento.localizar_blocos_setores(ws)]
    # lotação → setor: ACS mapeia; a "nova" fica sem setor (pendência proposital).
    mapa = {"5.0043.0000 - FMS, AGENTE COMUNITÁRIO DE SAÚDE - ACS": "ACS"}
    mapeador.aplicar_mapeamentos(lanc, mapa, cfg["regras"])
    mapeador.resolver_colunas(lanc, colunas, cfg["fora_de_escopo"], {})
    mapeador.resolver_tipos(lanc, tipos, {})
    return r, lanc, wb, ws, colunas, setores


def test_bloco_sem_linha_total_nao_engole_o_bloco_seguinte(tmp_path):
    """Regressão: a varredura parava no nome do próximo setor e o consumia.

    Em molde com linha TOTAL (o caso real) nada muda — mas um molde sem ela
    perdia um bloco inteiro em silêncio, e com ele todo o dinheiro do setor.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    linha = 1
    for setor in ("ADMINISTRATIVO", "ACS", "SAMU"):
        ws.cell(row=linha, column=1, value=setor)
        ws.cell(row=linha + 1, column=1, value="Tipo")
        ws.cell(row=linha + 1, column=2, value="PREVIBELOS")
        ws.cell(row=linha + 2, column=1, value="Mensal")
        ws.cell(row=linha + 3, column=1, value="13º salário")
        linha += 5  # sem linha TOTAL, com uma linha em branco entre blocos

    blocos = preenchimento.localizar_blocos_setores(ws)
    assert [b["setor"] for b in blocos] == ["ADMINISTRATIVO", "ACS", "SAMU"]
    assert all(b["tipos"] == ["Mensal", "13º salário"] for b in blocos)
    assert all(b["linha_total"] is None for b in blocos)


def test_total_do_evento_excluido(modelo):
    wb = load_workbook(modelo)
    colunas = preenchimento.listar_colunas_modelo(wb["MOLDE"])
    assert "TOTAL DO EVENTO" not in colunas
    assert "PREVIBELOS" in colunas and "CEF" in colunas


def test_reconciliacao_bate_ao_centavo(listagem, modelo):
    _, lanc, wb, ws, _, _ = _preparar(listagem, modelo)
    agregados, baldes = conferencia.agregar_lancamentos(lanc)
    blocos = preenchimento.localizar_blocos_setores(ws)
    preenchimento.limpar_area_lancamento(ws, blocos)
    rel = preenchimento.preencher_valores(ws, agregados, blocos)
    total_lido = conferencia.calcular_totais_lidos(lanc)
    rec = conferencia.reconciliar(total_lido, baldes, rel["pendencias_estrutura"])
    assert rec["confere"], f"diferença {rec['diferenca']}"
    # INSS é fora de escopo; a lotação nova fica sem setor.
    assert rec["total_fora_escopo"] == Decimal("120.00")
    assert rec["total_setor_nao_mapeado"] == Decimal("100.00")


def test_preenche_valor_e_preserva_formula(listagem, modelo):
    _, lanc, wb, ws, _, _ = _preparar(listagem, modelo)
    agregados, _ = conferencia.agregar_lancamentos(lanc)
    blocos = preenchimento.localizar_blocos_setores(ws)
    preenchimento.limpar_area_lancamento(ws, blocos)
    preenchimento.preencher_valores(ws, agregados, blocos)
    acs = next(b for b in blocos if b["setor"] == "ACS")
    col_prev = acs["colunas"][preenchimento.normalizar_texto("PREVIBELOS")]["coluna"]
    linha_mensal = acs["linhas_tipo"][preenchimento.chave_tipo("Mensal")]
    # PREVIBELOS ACS Mensal = 698.39 (o de FÉRIAS também soma em Mensal = +694.06)
    assert abs(ws.cell(row=linha_mensal, column=col_prev).value - (698.39 + 694.06)) < 0.001
    # A célula TOTAL continua sendo fórmula.
    tot = ws.cell(row=acs["linha_total"], column=col_prev).value
    assert isinstance(tot, str) and tot.startswith("=")
