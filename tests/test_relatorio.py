# -*- coding: utf-8 -*-
"""O compilado do histórico e seus dois formatos.

A arquitetura que estes testes protegem: `compilar()` faz a conta uma vez, e
o .xlsx e o .pdf apenas desenham. Se os dois pudessem divergir num centavo,
um deles estaria mentindo e ninguém saberia qual — por isso há um teste que
confere número por número entre os dois formatos.

Depois disso, o que mais importa:

  * o arquivo é uma PLANILHA, não uma foto dela — valores são números, os
    totais são fórmulas e o gráfico aponta para células visíveis;
  * o histórico atravessa versões, então um registro sem a dimensão nova
    tem de compilar junto — e o que falta precisa aparecer por escrito,
    nunca sumir em silêncio;
  * nada de PII e nada que o Excel execute.
"""
import io
import re
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from services import pdf_kit, relatorio, relatorio_pdf
from services.utils import formatar_moeda

_COMPLETA = {
    "id": "op1", "datahora": "01/07/2026 09:10:00",
    "perfil_nome": "Secretaria da Saúde (FMS)", "competencia": "06/2026",
    "aba_destino": "FMS JUNHO", "qtd_lancamentos": 812, "qtd_celulas": 73,
    "total_lido": "124530.55", "total_preenchido": "98120.30",
    "total_fora_escopo": "20410.25", "total_estrutura": "0.00",
    "total_pendente": "6000.00", "confere": True,
    "graficos": {
        "Rubrica": [{"label": "IR", "valor": 40000.0}, {"label": "CEF", "valor": 58120.3}],
        "Tipo": [{"label": "Mensal", "valor": 88120.3}, {"label": "13º salário", "valor": 10000.0}],
        "Setor": [{"label": "ACS", "valor": 50000.0}, {"label": "CAPS", "valor": 48120.3}],
        "Evento": [{"label": "IRRF", "valor": 40000.0},
                   {"label": "12/120 - EMPRESTIMO CEF 1", "valor": 58120.3}],
    },
}

# Como o registro saía de uma versão anterior: sem `graficos`, sem
# `total_estrutura`, e com um campo que virou lixo em disco.
_ANTIGA = {
    "id": "op0", "datahora": "15/05/2026 08:00:00", "perfil": "saude",
    "competencia": "04/2026", "aba_destino": "ABRIL",
    "total_lido": "1000.00", "total_preenchido": "corrompido",
    "total_pendente": None, "confere": False,
}


def _dimensao(dados, chave):
    return next((d for d in dados["dimensoes"] if d["chave"] == chave), None)


def _planilha(operacoes):
    buffer = io.BytesIO()
    relatorio.montar_relatorio(operacoes).save(buffer)
    buffer.seek(0)
    return load_workbook(buffer)


def _texto_do_pdf(operacoes) -> str:
    import pypdfium2 as pdfium

    doc = pdfium.PdfDocument(io.BytesIO(relatorio_pdf.montar_pdf(operacoes)))
    return "\n".join(pagina.get_textpage().get_text_range() for pagina in doc)


# ===========================================================================
# Compilação — a conta, feita uma vez só
# ===========================================================================

def test_resumo_soma_o_periodo_e_a_diferenca_fecha():
    dados = relatorio.compilar([_COMPLETA, _COMPLETA])
    resumo = dados["resumo"]
    assert resumo["lido"] == Decimal("249061.10")
    assert resumo["preenchido"] == Decimal("196240.60")
    assert resumo["diferenca"] == Decimal("0.00")
    assert resumo["operacoes"] == 2
    assert resumo["lancamentos"] == 1624


def test_por_secretaria_sai_das_operacoes_e_nunca_falta():
    """Não depende de série salva: funciona com registro de qualquer versão."""
    dados = relatorio.compilar([_COMPLETA, _ANTIGA])
    dimensao = _dimensao(dados, "Secretaria")
    assert dict(dimensao["itens"]) == {
        "Secretaria da Saúde (FMS)": Decimal("98120.30"),
        "saude": Decimal("0.00"),
    }
    assert dimensao["sem_detalhe"] == 0


@pytest.mark.parametrize("chave,esperado", [
    ("Setor", {"ACS": Decimal("50000.00"), "CAPS": Decimal("48120.30")}),
    ("Rubrica", {"IR": Decimal("40000.00"), "CEF": Decimal("58120.30")}),
    ("Evento", {"IRRF": Decimal("40000.00"),
                "12/120 - EMPRESTIMO CEF 1": Decimal("58120.30")}),
])
def test_dimensoes_saem_das_series_salvas(chave, esperado):
    dados = relatorio.compilar([_COMPLETA])
    assert dict(_dimensao(dados, chave)["itens"]) == esperado


def test_soma_a_mesma_categoria_entre_operacoes_e_ordena_por_valor():
    outra = dict(_COMPLETA, id="op2",
                 graficos={"Setor": [{"label": "ACS", "valor": 100.0},
                                     {"label": "SAMU", "valor": 5.0}]})
    itens = _dimensao(relatorio.compilar([_COMPLETA, outra]), "Setor")["itens"]
    assert itens[0] == ("ACS", Decimal("50100.00"))   # 50000 + 100
    assert itens[-1] == ("SAMU", Decimal("5.00"))


def test_operacao_sem_a_dimensao_e_denunciada_por_escrito():
    """O que falta não pode sumir: um compilado incompleto que parece
    completo é pior que não ter compilado nenhum."""
    dimensao = _dimensao(relatorio.compilar([_COMPLETA, _ANTIGA]), "Setor")
    assert dimensao["sem_detalhe"] == 1
    assert dimensao["valor_sem_detalhe"] == Decimal("0.00")   # a antiga tem lixo

    outra = dict(_COMPLETA, id="op3", graficos={})
    dimensao = _dimensao(relatorio.compilar([_COMPLETA, outra]), "Setor")
    assert dimensao["sem_detalhe"] == 1
    assert dimensao["valor_sem_detalhe"] == Decimal("98120.30")


def test_dimensao_que_ninguem_registrou_some_da_lista():
    """Seção vazia é ruído; a ausência já está dita nas outras dimensões."""
    dados = relatorio.compilar([_ANTIGA])
    assert _dimensao(dados, "Evento") is None
    assert _dimensao(dados, "Secretaria") is not None   # esta sempre existe


def test_valor_ilegivel_ou_ausente_vira_zero_nunca_numero_inventado():
    dados = relatorio.compilar([_ANTIGA])
    assert dados["operacoes"][0]["preenchido"] == Decimal("0.00")
    assert dados["operacoes"][0]["estrutura"] == Decimal("0.00")
    assert dados["operacoes"][0]["pendente"] == Decimal("0.00")


def test_serie_com_lixo_e_ignorada_item_a_item():
    op = dict(_COMPLETA, graficos={"Setor": [
        {"label": "ACS", "valor": 10.0}, {"label": "", "valor": 5.0},
        {"label": "X", "valor": "não é número"}, "nem é dicionário",
    ]})
    itens = _dimensao(relatorio.compilar([op]), "Setor")["itens"]
    assert itens == [("ACS", Decimal("10.00")), ("X", Decimal("0.00"))]


def test_graficos_nao_dicionario_nao_derruba():
    dados = relatorio.compilar([dict(_COMPLETA, graficos="lixo")])
    assert _dimensao(dados, "Setor") is None


def test_lista_vazia_e_recusada():
    with pytest.raises(relatorio.SemOperacoes):
        relatorio.compilar([])


def test_excesso_de_operacoes_e_cortado_no_teto():
    muitas = [dict(_COMPLETA, id=f"op{i}") for i in range(relatorio.MAX_OPERACOES + 25)]
    assert relatorio.compilar(muitas)["resumo"]["operacoes"] == relatorio.MAX_OPERACOES


# ---------------------------------------------------------------------------
# Tabela mostra tudo; só o gráfico corta
# ---------------------------------------------------------------------------

def test_compilado_nao_trunca_a_lista():
    """Quem pede um compilado quer o compilado inteiro."""
    serie = [{"label": f"S{i}", "valor": float(i + 1)} for i in range(40)]
    dimensao = _dimensao(relatorio.compilar([dict(_COMPLETA, graficos={"Setor": serie})]), "Setor")
    assert len(dimensao["itens"]) == 40
    assert dimensao["total"] == sum(Decimal(str(i["valor"])) for i in serie)


def test_grafico_corta_mas_preserva_a_soma():
    serie = [{"label": f"S{i}", "valor": float(i + 1)} for i in range(40)]
    dimensao = _dimensao(relatorio.compilar([dict(_COMPLETA, graficos={"Setor": serie})]), "Setor")
    recorte = relatorio.para_grafico(dimensao)

    assert len(recorte) == relatorio.TOP_GRAFICO + 1
    assert recorte[-1][0].startswith("outras (")
    assert sum((v for _k, v in recorte), Decimal("0")) == dimensao["total"]


# ===========================================================================
# XLSX
# ===========================================================================

def test_duas_abas_na_ordem_de_leitura():
    assert _planilha([_COMPLETA]).sheetnames == [relatorio.ABA_COMPILADO, relatorio.ABA_GRAFICO]


def test_valores_sao_numeros_formatados_e_nao_texto():
    """Gravar "R$ 1.234,56" deixaria a planilha bonita e inútil: não somaria."""
    ws = _planilha([_COMPLETA])[relatorio.ABA_COMPILADO]
    linha = next(r for r in range(1, ws.max_row + 1)
                 if ws.cell(row=r, column=1).value == "Total lido (bruto)")
    celula = ws.cell(row=linha, column=2)
    assert celula.value == pytest.approx(124530.55)
    assert isinstance(celula.value, float)
    assert celula.number_format == "R$ #,##0.00"


def test_totais_das_secoes_sao_formula_viva():
    """Apagar uma linha no Excel tem de corrigir o total sozinho."""
    ws = _planilha([_COMPLETA])[relatorio.ABA_COMPILADO]
    totais = [r for r in range(1, ws.max_row + 1)
              if ws.cell(row=r, column=1).value == "TOTAL"]
    # Uma por dimensão mais a das operações compiladas.
    assert len(totais) == len(relatorio.compilar([_COMPLETA])["dimensoes"]) + 1

    for linha in totais:
        # A seção de operações soma em duas colunas; as demais, em uma.
        celulas = [ws.cell(row=linha, column=c).value for c in range(2, 7)]
        formulas = [v for v in celulas if isinstance(v, str) and v.startswith("=SUM(")]
        assert formulas, f"linha TOTAL {linha} não tem fórmula de soma"


def test_planilha_traz_todas_as_secoes_do_compilado():
    ws = _planilha([_COMPLETA])[relatorio.ABA_COMPILADO]
    coluna_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    for titulo in ("RESUMO GERAL DO PERÍODO", "POR SECRETARIA", "POR SETOR",
                   "POR EVENTO DO RELATÓRIO", "POR RUBRICA (COLUNA DA PLANILHA)",
                   "POR LINHA DE TIPO DE FOLHA", "OPERAÇÕES COMPILADAS"):
        assert titulo in coluna_a


def test_aviso_de_dimensao_incompleta_aparece_na_planilha():
    ws = _planilha([_COMPLETA, dict(_COMPLETA, id="x", graficos={})])[relatorio.ABA_COMPILADO]
    textos = " ".join(str(ws.cell(row=r, column=1).value or "")
                      for r in range(1, ws.max_row + 1))
    assert "não têm este detalhamento" in textos


def test_graficos_apontam_para_celulas_visiveis_da_propria_aba():
    """Nenhum dado escondido dentro do desenho: toda barra tem uma célula."""
    ws = _planilha([_COMPLETA])[relatorio.ABA_GRAFICO]
    assert len(ws._charts) == 6   # composição + as 5 dimensões

    for grafico in ws._charts:
        for serie in grafico.series:
            referencia = serie.val.numRef.f
            assert relatorio.ABA_GRAFICO in referencia
            ultima = int(re.findall(r"\$(\d+)", referencia)[-1])
            assert ultima <= ws.max_row


def test_recalculo_marcado_para_a_abertura():
    """Sem isto o Excel abre as fórmulas de TOTAL em branco."""
    assert relatorio.montar_relatorio([_COMPLETA]).calculation.fullCalcOnLoad is True


# ===========================================================================
# PDF
# ===========================================================================

def test_pdf_e_um_pdf_valido_e_paginado():
    import pypdfium2 as pdfium

    bytes_pdf = relatorio_pdf.montar_pdf([_COMPLETA, _ANTIGA])
    assert bytes_pdf.startswith(b"%PDF-")
    assert len(pdfium.PdfDocument(io.BytesIO(bytes_pdf))) >= 2


def test_pdf_traz_as_mesmas_secoes():
    texto = _texto_do_pdf([_COMPLETA])
    for titulo in ("RELATÓRIO CONSOLIDADO DE RETENÇÕES", "RESUMO GERAL DO PERÍODO",
                   "POR SECRETARIA", "POR SETOR", "POR EVENTO DO RELATÓRIO",
                   "POR LINHA DE TIPO DE FOLHA", "OPERAÇÕES COMPILADAS"):
        assert titulo in texto


def test_pdf_e_planilha_nao_podem_divergir_num_centavo():
    """O teste que sustenta a arquitetura dos dois formatos.

    Ambos leem a mesma compilação; se um número aparecer só num deles, a
    separação entre calcular e desenhar foi rompida em algum ponto.
    """
    operacoes = [_COMPLETA, dict(_COMPLETA, id="op2", perfil_nome="Educação",
                                 competencia="07/2026")]
    dados = relatorio.compilar(operacoes)
    texto = _texto_do_pdf(operacoes)

    esperados = [dados["resumo"]["lido"], dados["resumo"]["preenchido"],
                 dados["resumo"]["fora"], dados["resumo"]["pendente"]]
    for dimensao in dados["dimensoes"]:
        esperados.append(dimensao["total"])
        esperados += [valor for _rotulo, valor in dimensao["itens"]]

    for valor in esperados:
        assert formatar_moeda(valor) in texto, f"{valor} não aparece no PDF"


def test_pdf_avisa_sobre_dimensao_incompleta():
    texto = _texto_do_pdf([_COMPLETA, dict(_COMPLETA, id="x", graficos={})])
    assert "não têm este detalhamento" in texto


def test_pdf_numera_as_paginas_e_declara_a_origem():
    texto = _texto_do_pdf([_COMPLETA])
    assert "Página 1" in texto
    assert "sem dados pessoais" in texto


def test_moeda_usa_o_formato_brasileiro_sem_locale():
    """Um formatador só: o número da tela e o do PDF não podem divergir."""
    assert formatar_moeda(Decimal("1234567.89")) == "R$ 1.234.567,89"
    assert formatar_moeda(Decimal("0")) == "R$ 0,00"
    assert formatar_moeda(Decimal("-42.10")) == "R$ -42,10"
    assert formatar_moeda("não é número") == "R$ 0,00"


# ===========================================================================
# Guardas de segurança
# ===========================================================================

_VENENOS = ("=cmd|'/c calc'!A1", "+1+1", "-2+3", "@SUM(A1)", "=HYPERLINK(\"http://x\")")


@pytest.mark.parametrize("veneno", _VENENOS)
def test_nenhuma_celula_de_texto_vira_formula(veneno):
    """Injeção de fórmula: o relatório é um documento que se compartilha.

    Um nome de secretaria digitado na tela, ou um rótulo de setor lido de um
    .xlsx enviado, chegam aqui como texto. Escritos crus, o Excel os executa
    na máquina de quem abre (DDE) ou exfiltra a planilha (=HYPERLINK).
    """
    op = dict(_COMPLETA, perfil_nome=veneno, aba_destino=veneno, competencia=veneno,
              graficos={"Setor": [{"label": veneno, "valor": 10.0}]})
    wb = _planilha([op])

    permitidas = ("=SUM(",)   # as fórmulas que o app escreve de propósito
    for aba in wb.sheetnames:
        ws = wb[aba]
        for linha in ws.iter_rows():
            for celula in linha:
                valor = celula.value
                if not isinstance(valor, str) or not valor.startswith("="):
                    continue
                assert valor.startswith(permitidas), (
                    f"{aba}!{celula.coordinate} virou fórmula: {valor!r}"
                )


@pytest.mark.parametrize("veneno", _VENENOS)
def test_texto_perigoso_continua_legivel(veneno):
    """Neutralizar não pode significar esconder: o texto tem de aparecer."""
    dados = relatorio.compilar([dict(_COMPLETA, perfil_nome=veneno)])
    assert dados["operacoes"][0]["secretaria"] == "'" + veneno


_MARCACOES = ("<font size=99>", "<b>NEGRITO</b>", "SETOR A & B",
              "<script>alert(1)</script>", "&nbsp;<para>")


@pytest.mark.parametrize("marcacao", _MARCACOES)
def test_marcacao_em_texto_nao_derruba_nem_reformata_o_pdf(marcacao):
    """`Paragraph` do reportlab recebe mini-XML, não texto puro.

    Um "<font size=99>" solto num nome de setor derrubava a geração inteira
    com erro de parse; um "<b>" reformataria o documento à revelia de quem o
    assina. É a contraparte em PDF da injeção de fórmula na planilha.
    """
    op = dict(_COMPLETA, perfil_nome=marcacao, aba_destino=marcacao,
              graficos={"Setor": [{"label": marcacao, "valor": 10.0}]})
    texto = _texto_do_pdf([op])          # não pode levantar
    assert "RELATÓRIO CONSOLIDADO" in texto
    # O texto sai literal, sem virar formatação.
    assert marcacao in texto


@pytest.mark.parametrize("caso,graficos", [
    ("tudo zero", {"Setor": [{"label": "A", "valor": 0.0}],
                   "Tipo": [{"label": "M", "valor": 0.0}]}),
    ("uma fatia só", {"Tipo": [{"label": "Mensal", "valor": 100.0}]}),
    ("estorno", {"Setor": [{"label": "ESTORNO", "valor": -50.0}],
                 "Tipo": [{"label": "Mensal", "valor": -50.0}]}),
    ("valor gigante", {"Setor": [{"label": "X", "valor": 999999999.99}]}),
    ("rótulo enorme", {"Setor": [{"label": "S" * 300, "valor": 10.0}]}),
    ("sem série alguma", {}),
])
def test_numeros_de_borda_nao_derrubam_nenhum_formato(caso, graficos):
    operacoes = [dict(_COMPLETA, graficos=graficos)]
    assert relatorio_pdf.montar_pdf(operacoes).startswith(b"%PDF-")
    assert _planilha(operacoes).sheetnames  # o .xlsx também


def test_estorno_nao_e_escondido_pelo_grafico():
    """Barra negativa com eixo ancorado em zero sumiria; pizza mentiria.

    Um gráfico que omite dinheiro é pior que gráfico nenhum.
    """
    dimensao = {"chave": "Setor", "titulo": "POR SETOR", "rotulo_coluna": "Setor",
                "itens": [("A", Decimal("100.00")), ("ESTORNO", Decimal("-50.00"))],
                "total": Decimal("50.00"), "sem_detalhe": 0,
                "valor_sem_detalhe": Decimal("0.00")}
    desenho = pdf_kit.grafico_barras(dimensao["itens"])
    grafico = desenho.contents[0]
    assert grafico.valueAxis.valueMin is None, "eixo fixo em zero esconderia o estorno"

    # Só positivo: aí sim ancora em zero, para a leitura não ser distorcida.
    desenho = pdf_kit.grafico_barras([("A", Decimal("100.00"))])
    assert desenho.contents[0].valueAxis.valueMin == 0


def test_tipo_com_valor_negativo_deixa_de_ser_pizza():
    """'Parte do todo' não existe com negativo — as fatias mentiriam."""
    from reportlab.graphics.charts.piecharts import Pie

    def _desenhos(valor):
        dim = {"chave": "Tipo", "titulo": "POR LINHA", "rotulo_coluna": "Linha",
               "itens": [("Mensal", Decimal("100.00")), ("Ajuste", Decimal(valor))],
               "total": Decimal("100.00"), "sem_detalhe": 0,
               "valor_sem_detalhe": Decimal("0.00")}
        partes = relatorio_pdf._dimensao(dim, pdf_kit.estilos())
        return [f for parte in partes for f in getattr(parte, "_content", [parte])]

    def _tem_pizza(elementos):
        return any(isinstance(c, Pie)
                   for e in elementos if hasattr(e, "contents") for c in e.contents)

    assert _tem_pizza(_desenhos("10.00"))       # tudo positivo -> pizza
    assert not _tem_pizza(_desenhos("-10.00"))  # com estorno -> barras


def test_pdf_com_muitas_linhas_pagina_em_vez_de_cortar():
    """Uma dimensão longa atravessa páginas; nada some no fim da folha."""
    import pypdfium2 as pdfium

    serie = [{"label": f"EVENTO NUMERO {i:03d}", "valor": float(i + 1)} for i in range(120)]
    op = dict(_COMPLETA, graficos={"Evento": serie})

    bytes_pdf = relatorio_pdf.montar_pdf([op])
    doc = pdfium.PdfDocument(io.BytesIO(bytes_pdf))
    texto = "\n".join(p.get_textpage().get_text_range() for p in doc)
    assert len(doc) >= 3
    assert "EVENTO NUMERO 000" in texto and "EVENTO NUMERO 119" in texto


@pytest.mark.parametrize("veneno", _VENENOS)
def test_aba_de_conferencia_tambem_e_blindada(veneno):
    """A outra superfície que gera planilha corre o mesmo risco.

    A aba de conferência acompanha todo arquivo preenchido e recebe nomes de
    setor e descrições de evento lidos de arquivos enviados — que nenhuma
    validação do construtor de molde alcança.
    """
    from openpyxl import Workbook

    from services import conferencia

    wb = Workbook()
    conferencia.criar_aba_conferencia(wb, {
        "competencia": veneno, "arquivo_origem": veneno, "arquivo_modelo": veneno,
        "aba_destino": veneno,
        "reconciliacao": {
            "total_lido": Decimal("10"), "total_preenchido": Decimal("10"),
            "total_fora_escopo": Decimal("0"), "total_folha_fora_escopo": Decimal("0"),
            "total_sem_vinculo": Decimal("0"), "total_folha_sem_vinculo": Decimal("0"),
            "total_setor_nao_mapeado": Decimal("0"), "total_estrutura": Decimal("0"),
            "diferenca": Decimal("0"), "confere": True,
        },
        "por_setor": {veneno: Decimal("10")}, "por_coluna": {}, "por_tipo": {},
        "decisoes_folhas": [{"rotulo": veneno, "tipo": veneno, "total": Decimal("10"),
                             "status": "ok", "motivo": veneno}],
        "decisoes_rubricas": [],
        "rubricas_sem_vinculo": {veneno: 1}, "folhas_sem_vinculo": {},
        "lotacoes_nao_mapeadas": {}, "rubricas_fora_escopo": {}, "folhas_sugeridas": {},
        "folhas_fora_escopo": {},
        "pendencias_estrutura": [{"motivo": veneno, "setor": veneno,
                                  "rubrica": veneno, "valor": Decimal("1")}],
    })

    ws = wb["CONFERÊNCIA_AUTOMAÇÃO"]
    formulas = [c.value for linha in ws.iter_rows() for c in linha
                if isinstance(c.value, str) and c.value.startswith("=")]
    assert formulas == []


@pytest.mark.parametrize("gerar", [
    lambda ops: io.BytesIO(_planilha_bytes(ops)).getvalue().decode("latin-1"),
    lambda ops: relatorio_pdf.montar_pdf(ops).decode("latin-1"),
])
def test_nenhum_formato_carrega_dados_pessoais(gerar):
    """O histórico não guarda PII; os relatórios derivados também não podem."""
    conteudo = gerar([_COMPLETA, _ANTIGA])
    assert not re.search(r"\d{3}\.\d{3}\.\d{3}-\d{2}", conteudo)   # CPF
    for campo in ("funcionario", "matricula"):
        assert campo not in conteudo.lower()


def _planilha_bytes(operacoes) -> bytes:
    buffer = io.BytesIO()
    relatorio.montar_relatorio(operacoes).save(buffer)
    return buffer.getvalue()
