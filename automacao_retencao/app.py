"""Automação de Retenções FMS — camada web (Flask).

Orquestra o fluxo: upload -> análise -> pré-visualização -> mapeamento ->
processamento -> resultado. Toda a regra de negócio vive em services/.
"""

from __future__ import annotations

import io
import os
import re
import tempfile
import traceback
from decimal import Decimal
from pathlib import Path
from urllib.parse import urlparse

# Carrega variáveis de um arquivo .env, se existir (opcional).
try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

from flask import (
    Flask,
    abort,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from services import conferencia, historico, mapeador, molde, perfis, preenchimento
from services.parser_listagem import extrair_lancamentos
from services.utils import (
    MODELOS_DIR,
    OUTPUTS_DIR,
    UPLOADS_DIR,
    carregar_sessao,
    configurar_logs,
    criar_pastas,
    gerar_nome_saida,
    limpar_temporarios,
    novo_id_sessao,
    salvar_sessao,
)

MAX_UPLOAD_MB = 30
# A spec do construtor de molde e texto puro: o molde real da secretaria dá
# ~4 KB. 512 KB é folga de sobra e mantém o corpo JSON longe do teto de upload.
MAX_SPEC_KB = 512
EXTENSOES_OK = {".xlsx"}

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24).hex())

criar_pastas()
log = configurar_logs()
limpar_temporarios()  # retenção: remove uploads/sessões >24h e saídas >7d


@app.before_request
def _bloqueia_origem_externa():
    """Barra requisições que mudam estado vindas de outra origem (drive-by).

    Como o app não tem auth e roda em localhost, um site aberto no navegador
    poderia POSTar para 127.0.0.1:5000. Só aceitamos POST/etc. cujo Origin
    seja a própria origem do app (ou ausente — curl, form same-origin).
    """
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return None
    origin = request.headers.get("Origin")
    if origin and urlparse(origin).netloc != request.host:
        log.warning("Origem externa bloqueada: %s (host=%s)", origin, request.host)
        abort(403)
    return None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extensao_valida(nome: str) -> bool:
    return os.path.splitext(nome.lower())[1] in EXTENSOES_OK


def _salvar_upload(file_storage, prefixo: str) -> tuple[str, str]:
    """Salva o upload dentro de uploads/ com nome seguro e único.

    Retorna (nome_original, caminho_absoluto). Levanta ValueError se a
    extensão for inválida.
    """
    nome_original = file_storage.filename or ""
    if not _extensao_valida(nome_original):
        raise ValueError(f"Arquivo '{nome_original}' não é .xlsx.")

    base = secure_filename(nome_original) or "arquivo.xlsx"
    destino = UPLOADS_DIR / f"{prefixo}_{base}"
    # Garante que o destino permaneça dentro de uploads/ (defesa em profundidade).
    destino = destino.resolve()
    if UPLOADS_DIR.resolve() not in destino.parents:
        raise ValueError("Caminho de upload inválido.")
    file_storage.save(destino)
    return nome_original, str(destino)


def _sessao_ou_404(sid: str) -> dict:
    dados = carregar_sessao(sid)
    if not dados:
        abort(404)
    return dados


def _detectar_modelo(caminho_modelo: str, aba_preferida: str | None):
    """Abre o modelo e detecta setores, colunas de rubrica e abas.

    Retorna (setores, colunas_modelo, abas, aba_analisada).
    """
    wb = load_workbook(caminho_modelo, data_only=False)
    try:
        abas = list(wb.sheetnames)
        alvo = aba_preferida if aba_preferida in abas else None
        if alvo is None:
            for nome in abas:  # primeira aba que não seja de conferência
                if "CONFER" not in nome.upper():
                    alvo = nome
                    break
        ws = wb[alvo] if alvo else wb.worksheets[0]
        blocos = preenchimento.localizar_blocos_setores(ws)
        setores = [b["setor"] for b in blocos]
        colunas = preenchimento.listar_colunas_modelo(ws)
        abas_compativeis = []
        for nome in abas:
            if "CONFER" in nome.upper():
                continue
            ws_aba = wb[nome]
            blocos_aba = preenchimento.localizar_blocos_setores(ws_aba)
            setores_aba = [b["setor"] for b in blocos_aba]
            colunas_aba = preenchimento.listar_colunas_modelo(ws_aba)
            if setores_aba == setores and colunas_aba == colunas:
                abas_compativeis.append(nome)
        return setores, colunas, abas_compativeis or ([alvo] if alvo else abas), alvo
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Rotas
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    lista = perfis.listar_perfis()
    moldes = {p["slug"]: perfis.info_molde(p["slug"]) for p in lista}
    return render_template(
        "index.html",
        max_mb=MAX_UPLOAD_MB,
        perfis=lista,
        perfil_padrao=perfis.slug_padrao(),
        moldes=moldes,
    )


@app.route("/secretarias")
def secretarias():
    lista = perfis.listar_perfis()
    return render_template(
        "secretarias.html",
        perfis=lista,
        padrao=perfis.slug_padrao(),
        moldes={p["slug"]: perfis.info_molde(p["slug"]) for p in lista},
    )


@app.route("/secretarias/salvar", methods=["POST"])
def secretarias_salvar():
    nome = (request.form.get("nome") or "").strip()
    slug = (request.form.get("slug") or "").strip()
    deteccao = [
        termo.strip()
        for termo in re.split(r"[,\n;]", request.form.get("deteccao") or "")
        if termo.strip()
    ]
    if not nome:
        return render_template(
            "erro.html",
            titulo="Nome obrigatorio",
            mensagem="Informe o nome da secretaria.",
        ), 400

    slug_final = slug if (slug and perfis.perfil_valido(slug)) else None
    perfis.registrar_perfil(nome, deteccao=deteccao, slug=slug_final)
    return redirect(url_for("secretarias"))


@app.route("/secretarias/<slug>/padrao", methods=["POST"])
def secretarias_padrao(slug):
    perfis.definir_padrao(slug)
    return redirect(url_for("secretarias"))


@app.route("/secretarias/<slug>/remover", methods=["POST"])
def secretarias_remover(slug):
    if not perfis.remover_perfil(slug):
        return render_template(
            "erro.html",
            titulo="Nao foi possivel remover",
            mensagem="Nao da para remover o perfil padrao nem o ultimo.",
        ), 400
    return redirect(url_for("secretarias"))


@app.route("/secretarias/<slug>/molde/limpar", methods=["POST"])
def secretarias_limpar_molde(slug):
    if not perfis.perfil_valido(slug):
        return render_template(
            "erro.html",
            titulo="Secretaria invalida",
            mensagem="Nao foi possivel localizar a secretaria informada.",
        ), 400
    perfis.remover_molde(slug)
    return redirect(url_for("secretarias"))


@app.route("/secretarias/<slug>/molde/baixar")
def secretarias_baixar_molde(slug):
    """Baixa o molde fixo da secretaria (para abrir no Excel ou reaproveitar)."""
    if not perfis.perfil_valido(slug) or not perfis.existe_molde(slug):
        abort(404)
    caminho = perfis.caminho_molde(slug).resolve()
    if MODELOS_DIR.resolve() not in caminho.parents:  # guarda anti path traversal
        abort(403)
    info = perfis.info_molde(slug) or {}
    nome = secure_filename(info.get("nome_original") or "") or f"molde_{slug}.xlsx"
    if not nome.lower().endswith(".xlsx"):
        nome += ".xlsx"
    return send_file(caminho, as_attachment=True, download_name=nome)


# ---------------------------------------------------------------------------
# Construtor de molde — a estrutura da planilha desenhada na interface
# ---------------------------------------------------------------------------

def _perfil_ou_404(slug: str) -> str:
    if not perfis.perfil_valido(slug):
        abort(404)
    return slug


def _spec_do_pedido() -> dict:
    """Le a spec do corpo JSON. Corpo ausente/invalido vira ErroDeMolde.

    O teto de 30 MB do app existe para planilhas; uma spec e texto e nunca
    passa de alguns KB. Recusar pelo Content-Length evita desserializar
    megabytes de JSON so para descobrir depois que era absurdo.
    """
    if (request.content_length or 0) > MAX_SPEC_KB * 1024:
        raise molde.ErroDeMolde(
            [f"Estrutura do molde grande demais (limite de {MAX_SPEC_KB} KB)."]
        )
    corpo = request.get_json(silent=True)
    if not isinstance(corpo, dict) or not isinstance(corpo.get("spec"), dict):
        raise molde.ErroDeMolde(["Estrutura do molde não recebida."])
    return corpo["spec"]


@app.route("/molde/<slug>")
def molde_editor(slug):
    """Tela do construtor: abre o desenho salvo, o molde atual ou uma folha em branco."""
    _perfil_ou_404(slug)
    nome = perfis.nome_perfil(slug)
    spec, origem = molde.spec_inicial(slug, nome)
    return render_template(
        "molde.html",
        slug=slug,
        perfil_nome=nome,
        spec=spec,
        origem=origem,
        spec_branco=molde.spec_em_branco(nome),
        tem_molde=perfis.existe_molde(slug),
        info_molde=perfis.info_molde(slug),
        tipos_suportados=list(molde.TIPOS_SUPORTADOS),
        limites={
            "abas": molde.MAX_ABAS,
            "setores": molde.MAX_SETORES,
            "colunas": molde.MAX_COLUNAS,
        },
    )


@app.route("/molde/<slug>/importar")
def molde_importar(slug):
    """Devolve a spec extraida do molde fixo atual (para editar em vez de redesenhar)."""
    _perfil_ou_404(slug)
    if not perfis.existe_molde(slug):
        return jsonify({"ok": False, "problemas": ["Esta secretaria ainda não tem molde fixo."]}), 404
    try:
        return jsonify({"ok": True, "spec": molde.extrair_spec(perfis.caminho_molde(slug))})
    except molde.ErroDeMolde as exc:
        return jsonify({"ok": False, "problemas": exc.problemas}), 422
    except Exception as exc:  # noqa: BLE001
        log.error("Falha ao importar molde do perfil '%s': %s", slug, exc)
        return jsonify({"ok": False, "problemas": ["Não foi possível ler o molde atual."]}), 500


@app.route("/molde/<slug>/previa", methods=["POST"])
def molde_previa(slug):
    """Valida o desenho e devolve a prévia — sem gravar nada em disco."""
    _perfil_ou_404(slug)
    try:
        spec, divergencias = molde.conferir(_spec_do_pedido())
    except molde.ErroDeMolde as exc:
        return jsonify({"ok": False, "problemas": exc.problemas, "divergencias": []}), 200
    except Exception as exc:  # noqa: BLE001
        log.error("Falha na prévia do molde '%s': %s\n%s", slug, exc, traceback.format_exc())
        return jsonify({"ok": False, "problemas": ["Erro inesperado ao montar a prévia."],
                        "divergencias": []}), 500

    return jsonify({
        "ok": not divergencias,
        "problemas": [],
        "divergencias": divergencias,
        "spec": spec,
        "resumo": molde.resumo(spec),
        "grade": molde.grade(spec),
    })


@app.route("/molde/<slug>/baixar", methods=["POST"])
def molde_baixar(slug):
    """Gera o .xlsx e devolve para download, sem tocar no molde fixo."""
    _perfil_ou_404(slug)
    try:
        _spec, wb, divergencias = molde.construir(_spec_do_pedido())
    except molde.ErroDeMolde as exc:
        return jsonify({"ok": False, "problemas": exc.problemas}), 422

    if divergencias:
        return jsonify({"ok": False, "problemas": divergencias}), 422

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"molde_{secure_filename(slug) or 'secretaria'}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/molde/<slug>/salvar", methods=["POST"])
def molde_salvar(slug):
    """Fixa o molde desenhado como padrão da secretaria.

    Só grava depois que o motor de preenchimento releu o arquivo gerado e
    confirmou setores, colunas e tipos. `perfis.definir_molde` guarda um
    backup timestampado do molde anterior.
    """
    _perfil_ou_404(slug)
    try:
        spec, wb, divergencias = molde.construir(_spec_do_pedido())
    except molde.ErroDeMolde as exc:
        return jsonify({"ok": False, "problemas": exc.problemas}), 422

    if divergencias:
        log.warning("Molde do perfil '%s' recusado na verificação: %s", slug, divergencias)
        return jsonify({"ok": False, "problemas": divergencias}), 422

    try:
        with tempfile.TemporaryDirectory() as pasta:
            caminho = Path(pasta) / "molde_construido.xlsx"
            wb.save(caminho)
            perfis.definir_molde(slug, caminho, f"molde_construido_{slug}.xlsx")
        molde.salvar_estrutura(slug, spec)
    except OSError as exc:
        log.error("Falha ao gravar molde do perfil '%s': %s", slug, exc)
        return jsonify({"ok": False, "problemas": ["Não foi possível gravar o molde em disco."]}), 500

    log.info("Molde construído para o perfil '%s': %d setor(es), %d coluna(s), %d aba(s).",
             slug, len(spec["setores"]), len(spec["colunas"]), len(spec["abas"]))
    return jsonify({"ok": True, "redirect": url_for("secretarias")})


@app.route("/analisar", methods=["POST"])
def analisar():
    try:
        f_origem = request.files.get("listagem")
        if not f_origem or not f_origem.filename:
            return render_template(
                "erro.html",
                titulo="Listagem faltando",
                mensagem="Envie a Listagem de Eventos (XLSX) exportada do sistema.",
            ), 400

        sid = novo_id_sessao()
        nome_origem, path_origem = _salvar_upload(f_origem, f"{sid}_origem")

        # --- Perfil (secretaria) ---
        perfil = (request.form.get("perfil") or "").strip() or perfis.slug_padrao()
        if not perfis.perfil_valido(perfil):
            perfil = perfis.slug_padrao()

        # --- Origem do molde: fixo (versionado por perfil) ou upload ---
        fonte_modelo = (request.form.get("fonte_modelo") or "").strip()
        f_modelo = request.files.get("modelo")
        usar_fixo = fonte_modelo == "fixo" or (
            not fonte_modelo and perfis.existe_molde(perfil) and (not f_modelo or not f_modelo.filename)
        )

        if usar_fixo:
            if not perfis.existe_molde(perfil):
                return render_template(
                    "erro.html",
                    titulo="Sem molde fixo",
                    mensagem="Este perfil não tem molde fixo definido. Envie uma planilha modelo.",
                ), 400
            info = perfis.info_molde(perfil) or {}
            nome_modelo = info.get("nome_original", "molde_padrao.xlsx")
            path_modelo = str(perfis.caminho_molde(perfil))
        else:
            if not f_modelo or not f_modelo.filename:
                return render_template(
                    "erro.html",
                    titulo="Molde faltando",
                    mensagem="Envie a planilha modelo (RETENÇÃO) ou escolha o molde fixo.",
                ), 400
            nome_modelo, path_modelo = _salvar_upload(f_modelo, f"{sid}_modelo")
            if request.form.get("definir_molde_fixo") == "on":
                perfis.definir_molde(perfil, path_modelo, nome_modelo)
                log.info("Molde fixo do perfil '%s' atualizado: %s", perfil, nome_modelo)

        # --- Parsing do relatório de origem ---
        cfg = mapeador.carregar_config_rubricas()
        regras_rubricas = cfg["regras"]
        fora_de_escopo = cfg["fora_de_escopo"]
        vinculos = mapeador.carregar_vinculos(perfil)

        resultado = extrair_lancamentos(path_origem)
        lancamentos = resultado["lancamentos"]

        # Autodeteccao de secretaria pelo cabecalho — apenas aviso, nao bloqueia.
        detectado = perfis.detectar_perfil(resultado.get("banners", []))
        aviso_perfil = None
        if detectado and detectado != perfil:
            aviso_perfil = (
                f"A Listagem parece ser de '{perfis.nome_perfil(detectado)}', "
                f"mas você selecionou '{perfis.nome_perfil(perfil)}'. Confira o perfil."
            )

        if not resultado.get("faixas"):
            return render_template(
                "erro.html",
                titulo="Estrutura não reconhecida",
                mensagem=(
                    "Não foi possível localizar os cabeçalhos de lançamento "
                    "(Descrição/Valor) no arquivo de origem. Confirme que é a "
                    "Listagem de Eventos exportada em XLSX."
                ),
            ), 422

        if not lancamentos:
            return render_template(
                "erro.html",
                titulo="Nenhum lançamento encontrado",
                mensagem="O relatório foi lido, mas não há linhas de lançamento válidas.",
            ), 422

        if not resultado["lotacoes"]:
            return render_template(
                "erro.html",
                titulo="Nenhuma lotação detectada",
                mensagem="Não foram encontradas lotações no formato '5.0000.0000 - FMS, ...'.",
            ), 422

        # --- Detecta setores/colunas/abas do modelo ---
        aba_sugerida = None
        wb_tmp = load_workbook(path_modelo, data_only=False, read_only=True)
        try:
            aba_sugerida = preenchimento.localizar_aba_destino(wb_tmp, resultado["competencia"])
        finally:
            wb_tmp.close()

        setores_planilha, colunas_modelo, abas_modelo, _ = _detectar_modelo(path_modelo, aba_sugerida)
        if not setores_planilha:
            return render_template(
                "erro.html",
                titulo="Modelo sem blocos de setor",
                mensagem=(
                    "A planilha modelo não apresentou blocos de setor no formato "
                    "esperado (nome do setor seguido de uma linha 'Tipo | rubricas...')."
                ),
            ), 422

        # --- Mapeamentos (eixo A: lotação→setor; eixo B: rubrica/evento→coluna) ---
        mapa_lotacoes = mapeador.carregar_mapeamento_lotacoes(perfil)
        mapeador.aplicar_mapeamentos(lancamentos, mapa_lotacoes, regras_rubricas)
        mapeador.resolver_colunas(lancamentos, colunas_modelo, fora_de_escopo, vinculos)
        pendencias = mapeador.detectar_pendencias(lancamentos)

        lotacoes_info = []
        for lot in resultado["lotacoes"]:
            setor = mapeador.mapear_lotacao(lot, mapa_lotacoes)
            sugestao = setor or mapeador.sugerir_setor(lot, setores_planilha)
            lotacoes_info.append(
                {"lotacao": lot, "setor_atual": setor, "sugestao": sugestao, "mapeada": bool(setor)}
            )

        grupos = mapeador.construir_grupos_rubricas(
            lancamentos, colunas_modelo, fora_de_escopo, vinculos
        )
        rubricas_identificadas = sorted(
            {g["coluna"] for g in grupos if g["status"] == "ok" and g["coluna"]}
        )

        dados = {
            "id": sid,
            "perfil": perfil,
            "perfil_nome": perfis.nome_perfil(perfil),
            "aviso_perfil": aviso_perfil,
            "competencia": resultado["competencia"],
            "aba_origem": resultado["aba"],
            "arquivo_origem_nome": nome_origem,
            "arquivo_origem_path": path_origem,
            "arquivo_modelo_nome": nome_modelo,
            "arquivo_modelo_path": path_modelo,
            "lancamentos": lancamentos,
            "lotacoes": resultado["lotacoes"],
            "lotacoes_info": lotacoes_info,
            "setores_planilha": setores_planilha,
            "colunas_modelo": colunas_modelo,
            "fora_de_escopo": fora_de_escopo,
            "grupos": grupos,
            "abas_modelo": abas_modelo,
            "aba_sugerida": aba_sugerida,
            "rubricas_identificadas": rubricas_identificadas,
            "pendencias": pendencias,
        }
        salvar_sessao(sid, dados)
        log.info("Análise concluída sid=%s lançamentos=%d lotações=%d grupos=%d",
                 sid, len(lancamentos), len(resultado["lotacoes"]), len(grupos))
        return redirect(url_for("preview", sid=sid))

    except ValueError as exc:
        log.warning("Validação falhou: %s", exc)
        return render_template("erro.html", titulo="Arquivo inválido", mensagem=str(exc)), 400
    except Exception as exc:  # noqa: BLE001 — nunca deixar a app cair
        log.error("Falha na análise: %s\n%s", exc, traceback.format_exc())
        return render_template(
            "erro.html",
            titulo="Erro ao analisar os arquivos",
            mensagem="Ocorreu um erro inesperado. Detalhes técnicos foram registrados em logs/app.log.",
        ), 500


@app.route("/preview/<sid>")
def preview(sid):
    dados = _sessao_ou_404(sid)
    p = dados["pendencias"]
    return render_template(
        "preview.html",
        sid=sid,
        perfil_nome=dados.get("perfil_nome"),
        aviso_perfil=dados.get("aviso_perfil"),
        competencia=dados["competencia"],
        qtd_lancamentos=len(dados["lancamentos"]),
        qtd_lotacoes=len(dados["lotacoes"]),
        rubricas=dados["rubricas_identificadas"],
        lotacoes_info=dados["lotacoes_info"],
        grupos=dados["grupos"],
        lotacoes_nao_mapeadas=p["lotacoes_nao_mapeadas"],
        rubricas_sem_vinculo=p["rubricas_sem_vinculo"],
        rubricas_fora_escopo=p["rubricas_fora_escopo"],
        folhas_desconhecidas=p["folhas_desconhecidas"],
    )


@app.route("/mapeamento/<sid>")
def mapeamento(sid):
    dados = _sessao_ou_404(sid)
    return render_template(
        "mapeamento.html",
        sid=sid,
        lotacoes_info=dados["lotacoes_info"],
        setores=dados["setores_planilha"],
        colunas=dados["colunas_modelo"],
        grupos=dados["grupos"],
        abas=dados["abas_modelo"],
        aba_sugerida=dados["aba_sugerida"],
        competencia=dados["competencia"],
        IGNORAR=mapeador.IGNORAR,
    )


@app.route("/processar/<sid>", methods=["POST"])
def processar(sid):
    dados = _sessao_ou_404(sid)
    try:
        lancamentos = dados["lancamentos"]
        lotacoes = dados["lotacoes"]
        grupos = dados["grupos"]
        fora_de_escopo = dados["fora_de_escopo"]
        perfil = dados.get("perfil") or perfis.slug_padrao()

        # --- Eixo A: lotação → setor (por índice, robusto a acentos) ---
        overrides_lot: dict[str, str] = {}
        for i, lot in enumerate(lotacoes):
            valor = (request.form.get(f"setor_{i}") or "").strip()
            if valor:
                overrides_lot[lot] = valor
        for reg in lancamentos:
            if reg["lotacao_original"] in overrides_lot:
                reg["setor_destino"] = overrides_lot[reg["lotacao_original"]]

        # --- Eixo B: evento/rubrica → coluna (menu suspenso), aprendido ---
        vinculos = mapeador.carregar_vinculos(perfil)
        for i, g in enumerate(grupos):
            escolha = (request.form.get(f"coluna_{i}") or "").strip()
            if escolha:  # vazio = manter automático
                vinculos[g["chave"]] = escolha

        # --- Aba de destino ---
        aba_destino = (request.form.get("aba_destino") or "").strip()
        if aba_destino not in dados["abas_modelo"]:
            return render_template(
                "erro.html", titulo="Aba inválida",
                mensagem="Selecione uma aba de destino existente na planilha modelo.",
            ), 400

        # --- Abre o modelo e usa as colunas REAIS da aba escolhida ---
        wb = load_workbook(dados["arquivo_modelo_path"], data_only=False)
        ws = wb[aba_destino]
        blocos = preenchimento.localizar_blocos_setores(ws)
        colunas_modelo = preenchimento.listar_colunas_modelo(ws)

        # Re-resolve rubrica→coluna já com os vínculos escolhidos.
        mapeador.resolver_colunas(lancamentos, colunas_modelo, fora_de_escopo, vinculos)

        # Persistir mapeamentos aprendidos no perfil, se solicitado.
        if request.form.get("salvar_mapeamento") == "on":
            if overrides_lot:
                mapa = mapeador.carregar_mapeamento_lotacoes(perfil)
                mapa.update(overrides_lot)
                mapeador.salvar_mapeamento_lotacoes(perfil, mapa)
            mapeador.salvar_vinculos(perfil, vinculos)
            log.info("Perfil '%s': %d lotações, %d rubricas salvas.", perfil, len(overrides_lot), len(vinculos))

        # --- Agregação e preenchimento ---
        agregados, baldes = conferencia.agregar_lancamentos(lancamentos)
        preenchimento.limpar_area_lancamento(ws, blocos)
        relatorio = preenchimento.preencher_valores(ws, agregados, blocos)

        # --- Totais, reconciliação e conferência ---
        total_lido = conferencia.calcular_totais_lidos(lancamentos)
        agregados_tot = conferencia.calcular_totais_agregados(agregados)
        rec = conferencia.reconciliar(total_lido, baldes, relatorio["pendencias_estrutura"])
        pendencias_map = mapeador.detectar_pendencias(lancamentos)

        dados_conf = {
            "competencia": dados["competencia"],
            "arquivo_origem": dados["arquivo_origem_nome"],
            "arquivo_modelo": dados["arquivo_modelo_nome"],
            "aba_destino": aba_destino,
            "reconciliacao": rec,
            "por_setor": agregados_tot["por_setor"],
            "por_coluna": agregados_tot["por_coluna"],
            "por_tipo": agregados_tot["por_tipo"],
            "lotacoes_nao_mapeadas": pendencias_map["lotacoes_nao_mapeadas"],
            "rubricas_sem_vinculo": pendencias_map["rubricas_sem_vinculo"],
            "rubricas_fora_escopo": pendencias_map["rubricas_fora_escopo"],
            "folhas_desconhecidas": pendencias_map["folhas_desconhecidas"],
            "pendencias_estrutura": relatorio["pendencias_estrutura"],
        }
        conferencia.criar_aba_conferencia(wb, dados_conf)
        preenchimento.preservar_formulas(wb)

        # --- Salvar saída ---
        nome_saida = gerar_nome_saida()
        caminho_saida = (OUTPUTS_DIR / nome_saida).resolve()
        if OUTPUTS_DIR.resolve() not in caminho_saida.parents:
            raise ValueError("Caminho de saída inválido.")
        wb.save(caminho_saida)
        wb.close()

        dados["output_nome"] = nome_saida
        dados["output_path"] = str(caminho_saida)
        dados["resumo"] = {
            "aba_destino": aba_destino,
            "reconciliacao": rec,
            "por_setor": agregados_tot["por_setor"],
            "por_coluna": agregados_tot["por_coluna"],
            "por_tipo": agregados_tot["por_tipo"],
            "qtd_celulas": len(relatorio["preenchidos"]),
            "pendencias_estrutura": relatorio["pendencias_estrutura"],
            "lotacoes_nao_mapeadas": pendencias_map["lotacoes_nao_mapeadas"],
            "rubricas_sem_vinculo": pendencias_map["rubricas_sem_vinculo"],
            "rubricas_fora_escopo": pendencias_map["rubricas_fora_escopo"],
            "folhas_desconhecidas": pendencias_map["folhas_desconhecidas"],
        }
        salvar_sessao(sid, dados)

        # --- Histórico local (nunca vai para o git) ---
        pendente = (rec["total_sem_vinculo"] + rec["total_setor_nao_mapeado"]
                    + rec["total_folha_desconhecida"])
        graficos_hist = {}
        for gnome, gchave in (("Setor", "por_setor"), ("Rubrica", "por_coluna"), ("Tipo", "por_tipo")):
            gserie = _serie_grafico(agregados_tot.get(gchave))
            if gserie:
                graficos_hist[gnome] = gserie
        historico.registrar_operacao({
            "sid": sid,
            "graficos": graficos_hist,
            "perfil": perfil,
            "perfil_nome": dados.get("perfil_nome"),
            "competencia": dados["competencia"],
            "aba_destino": aba_destino,
            "arquivo_origem": dados["arquivo_origem_nome"],
            "arquivo_modelo": dados["arquivo_modelo_nome"],
            "output_nome": nome_saida,
            "qtd_lancamentos": len(lancamentos),
            "qtd_celulas": len(relatorio["preenchidos"]),
            "total_lido": rec["total_lido"],
            "total_preenchido": rec["total_preenchido"],
            "total_fora_escopo": rec["total_fora_escopo"],
            "total_pendente": pendente,
            "confere": rec["confere"],
            "pendencias": {
                "lotacoes_nao_mapeadas": len(pendencias_map["lotacoes_nao_mapeadas"]),
                "rubricas_sem_vinculo": len(pendencias_map["rubricas_sem_vinculo"]),
                "folhas_desconhecidas": len(pendencias_map["folhas_desconhecidas"]),
            },
        })

        log.info("Processamento concluído sid=%s células=%d confere=%s saída=%s",
                 sid, len(relatorio["preenchidos"]), rec["confere"], nome_saida)
        return redirect(url_for("resultado", sid=sid))

    except KeyError as exc:
        log.error("Aba/estrutura ausente: %s", exc)
        return render_template(
            "erro.html", titulo="Aba não encontrada",
            mensagem=f"Não foi possível abrir a aba selecionada: {exc}.",
        ), 422
    except Exception as exc:  # noqa: BLE001
        log.error("Falha no processamento: %s\n%s", exc, traceback.format_exc())
        return render_template(
            "erro.html",
            titulo="Erro ao preencher a planilha",
            mensagem="Ocorreu um erro inesperado. Detalhes foram registrados em logs/app.log.",
        ), 500


def _serie_grafico(d: dict | None) -> list[dict]:
    """Transforma {rótulo: Decimal} em série [{label, valor(float)}], ordenada
    por valor (desc). float para serializar em JSON no gráfico SVG."""
    if not d:
        return []
    return [
        {"label": k, "valor": float(v)}
        for k, v in sorted(d.items(), key=lambda kv: kv[1], reverse=True)
    ]


@app.route("/resultado/<sid>")
def resultado(sid):
    dados = _sessao_ou_404(sid)
    if "resumo" not in dados:
        return redirect(url_for("mapeamento", sid=sid))
    resumo = dados["resumo"]
    graficos = {}
    for nome, chave in (("Setor", "por_setor"), ("Rubrica", "por_coluna"), ("Tipo", "por_tipo")):
        serie = _serie_grafico(resumo.get(chave))
        if serie:
            graficos[nome] = serie
    return render_template(
        "resultado.html",
        sid=sid,
        competencia=dados["competencia"],
        output_nome=dados["output_nome"],
        resumo=resumo,
        graficos=graficos,
    )


@app.route("/download/<sid>")
def download(sid):
    dados = _sessao_ou_404(sid)
    caminho = dados.get("output_path")
    if not caminho or not os.path.exists(caminho):
        abort(404)
    # Garante que o arquivo servido está dentro de outputs/ (anti path traversal).
    resolvido = Path(caminho).resolve()
    if OUTPUTS_DIR.resolve() not in resolvido.parents:
        abort(403)
    return send_file(resolvido, as_attachment=True, download_name=dados["output_nome"])


@app.route("/historico")
def historico_view():
    operacoes = historico.listar_operacoes()
    # Marca quais saídas ainda existem em disco (podem ter sido limpas após 7 dias).
    for op in operacoes:
        nome = op.get("output_nome") or ""
        op["disponivel"] = bool(nome) and (OUTPUTS_DIR / nome).exists()
    return render_template("historico.html", operacoes=operacoes)


@app.route("/guia")
def guia():
    return render_template("guia.html")


@app.route("/historico/<op_id>/remover", methods=["POST"])
def historico_remover(op_id):
    historico.remover_operacao(op_id)
    return redirect(url_for("historico_view"))


@app.route("/historico/<op_id>/graficos")
def historico_graficos(op_id):
    op = historico.buscar_operacao(op_id)
    if not op:
        abort(404)
    return jsonify(op.get("graficos") or {})


@app.route("/historico/limpar", methods=["POST"])
def historico_limpar():
    n = historico.limpar_historico()
    log.info("Histórico limpo: %d operação(ões) removida(s).", n)
    return redirect(url_for("historico_view"))


@app.route("/historico/baixar/<nome>")
def historico_baixar(nome):
    seguro = secure_filename(nome)
    if not seguro:
        abort(404)
    caminho = (OUTPUTS_DIR / seguro).resolve()
    if OUTPUTS_DIR.resolve() not in caminho.parents or not caminho.exists():
        abort(404)
    return send_file(caminho, as_attachment=True, download_name=seguro)


@app.errorhandler(413)
def arquivo_grande(_):
    return render_template(
        "erro.html",
        titulo="Arquivo muito grande",
        mensagem=f"O limite de upload é {MAX_UPLOAD_MB} MB por arquivo.",
    ), 413


@app.errorhandler(404)
def nao_encontrado(_):
    return render_template(
        "erro.html", titulo="Não encontrado",
        mensagem="A página ou sessão de trabalho não foi encontrada. Recomece o envio.",
    ), 404


@app.template_filter("moeda")
def _moeda(valor) -> str:
    """Formata Decimal/float como moeda brasileira (R$ 1.234,56)."""
    try:
        d = Decimal(str(valor))
    except Exception:
        return "R$ 0,00"
    inteiro, _, dec = f"{d:.2f}".partition(".")
    negativo = inteiro.startswith("-")
    inteiro = inteiro.lstrip("-")
    grupos = []
    while len(inteiro) > 3:
        grupos.insert(0, inteiro[-3:])
        inteiro = inteiro[:-3]
    grupos.insert(0, inteiro)
    formatado = ".".join(grupos)
    sinal = "-" if negativo else ""
    return f"R$ {sinal}{formatado},{dec}"


if __name__ == "__main__":
    # Debug desligado por padrão (o debugger do Werkzeug expõe tracebacks e
    # execução de código). Para desenvolvimento: FLASK_DEBUG=1 python app.py
    debug = os.environ.get("FLASK_DEBUG", "").lower() in ("1", "true", "yes")
    app.run(host="127.0.0.1", port=5000, debug=debug)
