# -*- coding: utf-8 -*-
"""Construtor de molde: validação, geração, verificação pelo motor e round-trip.

O teste que mais importa é `test_molde_construido_processa_de_verdade`: um molde
desenhado na tela tem de atravessar o fluxo real de preenchimento e reconciliar
ao centavo. Sem isso, o resto é decoração.
"""
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from services import conferencia, mapeador, molde, preenchimento
from services.parser_listagem import extrair_lancamentos


def spec_valida(**ajustes) -> dict:
    base = {
        "titulo": "PREFEITURA DE SÃO LUÍS DE MONTES BELOS",
        "subtitulo": "MODELO DE RETENÇÕES — BASE ZERADA",
        "abas": ["MOLDE"],
        "colunas": ["ARSEM", "CEF", "IR", "PREVIBELOS", "IPASGO"],
        "setores": [
            {"nome": "ADMINISTRATIVO", "apelido": "ADM"},
            {"nome": "ACS", "apelido": "ACS"},
        ],
        "tipos": ["Mensal", "13º salário"],
        "opcoes": {
            "coluna_total_evento": True,
            "linha_total_bloco": True,
            "linha_total_geral": True,
            "rodape_apelido": True,
        },
    }
    base.update(ajustes)
    return base


def problemas_de(**ajustes) -> list[str]:
    with pytest.raises(molde.ErroDeMolde) as erro:
        molde.validar_spec(spec_valida(**ajustes))
    return erro.value.problemas


# ---------------------------------------------------------------------------
# Validação — recusa tudo que o motor descartaria em silêncio
# ---------------------------------------------------------------------------

def test_spec_valida_passa_e_normaliza():
    spec = molde.validar_spec(spec_valida(colunas=["  ARSEM  ", "CEF\n", "IR", "PREVIBELOS", "IPASGO"]))
    assert spec["colunas"][0] == "ARSEM"
    assert spec["colunas"][1] == "CEF"
    assert spec["versao"] == molde.SPEC_VERSAO


def test_coluna_repetida_por_acento_ou_caixa_e_recusada():
    problemas = problemas_de(colunas=["SINDSAÚDE", "sindsaude"])
    assert any("repete" in p for p in problemas)


def test_coluna_total_e_recusada():
    # O motor ignora colunas iniciadas por TOTAL: aceitar seria prometer um
    # preenchimento que nunca aconteceria.
    assert any("TOTAL" in p for p in problemas_de(colunas=["ARSEM", "TOTAL DO EVENTO"]))


def test_coluna_tipo_e_recusada():
    assert any("Tipo" in p for p in problemas_de(colunas=["ARSEM", "Tipo"]))


def test_setor_repetido_e_recusado():
    problemas = problemas_de(setores=[{"nome": "SAÚDE BUCAL"}, {"nome": "saude bucal"}])
    assert any("repete" in p for p in problemas)


def test_setor_reservado_e_recusado():
    assert any("TOTAL" in p for p in problemas_de(setores=[{"nome": "TOTAL GERAL"}]))
    assert any("reservada" in p for p in problemas_de(setores=[{"nome": "Tipo"}]))


def test_setor_parecido_com_tipo_so_e_recusado_sem_linha_total():
    # Com a linha TOTAL fechando o bloco, a varredura para antes — nome livre.
    spec = molde.validar_spec(spec_valida(setores=[{"nome": "UBS 13 DE MAIO"}]))
    assert spec["setores"][0]["nome"] == "UBS 13 DE MAIO"

    # Sem ela, o setor seria lido como linha de tipo e engoliria o bloco seguinte.
    opcoes = dict(spec_valida()["opcoes"], linha_total_bloco=False)
    problemas = problemas_de(setores=[{"nome": "UBS 13 DE MAIO"}], opcoes=opcoes)
    assert any("linha de tipo" in p for p in problemas)


def test_texto_que_o_excel_leria_como_formula_e_recusado():
    for veneno in ("=SUM(A1)", "+1+1", "@SUM(A1)"):
        assert any("fórmula" in p for p in problemas_de(colunas=[veneno]))
    assert any("fórmula" in p for p in problemas_de(titulo="=cmd|'/c calc'!A1"))


def test_aba_com_confer_e_recusada():
    # "CONFERÊNCIA_AUTOMAÇÃO" é a aba que o app cria; abas assim são ignoradas.
    assert any("CONFER" in p for p in problemas_de(abas=["CONFERÊNCIA"]))


def test_aba_com_caractere_proibido_pelo_excel_e_recusada():
    assert any("não aceita" in p for p in problemas_de(abas=["FMS/JANEIRO"]))


def test_aba_repetida_ignorando_caixa_e_recusada():
    assert any("repetida" in p for p in problemas_de(abas=["MOLDE", "molde"]))


def test_listas_vazias_sao_recusadas():
    assert any("ao menos uma coluna" in p for p in problemas_de(colunas=[]))
    assert any("ao menos um setor" in p for p in problemas_de(setores=[]))
    assert any("ao menos uma aba" in p for p in problemas_de(abas=[]))
    assert any("ao menos um tipo" in p for p in problemas_de(tipos=[]))


def test_tipo_nao_suportado_e_descartado():
    spec = molde.validar_spec(spec_valida(tipos=["Mensal", "Rescisão"]))
    assert spec["tipos"] == ["Mensal"]


def test_bloco_grande_demais_e_recusado():
    """O bloco (setores × colunas) pesa na prévia, que roda a cada tecla."""
    problemas = problemas_de(
        setores=[{"nome": f"SETOR {i}"} for i in range(100)],
        colunas=[f"RUBRICA {i}" for i in range(100)],
    )
    assert any("células por aba" in p for p in problemas)


def test_arquivo_grande_demais_e_recusado():
    """O total de células pesa na gravação do .xlsx."""
    problemas = problemas_de(
        abas=[f"M{i}" for i in range(24)],
        setores=[{"nome": f"SETOR {i}"} for i in range(30)],
        colunas=[f"RUBRICA {i}" for i in range(30)],
    )
    assert any("molde inteiro" in p for p in problemas)


def test_um_ano_de_abas_no_tamanho_real_passa():
    """12 abas mensais com o porte do molde real não podem esbarrar no teto."""
    spec = molde.validar_spec(spec_valida(
        abas=[f"FMS MES {i}" for i in range(1, 13)],
        setores=[{"nome": f"SETOR {i}"} for i in range(16)],
        colunas=[f"RUBRICA {i}" for i in range(25)],
    ))
    assert len(spec["abas"]) == 12


def test_lista_absurda_e_recusada_sem_ser_percorrida():
    """Blindagem: 200 mil itens gerariam 200 mil mensagens de erro.

    Antes desta guarda, 1,4 MB de pedido viravam 27 MB de resposta e 98 MB de
    memória — o processo inteiro travava.
    """
    problemas = problemas_de(colunas=["CEF"] * (molde.MAX_ITENS_EXAMINADOS + 1))
    assert len(problemas) == 1
    assert "muito acima" in problemas[0]


def test_limite_normal_de_itens_da_mensagem_precisa():
    """Entre o limite e o teto de exame, a recusa diz o número exato."""
    problemas = problemas_de(colunas=[f"RUBRICA {i}" for i in range(molde.MAX_COLUNAS + 10)])
    assert any(f"Máximo de {molde.MAX_COLUNAS} colunas" in p for p in problemas)

    problemas = problemas_de(setores=[{"nome": f"S{i}"} for i in range(molde.MAX_SETORES + 5)])
    assert any(f"Máximo de {molde.MAX_SETORES} setores" in p for p in problemas)

    problemas = problemas_de(abas=[f"ABA {i}" for i in range(molde.MAX_ABAS + 3)])
    assert any(f"Máximo de {molde.MAX_ABAS} abas" in p for p in problemas)


def test_caractere_de_controle_e_recusado():
    assert any("controle" in p for p in problemas_de(colunas=["ARSEM\x07CEF"]))


def test_aba_com_apostrofo_nas_pontas_e_recusada():
    assert any("apóstrofo" in p for p in problemas_de(abas=["'MOLDE'"]))


def test_setor_aceita_texto_simples_alem_de_objeto():
    """Colar uma lista manda strings; a tela manda {nome, apelido}."""
    spec = molde.validar_spec(spec_valida(setores=["ADMINISTRATIVO", {"nome": "ACS"}]))
    assert spec["setores"] == [
        {"nome": "ADMINISTRATIVO", "apelido": ""},
        {"nome": "ACS", "apelido": ""},
    ]


def test_lista_de_problemas_tem_teto():
    repetidas = ["CEF"] * (molde.MAX_ITENS_EXAMINADOS - 50)
    problemas = problemas_de(colunas=repetidas)
    assert len(problemas) == molde.MAX_PROBLEMAS + 1
    assert problemas[-1].startswith("… e mais")


def test_todos_os_problemas_vem_de_uma_vez():
    problemas = problemas_de(colunas=["Tipo", "TOTAL X"], setores=[{"nome": "Tipo"}])
    assert len(problemas) >= 3


def test_spec_nao_dict_e_recusada():
    with pytest.raises(molde.ErroDeMolde):
        molde.validar_spec(["ARSEM"])


# ---------------------------------------------------------------------------
# Geração + verificação pelo motor real
# ---------------------------------------------------------------------------

def test_gera_e_o_motor_le_exatamente_o_desenhado():
    spec, wb, divergencias = molde.construir(spec_valida())
    assert divergencias == []

    ws = wb["MOLDE"]
    blocos = preenchimento.localizar_blocos_setores(ws)
    assert [b["setor"] for b in blocos] == ["ADMINISTRATIVO", "ACS"]
    assert preenchimento.listar_colunas_modelo(ws) == spec["colunas"]
    for bloco in blocos:
        assert set(bloco["linhas_tipo"]) == {"Mensal", "13º salário"}
        assert bloco["linha_total"] is not None


def test_total_do_evento_nao_vira_rubrica():
    _, wb, _ = molde.construir(spec_valida())
    assert "TOTAL DO EVENTO" not in preenchimento.listar_colunas_modelo(wb["MOLDE"])


def test_linha_total_e_formula_e_nunca_recebe_valor():
    _, wb, _ = molde.construir(spec_valida())
    ws = wb["MOLDE"]
    bloco = preenchimento.localizar_blocos_setores(ws)[0]
    coluna = bloco["colunas"]["ARSEM"]["coluna"]
    valor = ws.cell(row=bloco["linha_total"], column=coluna).value
    assert isinstance(valor, str) and valor.startswith("=SUM(")

    # A área de lançamento nasce zerada e editável (nunca fórmula).
    linha_mensal = bloco["linhas_tipo"]["Mensal"]
    assert ws.cell(row=linha_mensal, column=coluna).value == 0


def test_total_geral_soma_todos_os_blocos():
    _, wb, _ = molde.construir(spec_valida())
    ws = wb["MOLDE"]
    totais = [b["linha_total"] for b in preenchimento.localizar_blocos_setores(ws)]
    formula = next(
        ws.cell(row=linha, column=2).value
        for linha in range(1, ws.max_row + 1)
        if str(ws.cell(row=linha, column=1).value or "").strip() == "TOTAL GERAL"
    )
    for linha_total in totais:
        assert f"B{linha_total}" in formula


def test_abas_multiplas_saem_compativeis_entre_si():
    spec, wb, divergencias = molde.construir(spec_valida(abas=["FMS JANEIRO", "FMS FEVEREIRO"]))
    assert divergencias == []
    assert wb.sheetnames == ["FMS JANEIRO", "FMS FEVEREIRO"]
    primeira = preenchimento.listar_colunas_modelo(wb["FMS JANEIRO"])
    assert primeira == preenchimento.listar_colunas_modelo(wb["FMS FEVEREIRO"])
    # A competência 01/2026 tem de encontrar a aba de janeiro sozinha.
    assert preenchimento.localizar_aba_destino(wb, "01/2026") == "FMS JANEIRO"


def test_opcoes_desligadas_geram_molde_mais_simples():
    opcoes = {
        "coluna_total_evento": False,
        "linha_total_bloco": False,
        "linha_total_geral": False,
        "rodape_apelido": False,
    }
    spec, wb, divergencias = molde.construir(spec_valida(opcoes=opcoes))
    assert divergencias == []
    ws = wb["MOLDE"]
    blocos = preenchimento.localizar_blocos_setores(ws)
    assert [b["setor"] for b in blocos] == ["ADMINISTRATIVO", "ACS"]
    assert all(b["linha_total"] is None for b in blocos)
    assert preenchimento.listar_colunas_modelo(ws) == spec["colunas"]


def test_um_unico_tipo_de_folha_funciona():
    spec, wb, divergencias = molde.construir(spec_valida(tipos=["Mensal"]))
    assert divergencias == []
    bloco = preenchimento.localizar_blocos_setores(wb["MOLDE"])[0]
    assert set(bloco["linhas_tipo"]) == {"Mensal"}


def test_conferir_concorda_com_construir():
    """A prévia é o caminho rápido — não pode ser um veredito diferente.

    `conferir` dispensa estilos e confere uma aba só; o que o motor enxerga
    tem de ser exatamente o mesmo que no arquivo entregue.
    """
    entrada = spec_valida(abas=["JANEIRO", "FEVEREIRO"])
    spec_rapida, div_rapidas = molde.conferir(entrada)
    spec_cheia, _wb, div_cheias = molde.construir(entrada)
    assert spec_rapida == spec_cheia
    assert div_rapidas == div_cheias == []


def test_estilos_em_cache_nao_vazam_entre_workbooks(tmp_path):
    """`_estilo` é memoizado: o MESMO objeto Font/Fill vai para todo arquivo.

    Se o openpyxl mutasse esses objetos ao indexá-los, o segundo molde gerado
    sairia com as cores do primeiro. Ele indexa por valor — este teste trava
    essa premissa.
    """
    from openpyxl import load_workbook
    caminhos = []
    for i, colunas in enumerate([["ARSEM", "CEF"], ["IR", "IPASGO", "PREVIBELOS"], ["ARSEM"]]):
        spec = molde.validar_spec(spec_valida(colunas=colunas))
        caminho = tmp_path / f"w{i}.xlsx"
        molde.gerar_workbook(spec).save(caminho)
        caminhos.append(caminho)

    for caminho in caminhos:
        ws = load_workbook(caminho)["MOLDE"]
        bloco = preenchimento.localizar_blocos_setores(ws)[0]
        cabecalho = ws.cell(row=bloco["linha_cabecalho"], column=2)
        valor = ws.cell(row=bloco["linhas_tipo"]["Mensal"], column=2)
        total = ws.cell(row=bloco["linha_total"], column=2)
        assert cabecalho.fill.fgColor.rgb == "FF1F4E78"
        assert cabecalho.font.b is True and cabecalho.font.color.rgb == "FFFFFFFF"
        assert valor.fill.fgColor.rgb == "FFFFF2CC"
        assert valor.number_format == "\\R\\$\\ #,##0.00"
        assert total.fill.fgColor.rgb == "FFE2F0D9"


def test_estrutura_gravada_de_forma_atomica(perfis_tmp):
    """Nada de .tmp sobrando, e o conteúdo é o desenho completo."""
    spec = molde.validar_spec(spec_valida())
    molde.salvar_estrutura("saude", spec)
    pasta = molde.caminho_estrutura("saude").parent
    assert list(pasta.glob("*.tmp")) == []
    assert molde.carregar_estrutura("saude") == spec


def test_workbook_sem_estilo_tem_o_mesmo_conteudo():
    spec = molde.validar_spec(spec_valida())
    com = molde.gerar_workbook(spec)["MOLDE"]
    sem = molde.gerar_workbook(spec, com_estilo=False)["MOLDE"]

    assert preenchimento.listar_colunas_modelo(com) == preenchimento.listar_colunas_modelo(sem)
    blocos_com = preenchimento.localizar_blocos_setores(com)
    blocos_sem = preenchimento.localizar_blocos_setores(sem)
    assert [b["setor"] for b in blocos_com] == [b["setor"] for b in blocos_sem]
    assert [b["linhas_tipo"] for b in blocos_com] == [b["linhas_tipo"] for b in blocos_sem]
    assert [b["linha_total"] for b in blocos_com] == [b["linha_total"] for b in blocos_sem]


def test_verificacao_denuncia_molde_adulterado():
    # Se alguém quebrar a geração, a verificação tem de acusar — é a rede de
    # segurança que autoriza a gravação.
    spec, wb, _ = molde.construir(spec_valida())
    ws = wb["MOLDE"]
    bloco = preenchimento.localizar_blocos_setores(ws)[1]
    ws.cell(row=bloco["linha_setor"], column=1).value = "OUTRO NOME"
    assert molde.verificar_workbook(wb, spec) != []


# ---------------------------------------------------------------------------
# Engenharia reversa (.xlsx -> spec)
# ---------------------------------------------------------------------------

def test_round_trip_spec_xlsx_spec(tmp_path):
    spec = molde.validar_spec(spec_valida())
    caminho = tmp_path / "m.xlsx"
    molde.gerar_workbook(spec).save(caminho)
    assert molde.extrair_spec(caminho) == spec


def test_round_trip_preserva_opcoes_desligadas(tmp_path):
    spec = molde.validar_spec(spec_valida(
        setores=[{"nome": "ADMINISTRATIVO"}, {"nome": "ACS"}],
        opcoes={
            "coluna_total_evento": False,
            "linha_total_bloco": True,
            "linha_total_geral": False,
            "rodape_apelido": False,
        },
    ))
    caminho = tmp_path / "m.xlsx"
    molde.gerar_workbook(spec).save(caminho)
    assert molde.extrair_spec(caminho) == spec


def test_apelido_vive_no_rodape_e_some_sem_ele(tmp_path):
    """Sem rodapé, o apelido não tem onde morar no .xlsx.

    Quem guarda o desenho completo é o molde_estrutura.json do perfil — o
    .xlsx só carrega o que ele consegue representar.
    """
    spec = molde.validar_spec(spec_valida(opcoes={
        "coluna_total_evento": True,
        "linha_total_bloco": True,
        "linha_total_geral": True,
        "rodape_apelido": False,
    }))
    caminho = tmp_path / "m.xlsx"
    molde.gerar_workbook(spec).save(caminho)

    relido = molde.extrair_spec(caminho)
    assert [s["apelido"] for s in relido["setores"]] == ["", ""]
    assert [s["nome"] for s in relido["setores"]] == [s["nome"] for s in spec["setores"]]


def test_extrair_do_modelo_sintetico_do_conftest(modelo):
    spec = molde.extrair_spec(modelo)
    assert [s["nome"] for s in spec["setores"]] == ["ADMINISTRATIVO", "ACS"]
    assert "PREVIBELOS" in spec["colunas"]
    assert "TOTAL DO EVENTO" not in spec["colunas"]
    assert spec["opcoes"]["coluna_total_evento"] is True


def test_extrair_de_planilha_so_com_conferencia_falha_com_motivo(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = "CONFERÊNCIA_AUTOMAÇÃO"
    caminho = tmp_path / "so_conferencia.xlsx"
    wb.save(caminho)
    with pytest.raises(molde.ErroDeMolde) as erro:
        molde.extrair_spec(caminho)
    assert any("conferência" in p for p in erro.value.problemas)


def test_verificacao_aponta_setor_que_sumiu():
    spec, wb, _ = molde.construir(spec_valida())
    ws = wb["MOLDE"]
    # Apaga o rótulo "Tipo" do 2º bloco: sem ele o motor não enxerga o bloco.
    bloco = preenchimento.localizar_blocos_setores(ws)[1]
    ws.cell(row=bloco["linha_cabecalho"], column=1).value = None
    divergencias = molde.verificar_workbook(wb, spec)
    assert divergencias and "ausente" in divergencias[0]


def test_extrair_de_planilha_sem_blocos_falha_com_motivo(tmp_path):
    from openpyxl import Workbook
    caminho = tmp_path / "vazio.xlsx"
    Workbook().save(caminho)
    with pytest.raises(molde.ErroDeMolde) as erro:
        molde.extrair_spec(caminho)
    assert any("blocos de setor" in p for p in erro.value.problemas)


# ---------------------------------------------------------------------------
# Prévia e resumo
# ---------------------------------------------------------------------------

def test_grade_corta_e_avisa_quantos_setores_ficaram_de_fora():
    spec = molde.validar_spec(spec_valida(
        setores=[{"nome": f"SETOR {i}"} for i in range(6)]
    ))
    grade = molde.grade(spec, limite_setores=2)
    assert grade["setores_ocultos"] == 4
    # Sem todos os blocos, o TOTAL GERAL seria uma soma falsa: fica de fora.
    assert not any(linha["papel"] == "total_geral" for linha in grade["linhas"])


def test_grade_completa_mantem_total_geral():
    grade = molde.grade(molde.validar_spec(spec_valida()), limite_setores=10)
    assert grade["setores_ocultos"] == 0
    assert any(linha["papel"] == "total_geral" for linha in grade["linhas"])


def test_resumo_conta_celulas_de_valor():
    resumo = molde.resumo(molde.validar_spec(spec_valida()))
    assert resumo == {"abas": 1, "setores": 2, "colunas": 5, "tipos": 2,
                      "celulas_preenchiveis": 20}


# ---------------------------------------------------------------------------
# Persistência por perfil
# ---------------------------------------------------------------------------

@pytest.fixture
def perfis_tmp(tmp_path, monkeypatch):
    from services import perfis
    monkeypatch.setattr(perfis, "PERFIS_CONFIG_DIR", tmp_path / "config" / "perfis")
    monkeypatch.setattr(perfis, "PERFIS_MODELOS_DIR", tmp_path / "modelos" / "perfis")
    monkeypatch.setattr(perfis, "REGISTRO", tmp_path / "config" / "perfis.json")
    (tmp_path / "config").mkdir(parents=True, exist_ok=True)
    return tmp_path


def test_estrutura_salva_e_relida(perfis_tmp):
    spec = molde.validar_spec(spec_valida())
    molde.salvar_estrutura("saude", spec)
    assert molde.carregar_estrutura("saude") == spec
    assert molde.spec_inicial("saude") == (spec, "estrutura")


def test_estrutura_corrompida_nao_derruba_a_tela(perfis_tmp):
    molde.caminho_estrutura("saude").write_text("{ isto não é json", encoding="utf-8")
    assert molde.carregar_estrutura("saude") is None
    spec, origem = molde.spec_inicial("saude", "Secretaria da Saúde")
    assert origem == "branco"
    assert spec["setores"] == []


def test_spec_inicial_cai_para_o_molde_fixo_quando_nao_ha_desenho(perfis_tmp):
    from services import perfis
    caminho = perfis_tmp / "origem.xlsx"
    molde.gerar_workbook(molde.validar_spec(spec_valida())).save(caminho)
    perfis.definir_molde("saude", caminho, "RETENCAO.xlsx")

    spec, origem = molde.spec_inicial("saude")
    assert origem == "molde"
    assert [s["nome"] for s in spec["setores"]] == ["ADMINISTRATIVO", "ACS"]


def test_spec_em_branco_e_folha_limpa():
    spec = molde.spec_em_branco("Secretaria da Saúde (FMS)")
    assert spec["setores"] == [] and spec["colunas"] == []
    assert spec["abas"] == ["MOLDE"]
    assert spec["titulo"] == "Secretaria da Saúde (FMS)"
    # Folha em branco ainda não é um molde válido — falta o conteúdo.
    with pytest.raises(molde.ErroDeMolde):
        molde.validar_spec(spec)


# ---------------------------------------------------------------------------
# O que realmente importa: o molde construído processa
# ---------------------------------------------------------------------------

def test_molde_construido_processa_de_verdade(listagem, tmp_path):
    """Desenha na interface → gera → preenche a Listagem real → reconcilia."""
    spec = molde.validar_spec(spec_valida(
        colunas=["ARSEM", "CEF", "IR", "PREVIBELOS", "IPASGO"],
        setores=[{"nome": "ADMINISTRATIVO", "apelido": "ADM"}, {"nome": "ACS", "apelido": "ACS"}],
    ))
    caminho = tmp_path / "molde_construido.xlsx"
    molde.gerar_workbook(spec).save(caminho)

    lancamentos = extrair_lancamentos(listagem)["lancamentos"]
    cfg = mapeador.carregar_config_rubricas()
    wb = load_workbook(caminho)
    ws = wb["MOLDE"]

    colunas = preenchimento.listar_colunas_modelo(ws)
    mapa = {"5.0043.0000 - FMS, AGENTE COMUNITÁRIO DE SAÚDE - ACS": "ACS"}
    mapeador.aplicar_mapeamentos(lancamentos, mapa, cfg["regras"])
    mapeador.resolver_colunas(lancamentos, colunas, cfg["fora_de_escopo"], {})

    agregados, baldes = conferencia.agregar_lancamentos(lancamentos)
    blocos = preenchimento.localizar_blocos_setores(ws)
    preenchimento.limpar_area_lancamento(ws, blocos)
    relatorio = preenchimento.preencher_valores(ws, agregados, blocos)

    total_lido = conferencia.calcular_totais_lidos(lancamentos)
    rec = conferencia.reconciliar(total_lido, baldes, relatorio["pendencias_estrutura"])
    assert rec["confere"], f"diferença {rec['diferenca']}"
    assert relatorio["pendencias_estrutura"] == []

    # PREVIBELOS/ACS/Mensal recebe o lançamento MENSAL + o de FÉRIAS.
    acs = next(b for b in blocos if b["setor"] == "ACS")
    coluna = acs["colunas"]["PREVIBELOS"]["coluna"]
    escrito = ws.cell(row=acs["linhas_tipo"]["Mensal"], column=coluna).value
    assert abs(escrito - (698.39 + 694.06)) < 0.001

    # A fórmula do TOTAL do bloco continua intacta depois do preenchimento.
    total = ws.cell(row=acs["linha_total"], column=coluna).value
    assert isinstance(total, str) and total.startswith("=SUM(")

    # INSS é fora de escopo; a lotação nova não tem setor.
    assert rec["total_fora_escopo"] == Decimal("120.00")
    assert rec["total_setor_nao_mapeado"] == Decimal("100.00")
