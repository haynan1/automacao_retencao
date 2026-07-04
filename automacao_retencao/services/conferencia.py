"""Agregacao e conferencia com reconciliacao exata.

Invariante central (garantia "sem erros"):

    total_lido = total_preenchido
               + total_fora_escopo        (ignorado de proposito)
               + total_sem_vinculo         (rubrica sem coluna — precisa decisao)
               + total_setor_nao_mapeado   (lotacao sem setor)
               + total_folha_desconhecida  (tipo de folha nao reconhecido)

Se essa soma nao bater com o total lido, algo escapou — e o sistema mostra.
Todos os totais sao calculados na aplicacao com Decimal, sem depender do
Excel recalcular.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_ZERO = Decimal("0.00")


def agregar_lancamentos(lancamentos: list[dict]) -> tuple[dict, dict]:
    """Agrupa lancamentos preenchiveis por (setor, tipo, coluna) somando valor.

    So entram lancamentos com setor + tipo + coluna_destino e status 'ok'.
    Retorna (agregados, baldes) onde baldes classifica o que NAO foi agregado.
    """
    agregados: dict[tuple, Decimal] = {}
    baldes = {
        "preenchivel": _ZERO,
        "fora_escopo": _ZERO,
        "sem_vinculo": _ZERO,
        "setor_nao_mapeado": _ZERO,
        "folha_desconhecida": _ZERO,
    }

    for reg in lancamentos:
        valor = reg["valor"]
        setor = reg.get("setor_destino")
        tipo = reg.get("tipo_destino")
        coluna = reg.get("coluna_destino")
        status = reg.get("rubrica_status")

        if not reg.get("folha_reconhecida"):
            baldes["folha_desconhecida"] += valor
            continue
        if not setor:
            baldes["setor_nao_mapeado"] += valor
            continue
        if status == "fora_escopo":
            baldes["fora_escopo"] += valor
            continue
        if status != "ok" or not coluna or not tipo:
            baldes["sem_vinculo"] += valor
            continue

        chave = (setor, tipo, coluna)
        agregados[chave] = agregados.get(chave, _ZERO) + valor
        baldes["preenchivel"] += valor

    return agregados, baldes


def calcular_totais_lidos(lancamentos: list[dict]) -> Decimal:
    total = _ZERO
    for reg in lancamentos:
        total += reg["valor"]
    return total


def calcular_totais_agregados(agregados: dict) -> dict:
    """Totais por setor, por coluna (rubrica) e por tipo."""
    por_setor: dict[str, Decimal] = {}
    por_coluna: dict[str, Decimal] = {}
    por_tipo: dict[str, Decimal] = {}
    total = _ZERO
    for (setor, tipo, coluna), valor in agregados.items():
        por_setor[setor] = por_setor.get(setor, _ZERO) + valor
        por_coluna[coluna] = por_coluna.get(coluna, _ZERO) + valor
        por_tipo[tipo] = por_tipo.get(tipo, _ZERO) + valor
        total += valor
    return {
        "por_setor": dict(sorted(por_setor.items())),
        "por_coluna": dict(sorted(por_coluna.items())),
        "por_tipo": dict(sorted(por_tipo.items())),
        "total": total,
    }


def reconciliar(total_lido: Decimal, baldes: dict, pendencias_estrutura: list[dict]) -> dict:
    """Verifica a invariante de reconciliacao. Retorna diagnostico completo."""
    total_estrutura = sum((i.get("valor", _ZERO) for i in pendencias_estrutura), _ZERO)
    soma_baldes = (
        baldes["preenchivel"] + baldes["fora_escopo"] + baldes["sem_vinculo"]
        + baldes["setor_nao_mapeado"] + baldes["folha_desconhecida"]
    )
    # Itens que estavam preenchiveis mas nao acharam lugar na planilha migram
    # de 'preenchido' para 'estrutura'.
    total_preenchido = baldes["preenchivel"] - total_estrutura
    diferenca = total_lido - soma_baldes
    return {
        "total_lido": total_lido,
        "total_preenchido": total_preenchido,
        "total_fora_escopo": baldes["fora_escopo"],
        "total_sem_vinculo": baldes["sem_vinculo"],
        "total_setor_nao_mapeado": baldes["setor_nao_mapeado"],
        "total_folha_desconhecida": baldes["folha_desconhecida"],
        "total_estrutura": total_estrutura,
        "diferenca": diferenca,
        "confere": abs(diferenca) < Decimal("0.01"),
    }


# ---------------------------------------------------------------------------
# Aba de conferencia
# ---------------------------------------------------------------------------

_FONTE_TITULO = Font(bold=True, size=14, color="FFFFFF")
_FONTE_SECAO = Font(bold=True, size=11, color="FFFFFF")
_FONTE_CAB = Font(bold=True, color="FFFFFF")
_FILL_TITULO = PatternFill("solid", fgColor="0F172A")
_FILL_SECAO = PatternFill("solid", fgColor="1E293B")
_FILL_CAB = PatternFill("solid", fgColor="334155")
_FILL_ALERTA = PatternFill("solid", fgColor="7F1D1D")
_FILL_OK = PatternFill("solid", fgColor="14532D")
_BORDA = Border(*(Side(style="thin", color="CBD5E1"),) * 4)
_MOEDA = 'R$ #,##0.00'


def criar_aba_conferencia(wb, dados: dict) -> None:
    nome = "CONFERÊNCIA_AUTOMAÇÃO"
    if nome in wb.sheetnames:
        del wb[nome]
    ws = wb.create_sheet(title=nome)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16

    linha = _titulo(ws, 1, "CONFERÊNCIA DA AUTOMAÇÃO DE RETENÇÕES")
    linha += 1

    for rotulo, valor in (
        ("Processado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Competência", dados.get("competencia") or "—"),
        ("Arquivo de origem", dados.get("arquivo_origem") or "—"),
        ("Planilha modelo", dados.get("arquivo_modelo") or "—"),
        ("Aba preenchida", dados.get("aba_destino") or "—"),
    ):
        ws.cell(row=linha, column=1, value=rotulo).font = Font(bold=True)
        ws.cell(row=linha, column=2, value=valor)
        linha += 1
    linha += 1

    # Reconciliacao ------------------------------------------------------
    rec = dados["reconciliacao"]
    linha = _secao(ws, linha, "RECONCILIAÇÃO", alerta=not rec["confere"])
    status = "✔ CONFERE (bate ao centavo)" if rec["confere"] else f"✘ DIVERGÊNCIA de {rec['diferenca']}"
    c = ws.cell(row=linha, column=1, value=status)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = _FILL_OK if rec["confere"] else _FILL_ALERTA
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
    linha += 1
    for rotulo, chave in (
        ("Total lido (bruto)", "total_lido"),
        ("Total preenchido na planilha", "total_preenchido"),
        ("Fora de escopo (ignorado)", "total_fora_escopo"),
        ("Sem vínculo (rubrica sem coluna)", "total_sem_vinculo"),
        ("Setor não mapeado", "total_setor_nao_mapeado"),
        ("Tipo de folha desconhecido", "total_folha_desconhecida"),
    ):
        ws.cell(row=linha, column=1, value=rotulo).font = Font(bold=(chave in ("total_lido", "total_preenchido")))
        cc = ws.cell(row=linha, column=2, value=float(rec[chave]))
        cc.number_format = _MOEDA
        linha += 1
    linha += 1

    linha = _tabela_valor(ws, linha, "TOTAL POR SETOR", dados.get("por_setor", {}))
    linha = _tabela_valor(ws, linha, "TOTAL POR RUBRICA (COLUNA)", dados.get("por_coluna", {}))
    linha = _tabela_valor(ws, linha, "TOTAL POR TIPO", dados.get("por_tipo", {}))

    linha = _secao(ws, linha, "PENDÊNCIAS E OBSERVAÇÕES", alerta=True)
    linha = _tabela_contagem(ws, linha, "Lotações não mapeadas", dados.get("lotacoes_nao_mapeadas", {}))
    linha = _tabela_contagem(ws, linha, "Rubricas sem vínculo (precisam de decisão)", dados.get("rubricas_sem_vinculo", {}))
    linha = _tabela_contagem(ws, linha, "Rubricas fora de escopo (ignoradas)", dados.get("rubricas_fora_escopo", {}))
    linha = _tabela_contagem(ws, linha, "Tipos de folha desconhecidos", dados.get("folhas_desconhecidas", {}))

    estrutura = dados.get("pendencias_estrutura", [])
    if estrutura:
        ws.cell(row=linha, column=1, value="Itens sem lugar na planilha").font = Font(bold=True, italic=True)
        linha += 1
        for cab, col in (("Motivo", 1), ("Setor / Rubrica", 2), ("Valor", 3)):
            cel = ws.cell(row=linha, column=col, value=cab)
            cel.font, cel.fill = _FONTE_CAB, _FILL_CAB
        linha += 1
        for item in estrutura:
            ws.cell(row=linha, column=1, value=item.get("motivo"))
            ws.cell(row=linha, column=2, value=f"{item.get('setor')} / {item.get('rubrica')}")
            cc = ws.cell(row=linha, column=3, value=float(item.get("valor", _ZERO)))
            cc.number_format = _MOEDA
            linha += 1


def _titulo(ws, linha, texto):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
    c = ws.cell(row=linha, column=1, value=texto)
    c.font, c.fill = _FONTE_TITULO, _FILL_TITULO
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[linha].height = 26
    return linha + 1


def _secao(ws, linha, texto, alerta=False):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
    c = ws.cell(row=linha, column=1, value=texto)
    c.font = _FONTE_SECAO
    c.fill = _FILL_ALERTA if alerta else _FILL_SECAO
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[linha].height = 20
    return linha + 1


def _tabela_valor(ws, linha, titulo, dados: dict):
    linha = _secao(ws, linha, titulo)
    if not dados:
        ws.cell(row=linha, column=1, value="— nada a exibir —").font = Font(italic=True, color="94A3B8")
        return linha + 2
    for cab, col in (("Descrição", 1), ("Valor", 2)):
        cel = ws.cell(row=linha, column=col, value=cab)
        cel.font, cel.fill = _FONTE_CAB, _FILL_CAB
    linha += 1
    total = _ZERO
    for nome, valor in dados.items():
        ws.cell(row=linha, column=1, value=nome).border = _BORDA
        cc = ws.cell(row=linha, column=2, value=float(valor))
        cc.number_format, cc.border = _MOEDA, _BORDA
        total += valor
        linha += 1
    ws.cell(row=linha, column=1, value="TOTAL").font = Font(bold=True)
    cc = ws.cell(row=linha, column=2, value=float(total))
    cc.number_format, cc.font = _MOEDA, Font(bold=True)
    return linha + 2


def _tabela_contagem(ws, linha, titulo, dados: dict):
    ws.cell(row=linha, column=1, value=titulo).font = Font(bold=True, italic=True)
    linha += 1
    if not dados:
        ws.cell(row=linha, column=1, value="Nenhuma").font = Font(italic=True, color="16A34A")
        return linha + 2
    for cab, col in (("Item", 1), ("Ocorrências", 2)):
        cel = ws.cell(row=linha, column=col, value=cab)
        cel.font, cel.fill = _FONTE_CAB, _FILL_CAB
    linha += 1
    for nome, qtd in dados.items():
        ws.cell(row=linha, column=1, value=nome).border = _BORDA
        ws.cell(row=linha, column=2, value=qtd).border = _BORDA
        linha += 1
    return linha + 2
