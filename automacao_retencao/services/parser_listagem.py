"""Leitura e interpretacao do relatorio de origem (Listagem de Eventos).

O relatorio real e um XLSX paginado, com forte uso de CELULAS MESCLADAS:
- Banners, linhas de lotacao e de evento sao mesclados na largura toda (A:AC).
- Cada campo do lancamento e uma celula mesclada (ex.: Descricao ocupa Q:V no
  cabecalho, mas R:W nos dados — deslocada 1 coluna).

Por isso NAO da para confiar na letra da coluna. A estrategia e:
1. Resolver toda celula ao valor-ancora da sua mesclagem (segmentos da linha).
2. Localizar o cabecalho e guardar a FAIXA de colunas de cada campo.
3. Em cada linha, escolher o segmento com maior SOBREPOSICAO com a faixa do
   campo. Isso torna a leitura imune a deslocamentos por mesclagem.
4. Distinguir linha tabular de banner/lotacao/evento pela estrutura: se o
   segmento que cobre a Descricao tambem comeca na coluna A, a linha inteira
   e uma so mesclagem -> nao e lancamento.
"""

from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from .normalizador import normalizar_folha, normalizar_texto

# Lotacao no formato "5.0037.0000 - FMS, ...".
_RE_LOTACAO = re.compile(r"^\s*\d+\.\d+\.\d+\s+-\s+")
# Competencia MM/AAAA.
_RE_COMPETENCIA = re.compile(r"\b(0[1-9]|1[0-2])/(\d{4})\b")

# Rotulos possiveis de cada campo (normalizados) para localizar o cabecalho.
_ROTULOS = {
    "matricula": ("MAT.", "MAT", "MATRICULA"),
    "funcionario": ("FUNCIONARIO", "NOME", "SERVIDOR"),
    "cpf": ("CPF",),
    "folha": ("FOLHA",),
    "descricao": ("DESCRICAO", "DESCR", "EVENTO"),
    "base_calculo": ("BASE DE CALC.", "BASE DE CALCULO", "BASE CALC", "BASE"),
    "referencia": ("REFER.", "REFERENCIA", "REFER"),
    "valor": ("VALOR",),
}

# Texto que, em linha mesclada de largura total, deve ser ignorado.
_BANNER_MARCADORES = (
    "MUNICIPIO", "FUNDO MUNICIPAL", "LISTAGEM DE EVENTOS", "EMITIDO EM",
    "PAGINA", "FP038", "CNPJ",
)
# Descricoes que nao sao lancamento (ruido tabular / resumo).
_DESC_RUIDO = {"QUANTIDADE:", "QUANTIDADE", "RESUMO", "TOTAL", "TOTAL GERAL", "DESCRICAO"}


def _texto(v) -> str:
    return "" if v is None else str(v).strip()


def limpar_valor(valor) -> Decimal:
    """Converte valor de celula (numero ou string BR) em Decimal(2 casas)."""
    if valor is None or valor == "":
        return Decimal("0.00")
    if isinstance(valor, (int, float)):
        try:
            return Decimal(str(valor)).quantize(Decimal("0.01"))
        except InvalidOperation:
            return Decimal("0.00")

    texto = str(valor).strip()
    if not texto:
        return Decimal("0.00")

    negativo = False
    if texto.startswith("(") and texto.endswith(")"):
        negativo, texto = True, texto[1:-1]
    if texto.startswith("-"):
        negativo, texto = True, texto[1:]

    texto = re.sub(r"[^\d.,]", "", texto)
    if not texto:
        return Decimal("0.00")
    if "," in texto:  # formato brasileiro: ponto milhar, virgula decimal
        texto = texto.replace(".", "").replace(",", ".")

    try:
        numero = Decimal(texto).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")
    return -numero if negativo else numero


def _eh_numero(texto: str) -> bool:
    """True se o texto e apenas um numero (usado para descartar ruido)."""
    if not texto:
        return False
    t = texto.replace(".", "").replace(",", "").replace("-", "").strip()
    return t.isdigit()


# ---------------------------------------------------------------------------
# Modelo de segmentos (mesclagens) por linha
# ---------------------------------------------------------------------------

def _indexar_mesclagens(ws) -> dict:
    """row -> lista de (min_col, max_col, valor_ancora) das mesclagens dessa linha."""
    por_linha: dict[int, list] = {}
    for mc in ws.merged_cells.ranges:
        ancora = ws.cell(mc.min_row, mc.min_col).value
        por_linha.setdefault(mc.min_row, []).append((mc.min_col, mc.max_col, ancora))
    return por_linha


def _segmentos_da_linha(ws, r: int, max_col: int, mescladas: dict) -> list:
    """Segmentos da linha r: mesclagens (ancora) + celulas isoladas nao vazias."""
    segs = list(mescladas.get(r, []))
    cobertas = set()
    for mn, mx, _ in segs:
        cobertas.update(range(mn, mx + 1))
    for c in range(1, max_col + 1):
        if c in cobertas:
            continue
        v = ws.cell(r, c).value
        if v is not None and str(v).strip() != "":
            segs.append((c, c, v))
    return segs


def _melhor_sobreposicao(segs: list, faixa: tuple):
    """Segmento (mn,mx,val) com maior sobreposicao com a faixa (smin,smax)."""
    if not faixa:
        return None
    smin, smax = faixa
    melhor, melhor_ov = None, 0
    for mn, mx, val in segs:
        ov = min(mx, smax) - max(mn, smin) + 1
        if ov > melhor_ov:
            melhor, melhor_ov = (mn, mx, val), ov
    return melhor


# ---------------------------------------------------------------------------
# Deteccao de cabecalho, competencia e lotacoes
# ---------------------------------------------------------------------------

def detectar_competencia(ws) -> str | None:
    for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
        for celula in row:
            if celula is None:
                continue
            m = _RE_COMPETENCIA.search(str(celula))
            if m:
                return f"{m.group(1)}/{m.group(2)}"
    return None


def detectar_cabecalhos(ws, max_col: int, mescladas: dict) -> tuple[int | None, dict]:
    """Localiza a linha de cabecalho e as FAIXAS (min,max) de cada campo.

    Retorna (linha_cabecalho, {campo: (min_col, max_col)}). A linha valida e a
    primeira que contenha, ao mesmo tempo, os rotulos de Descricao e Valor.
    """
    for r in range(1, min(ws.max_row, 200) + 1):
        segs = _segmentos_da_linha(ws, r, max_col, mescladas)
        faixas: dict[str, tuple] = {}
        for mn, mx, val in segs:
            rot = normalizar_texto(_texto(val))
            if not rot:
                continue
            for campo, opcoes in _ROTULOS.items():
                if campo in faixas:
                    continue
                if rot in opcoes:
                    faixas[campo] = (mn, mx)
        if "descricao" in faixas and "valor" in faixas:
            return r, faixas
    return None, {}


# ---------------------------------------------------------------------------
# Extracao principal
# ---------------------------------------------------------------------------

def extrair_lancamentos(caminho_xlsx: str | Path) -> dict:
    """Le o relatorio e devolve a estrutura de trabalho.

    Retorna:
        {
          "competencia": "06/2026" | None,
          "aba": "Page1",
          "lotacoes": [...],
          "lancamentos": [ {registro}, ... ],
          "faixas": {campo: (min,max)},
          "total_relatorio": Decimal | None,   # total proprio do relatorio (cross-check)
        }
    """
    caminho = Path(caminho_xlsx)
    # Precisamos das mesclagens -> nao usar read_only.
    wb = load_workbook(caminho, data_only=True)
    ws = wb.worksheets[0]
    max_col = ws.max_column or 30

    mescladas = _indexar_mesclagens(ws)
    competencia = detectar_competencia(ws)
    linha_cab, faixas = detectar_cabecalhos(ws, max_col, mescladas)

    if linha_cab is None:
        wb.close()
        return {
            "competencia": competencia, "aba": ws.title, "lotacoes": [],
            "lancamentos": [], "faixas": {}, "total_relatorio": None,
        }

    faixa_desc = faixas["descricao"]
    lancamentos: list[dict] = []
    lotacao_atual: str | None = None
    em_resumo = False
    total_relatorio: Decimal | None = None
    banners: list[str] = []  # cabecalhos (orgao/secretaria) para autodeteccao

    for r in range(1, ws.max_row + 1):
        segs = _segmentos_da_linha(ws, r, max_col, mescladas)
        if not segs:
            continue

        seg_desc = _melhor_sobreposicao(segs, faixa_desc)
        seg_folha = _melhor_sobreposicao(segs, faixas.get("folha"))

        # Linha NAO tabular (banner / lotacao / evento / rodape / resumo):
        # nao ha descricao, OU o MESMO segmento mesclado cobre Descricao e Folha
        # (mesclagem larga de titulo), OU o segmento comeca na coluna A.
        nao_tabular = (
            seg_desc is None
            or seg_desc[0] <= 1
            or (seg_folha is not None and seg_folha[:2] == seg_desc[:2])
        )
        if nao_tabular:
            texto = _texto(seg_desc[2]) if seg_desc else _texto(segs[0][2])
            if _RE_LOTACAO.match(texto):
                lotacao_atual = texto
            elif normalizar_texto(texto).startswith("RESUMO"):
                em_resumo = True
            elif texto and not texto.startswith("Evento:") and len(banners) < 12:
                # cabecalho (municipio/orgao/secretaria) — util p/ autodeteccao
                if texto not in banners:
                    banners.append(texto)
            continue

        if em_resumo:
            continue

        # --- linha tabular: candidata a lancamento ---
        descricao = _texto(seg_desc[2])
        norm_desc = normalizar_texto(descricao)
        if not descricao or norm_desc in _DESC_RUIDO or _eh_numero(descricao):
            continue

        folha_bruta = _texto(seg_folha[2]) if seg_folha else ""
        if not folha_bruta:  # sem folha nao e lancamento valido
            continue

        info_folha = normalizar_folha(folha_bruta)
        registro = {
            "lotacao_original": lotacao_atual or "",
            "setor_destino": None,
            "matricula": _texto(_valor_campo(segs, faixas.get("matricula"))),
            "funcionario": _texto(_valor_campo(segs, faixas.get("funcionario"))),
            "cpf": _texto(_valor_campo(segs, faixas.get("cpf"))),
            "folha": folha_bruta,
            "tipo_destino": info_folha["tipo_destino"],
            "folha_reconhecida": info_folha["reconhecida"],
            "observacao": info_folha["observacao"],
            "descricao_original": descricao,
            "rubrica_destino": None,
            "coluna_destino": None,
            "base_calculo": limpar_valor(_valor_campo(segs, faixas.get("base_calculo"))),
            "referencia": limpar_valor(_valor_campo(segs, faixas.get("referencia"))),
            "valor": limpar_valor(_valor_campo(segs, faixas.get("valor"))),
            "linha_origem": r,
        }
        lancamentos.append(registro)

    wb.close()

    lotacoes: list[str] = []
    for reg in lancamentos:
        lot = reg["lotacao_original"]
        if lot and lot not in lotacoes:
            lotacoes.append(lot)

    return {
        "competencia": competencia,
        "aba": ws.title,
        "lotacoes": lotacoes,
        "lancamentos": lancamentos,
        "faixas": faixas,
        "banners": banners,
        "total_relatorio": total_relatorio,
    }


def _valor_campo(segs: list, faixa: tuple | None):
    """Valor do campo cuja faixa de colunas e 'faixa', por maior sobreposicao."""
    seg = _melhor_sobreposicao(segs, faixa) if faixa else None
    return seg[2] if seg else None
