"""Preenchimento da planilha modelo de Retencao.

Nada aqui e feito por coordenada fixa. Os blocos de setor, as colunas de
rubrica e as linhas de tipo (Mensal / 13o salario) sao localizados
dinamicamente pela estrutura da planilha, preservando formulas, estilos,
mesclagens e dashboard.
"""

from __future__ import annotations

from decimal import Decimal

from openpyxl.worksheet.worksheet import Worksheet

from .normalizador import normalizar_texto

# Tipos que o app preenche (a linha TOTAL nunca e tocada — mantem formula).
_TIPOS_CANONICOS = {
    "MENSAL": "Mensal",
    "13": "13º salário",
    "13O SALARIO": "13º salário",
    "13 SALARIO": "13º salário",
    "DECIMO TERCEIRO": "13º salário",
}

# Meses por numero para casar a competencia com o nome da aba (FMS JUNHO).
_MESES_PT = {
    "01": "JANEIRO", "02": "FEVEREIRO", "03": "MARÇO", "04": "ABRIL",
    "05": "MAIO", "06": "JUNHO", "07": "JULHO", "08": "AGOSTO",
    "09": "SETEMBRO", "10": "OUTUBRO", "11": "NOVEMBRO", "12": "DEZEMBRO",
}


def localizar_aba_destino(wb, competencia: str | None) -> str | None:
    """Sugere a aba de destino a partir da competencia (MM/AAAA).

    06/2026 -> procura uma aba cujo nome contenha 'JUNHO'. Retorna o nome
    exato da aba encontrada, ou None se nao houver correspondencia.
    """
    if not competencia or "/" not in competencia:
        return None
    mes = competencia.split("/")[0]
    nome_mes = _MESES_PT.get(mes)
    if not nome_mes:
        return None

    alvo = normalizar_texto(nome_mes)
    for nome in wb.sheetnames:
        if alvo in normalizar_texto(nome):
            return nome
    return None


def _tipo_canonico(texto: str) -> str | None:
    """Converte o rotulo de uma linha em 'Mensal'/'13º salário' ou None."""
    norm = normalizar_texto(texto)
    if not norm:
        return None
    if "MENSAL" in norm:
        return "Mensal"
    if "13" in norm or "DECIMO TERCEIRO" in norm:
        return "13º salário"
    return None


def localizar_colunas_rubricas(ws: Worksheet, linha_cabecalho: int) -> dict:
    """Mapeia rubrica_normalizada -> {'coluna': idx, 'rotulo': texto}.

    Lê a linha de cabecalho do bloco ('Tipo | ARSEM | CEF | ...') e indexa
    cada coluna pela rubrica normalizada, ignorando a coluna 'Tipo'.
    """
    colunas: dict[str, dict] = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=linha_cabecalho, column=col).value
        rotulo = "" if valor is None else str(valor).strip()
        norm = normalizar_texto(rotulo)
        # Ignora a coluna 'Tipo' e colunas de total (ex.: 'TOTAL DO EVENTO'),
        # que nao sao rubricas e nunca devem receber lancamento.
        if not norm or norm == "TIPO" or norm.startswith("TOTAL"):
            continue
        # Se a mesma rubrica aparecer duas vezes, mantem a primeira ocorrencia.
        colunas.setdefault(norm, {"coluna": col, "rotulo": rotulo})
    return colunas


def listar_colunas_modelo(ws: Worksheet) -> list[str]:
    """Lista, na ordem das colunas, os rotulos de rubrica do modelo.

    As colunas sao identicas em todos os blocos; usamos o bloco com mais
    colunas como referencia. Esta lista e a FONTE DA VERDADE para o vinculo
    rubrica -> coluna (nunca se escreve numa coluna que nao exista).
    """
    blocos = localizar_blocos_setores(ws)
    if not blocos:
        return []
    ref = max(blocos, key=lambda b: len(b["colunas"]))
    itens = sorted(ref["colunas"].values(), key=lambda x: x["coluna"])
    return [i["rotulo"] for i in itens]


def localizar_blocos_setores(ws: Worksheet) -> list[dict]:
    """Localiza dinamicamente todos os blocos de setor da aba.

    Um bloco e identificado quando a coluna A tem um nome (setor) e a
    linha imediatamente abaixo comeca com 'Tipo'. A partir dai, as linhas
    'Mensal', '13º salário' e 'TOTAL' sao localizadas ate o fim do bloco.
    """
    blocos: list[dict] = []
    max_row = ws.max_row

    linha = 1
    while linha <= max_row:
        valor_a = ws.cell(row=linha, column=1).value
        setor = "" if valor_a is None else str(valor_a).strip()

        if setor:
            proximo = ws.cell(row=linha + 1, column=1).value
            if proximo is not None and normalizar_texto(str(proximo)) == "TIPO":
                linha_cabecalho = linha + 1
                colunas = localizar_colunas_rubricas(ws, linha_cabecalho)
                linhas_tipo, linha_total, fim = _localizar_linhas_tipo(
                    ws, linha_cabecalho, max_row
                )
                blocos.append(
                    {
                        "setor": setor,
                        "setor_norm": normalizar_texto(setor),
                        "linha_setor": linha,
                        "linha_cabecalho": linha_cabecalho,
                        "colunas": colunas,
                        "linhas_tipo": linhas_tipo,
                        "linha_total": linha_total,
                    }
                )
                linha = max(fim, linha_cabecalho) + 1
                continue
        linha += 1

    return blocos


def _localizar_linhas_tipo(ws, linha_cabecalho: int, max_row: int):
    """A partir do cabecalho, encontra as linhas Mensal/13º e a linha TOTAL.

    Retorna (linhas_tipo: dict, linha_total: int|None, ultima_linha: int).
    Para de varrer ao encontrar TOTAL ou o inicio de um novo bloco.
    """
    linhas_tipo: dict[str, int] = {}
    linha_total = None
    linha = linha_cabecalho + 1
    limite = min(max_row, linha_cabecalho + 12)  # janela de seguranca

    while linha <= limite:
        rotulo = ws.cell(row=linha, column=1).value
        texto = "" if rotulo is None else str(rotulo).strip()
        norm = normalizar_texto(texto)

        if not texto:
            linha += 1
            continue
        if "TOTAL" in norm:
            linha_total = linha
            break

        tipo = _tipo_canonico(texto)
        if tipo and tipo not in linhas_tipo:
            linhas_tipo[tipo] = linha
        elif not tipo:
            # Rotulo desconhecido: provavelmente inicio de outro bloco.
            break
        linha += 1

    return linhas_tipo, linha_total, linha


def limpar_area_lancamento(ws: Worksheet, blocos: list[dict]) -> None:
    """Zera as celulas de valor (Mensal / 13º) antes de preencher, para
    garantir idempotencia. Nao toca em TOTAL nem em nenhuma formula."""
    for bloco in blocos:
        for tipo, linha in bloco["linhas_tipo"].items():
            for info in bloco["colunas"].values():
                celula = ws.cell(row=linha, column=info["coluna"])
                # So limpa celulas que NAO contem formula.
                if not (isinstance(celula.value, str) and celula.value.startswith("=")):
                    celula.value = None


def preencher_valores(ws: Worksheet, agregados: dict, blocos: list[dict]) -> dict:
    """Escreve os valores agregados nas celulas corretas.

    'agregados' e um dict {(setor, tipo, rubrica): Decimal}.

    Retorna um relatorio de preenchimento:
        {
          "preenchidos": [ {setor, tipo, rubrica, valor, celula} ],
          "total_preenchido": Decimal,
          "pendencias_estrutura": [ {motivo, setor, tipo, rubrica, valor} ],
        }
    """
    indice_setor = {b["setor_norm"]: b for b in blocos}

    preenchidos: list[dict] = []
    pendencias: list[dict] = []
    total = Decimal("0.00")

    for (setor, tipo, rubrica), valor in agregados.items():
        bloco = indice_setor.get(normalizar_texto(setor))
        if bloco is None:
            pendencias.append(_pend("setor não encontrado na planilha", setor, tipo, rubrica, valor))
            continue

        linha = bloco["linhas_tipo"].get(tipo)
        if linha is None:
            pendencias.append(_pend(f"linha '{tipo}' não encontrada no setor", setor, tipo, rubrica, valor))
            continue

        info_col = bloco["colunas"].get(normalizar_texto(rubrica))
        if info_col is None:
            pendencias.append(_pend("rubrica sem coluna na planilha", setor, tipo, rubrica, valor))
            continue

        celula = ws.cell(row=linha, column=info_col["coluna"])
        # Seguranca extra: jamais sobrescrever uma formula.
        if isinstance(celula.value, str) and celula.value.startswith("="):
            pendencias.append(_pend("célula contém fórmula (protegida)", setor, tipo, rubrica, valor))
            continue

        celula.value = float(valor)
        total += valor
        preenchidos.append(
            {
                "setor": setor,
                "tipo": tipo,
                "rubrica": rubrica,
                "valor": valor,
                "celula": f"{celula.coordinate}",
            }
        )

    return {
        "preenchidos": preenchidos,
        "total_preenchido": total,
        "pendencias_estrutura": pendencias,
    }


def _pend(motivo, setor, tipo, rubrica, valor) -> dict:
    return {"motivo": motivo, "setor": setor, "tipo": tipo, "rubrica": rubrica, "valor": valor}


def preservar_formulas(wb) -> None:
    """Força o Excel a recalcular todas as formulas ao abrir o arquivo.

    openpyxl nao recalcula formulas; garantimos que o Excel o faca na
    abertura, mantendo dashboard e totais coerentes.
    """
    wb.calculation.fullCalcOnLoad = True
    try:
        wb.calculation.forceFullCalc = True
    except Exception:  # atributo pode variar entre versoes do openpyxl
        pass
    wb.calculation.calcMode = "auto"
