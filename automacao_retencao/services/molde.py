"""Construtor de molde — a estrutura da planilha de Retencao vira DADO.

Ate aqui o molde so existia como um .xlsx editado a mao no Excel. Qualquer
escorregao de digitacao — uma coluna repetida, um setor chamado "Tipo", um
rotulo que comeca com "=" — nao dava erro: dava preenchimento silenciosamente
perdido, tres telas adiante. Aqui a estrutura e uma *spec* validada e o .xlsx
passa a ser um artefato gerado a partir dela.

Duas garantias, nesta ordem:

1. `validar_spec` recusa tudo que o motor de preenchimento descartaria em
   silencio: coluna duplicada apos normalizacao, rotulo iniciado por "TOTAL",
   aba com "CONFER" no nome, texto que o Excel leria como formula.
2. `verificar_workbook` abre o arquivo gerado com o MESMO codigo do
   processamento (`preenchimento.localizar_blocos_setores` e
   `listar_colunas_modelo`) e confere que ele devolve exatamente os setores,
   as colunas e os tipos desenhados. Nada e salvo sem passar por isso.

O layout existe uma unica vez, em `_linhas_layout`. O escritor de .xlsx e a
previa da tela consomem a mesma sequencia de linhas — a previa nao e uma
reproducao aproximada da planilha, e a propria planilha.
"""

from __future__ import annotations

import json
import logging
import os
import re
from functools import lru_cache
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import perfis, preenchimento
from .normalizador import normalizar_texto

log = logging.getLogger("automacao_retencao")

SPEC_VERSAO = 1

# Os unicos tipos que `preenchimento.tipo_canonico` sabe reconhecer. Oferecer
# outros na interface seria prometer um preenchimento que nunca aconteceria.
TIPOS_SUPORTADOS = ("Mensal", "13º salário")

ROTULO_TIPO = "Tipo"
ROTULO_TOTAL = "TOTAL"
ROTULO_TOTAL_EVENTO = "TOTAL DO EVENTO"
ROTULO_TOTAL_GERAL = "TOTAL GERAL"
LINHA_SECAO = "LANÇAMENTO DAS RETENÇÕES POR SETOR — BASE ZERADA"

# Limites de sanidade, calibrados por medicao — nao por chute. Referencia: o
# molde real da secretaria tem 16 setores x 23 colunas em 2 abas (~2,9 mil
# celulas), conferido em 27 ms e gravado em 175 ms.
#
#   MAX_BLOCO_CELULAS limita setores x colunas de UMA aba. E o que pesa na
#   previa, que roda a cada digitacao: 6 mil mantem a resposta perto de 1 s.
#   MAX_CELULAS limita o arquivo inteiro. E o que pesa na gravacao, acao
#   pontual: 100 mil mantem a escrita abaixo de ~8 s.
MAX_ABAS = 24
MAX_SETORES = 200
MAX_COLUNAS = 200
MAX_BLOCO_CELULAS = 6_000
MAX_CELULAS = 100_000

# Teto de itens que chegam a ser examinados um a um. Sem ele, uma lista com
# 200 mil repeticoes gerava 200 mil mensagens de erro — 1,4 MB de pedido
# viravam 27 MB de resposta e 98 MB de memoria. Acima disto a lista e recusada
# inteira, sem ser percorrida.
MAX_ITENS_EXAMINADOS = 500
# Ninguem corrige quarenta pontos de uma vez; alem disso e ruido.
MAX_PROBLEMAS = 40

_LIMITE_TEXTO = 200
_LIMITE_ABA = 31  # limite do proprio Excel

# Caracteres de controle que o Excel recusa dentro de uma celula.
_RE_CONTROLE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
# Excel reservou estes caracteres para a sintaxe de referencia entre abas.
_RE_ABA_INVALIDA = re.compile(r"[\[\]:*?/\\]")
# Prefixos que transformam texto em formula/comando ao abrir a planilha.
_PREFIXOS_FORMULA = ("=", "+", "@")

_FORMATO_MOEDA = "\\R\\$\\ #,##0.00"

# Paleta e pesos copiados do molde real da secretaria: o arquivo gerado sai
# indistinguivel do que o Departamento Pessoal ja imprime e assina.
_C_ESCURO = "FF1F4E78"
_C_MEDIO = "FF375A7F"
_C_CLARO = "FFD9EAF7"
_C_ROTULO = "FFF3F4F6"
_C_VALOR = "FFFFF2CC"
_C_TOTAL = "FFE2F0D9"


class ErroDeMolde(ValueError):
    """Spec invalida. `problemas` traz a lista completa, nao so a primeira."""

    def __init__(self, problemas: list[str]):
        self.problemas = list(problemas)
        super().__init__("; ".join(self.problemas))


# ---------------------------------------------------------------------------
# Spec: normalizacao e validacao
# ---------------------------------------------------------------------------

def _texto(valor, limite: int = _LIMITE_TEXTO) -> str:
    """Achata para uma linha unica, colapsa espacos e corta no limite."""
    if valor is None:
        return ""
    bruto = str(valor).replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return " ".join(bruto.split())[:limite]


def _checar_texto(rotulo: str, texto: str, problemas: list[str]) -> None:
    """Recusa o que o Excel leria como formula ou nao conseguiria guardar."""
    if not texto:
        return
    if texto[0] in _PREFIXOS_FORMULA:
        problemas.append(
            f"{rotulo}: não pode começar com “{texto[0]}” — o Excel leria como fórmula."
        )
    if _RE_CONTROLE.search(texto):
        problemas.append(f"{rotulo}: contém caractere de controle não suportado pelo Excel.")


def _itens_examinaveis(bruto, rotulo: str, problemas: list[str]) -> list | None:
    """Devolve a lista a examinar, ou None se ela for grande demais para isso.

    Recusar sem percorrer e o que impede uma lista absurda de virar uma
    avalanche de mensagens de erro.
    """
    itens = bruto if isinstance(bruto, list) else []
    if len(itens) > MAX_ITENS_EXAMINADOS:
        problemas.append(
            f"{rotulo}: {len(itens)} itens recebidos — muito acima do que um molde comporta. "
            "Revise a lista antes de enviar."
        )
        return None
    return itens


def spec_em_branco(nome_secretaria: str = "") -> dict:
    """Folha em branco: um esqueleto valido para comecar do zero na tela."""
    return {
        "versao": SPEC_VERSAO,
        "titulo": _texto(nome_secretaria) or "MODELO DE RETENÇÕES",
        "subtitulo": "MODELO DE RETENÇÕES — BASE ZERADA",
        "abas": ["MOLDE"],
        "colunas": [],
        "setores": [],
        "tipos": list(TIPOS_SUPORTADOS),
        "opcoes": {
            "coluna_total_evento": True,
            "linha_total_bloco": True,
            "linha_total_geral": True,
            "rodape_apelido": True,
        },
    }


def _validar_abas(bruto, problemas: list[str]) -> list[str]:
    itens = _itens_examinaveis(bruto, "Abas", problemas)
    if itens is None:
        return []

    abas: list[str] = []
    vistas: dict[str, str] = {}
    for item in itens:
        nome = _texto(item, _LIMITE_ABA)
        if not nome:
            continue
        rotulo = f"Aba “{nome}”"
        _checar_texto(rotulo, nome, problemas)
        if _RE_ABA_INVALIDA.search(nome):
            problemas.append(f"{rotulo}: o Excel não aceita os caracteres  [ ] : * ? / \\  em nome de aba.")
            continue
        if nome.startswith("'") or nome.endswith("'"):
            problemas.append(f"{rotulo}: nome de aba não pode começar nem terminar com apóstrofo.")
            continue
        if "CONFER" in normalizar_texto(nome):
            problemas.append(
                f"{rotulo}: nomes com “CONFER” são reservados para a aba de conferência "
                "e seriam ignorados no processamento."
            )
            continue
        chave = nome.casefold()
        if chave in vistas:
            problemas.append(f"{rotulo}: aba repetida (o Excel não diferencia maiúsculas de minúsculas).")
            continue
        vistas[chave] = nome
        abas.append(nome)

    if not abas:
        problemas.append("Defina ao menos uma aba para a planilha.")
    elif len(abas) > MAX_ABAS:
        problemas.append(f"Máximo de {MAX_ABAS} abas por molde (recebidas {len(abas)}).")
    return abas


def _validar_colunas(bruto, problemas: list[str]) -> list[str]:
    itens = _itens_examinaveis(bruto, "Colunas de rubrica", problemas)
    if itens is None:
        return []

    colunas: list[str] = []
    vistas: dict[str, str] = {}
    for item in itens:
        nome = _texto(item)
        if not nome:
            continue
        rotulo = f"Coluna “{nome}”"
        _checar_texto(rotulo, nome, problemas)
        norm = normalizar_texto(nome)
        if norm == "TIPO":
            problemas.append(f"{rotulo}: “Tipo” é o rótulo da primeira coluna e não pode ser uma rubrica.")
            continue
        if norm.startswith("TOTAL"):
            problemas.append(
                f"{rotulo}: colunas iniciadas por “TOTAL” são tratadas como totalizadores "
                "e nunca receberiam lançamento. Use a opção “coluna TOTAL DO EVENTO”."
            )
            continue
        if norm in vistas:
            problemas.append(
                f"{rotulo}: repete “{vistas[norm]}” (acentos e caixa não diferenciam) — "
                "a segunda coluna nunca seria preenchida."
            )
            continue
        vistas[norm] = nome
        colunas.append(nome)

    if not colunas:
        problemas.append("Defina ao menos uma coluna de rubrica.")
    elif len(colunas) > MAX_COLUNAS:
        problemas.append(f"Máximo de {MAX_COLUNAS} colunas por molde (recebidas {len(colunas)}).")
    return colunas


def _validar_setores(bruto, com_total_bloco: bool, problemas: list[str]) -> list[dict]:
    itens = _itens_examinaveis(bruto, "Setores", problemas)
    if itens is None:
        return []

    setores: list[dict] = []
    vistos: dict[str, str] = {}
    for item in itens:
        if isinstance(item, dict):
            nome = _texto(item.get("nome"))
            apelido = _texto(item.get("apelido"), 60)
        else:
            nome, apelido = _texto(item), ""
        if not nome:
            continue

        rotulo = f"Setor “{nome}”"
        _checar_texto(rotulo, nome, problemas)
        _checar_texto(f"Apelido de {rotulo}", apelido, problemas)
        norm = normalizar_texto(nome)
        if norm == "TIPO":
            problemas.append(f"{rotulo}: “Tipo” é palavra reservada da estrutura do bloco.")
            continue
        if norm.startswith("TOTAL"):
            problemas.append(
                f"{rotulo}: nomes iniciados por “TOTAL” se confundem com a linha de total "
                "e quebrariam a leitura do bloco."
            )
            continue
        # Sem a linha TOTAL fechando o bloco, a varredura de tipos segue ate a
        # proxima linha rotulada — e um setor que "parece" um tipo (contem
        # MENSAL ou 13) faria o bloco seguinte ser engolido.
        if not com_total_bloco and preenchimento.tipo_canonico(nome) is not None:
            problemas.append(
                f"{rotulo}: com a linha TOTAL desligada, um setor que contenha “MENSAL” ou “13” "
                "se confunde com uma linha de tipo. Ligue a linha TOTAL ou renomeie o setor."
            )
            continue
        if norm in vistos:
            problemas.append(
                f"{rotulo}: repete “{vistos[norm]}” (acentos e caixa não diferenciam) — "
                "os lançamentos cairiam sempre no primeiro bloco."
            )
            continue
        vistos[norm] = nome
        setores.append({"nome": nome, "apelido": apelido})

    if not setores:
        problemas.append("Defina ao menos um setor.")
    elif len(setores) > MAX_SETORES:
        problemas.append(f"Máximo de {MAX_SETORES} setores por molde (recebidos {len(setores)}).")
    return setores


def _validar_tipos(bruto, problemas: list[str]) -> list[str]:
    escolhidos = [t for t in TIPOS_SUPORTADOS if t in (bruto or [])]
    if not escolhidos:
        problemas.append(
            "Selecione ao menos um tipo de folha (Mensal e/ou 13º salário) — "
            "são os únicos que o preenchimento sabe destinar."
        )
    return escolhidos


def contar_celulas(spec: dict) -> int:
    """Celulas que o .xlsx tera. Serve de orcamento antes de gerar.

    Conta pelo teto (todas as opcoes ligadas): setor, cabecalho, os tipos,
    TOTAL, rodape e a linha em branco — mais os 3 banners e o total geral.
    """
    por_bloco = len(spec["tipos"]) + 5
    linhas = 3 + len(spec["setores"]) * por_bloco + 2
    colunas = len(spec["colunas"]) + 2
    return len(spec["abas"]) * linhas * colunas


def _mil(n: int) -> str:
    return f"{n:,}".replace(",", ".")


def _problemas_de_tamanho(spec: dict) -> list[str]:
    """Recusa moldes que travariam a tela ou a gravacao.

    Separa as duas grandezas porque elas doem em momentos diferentes: o bloco
    (setores x colunas) pesa na previa, que roda a cada tecla; o total de
    celulas pesa na gravacao do arquivo.
    """
    problemas: list[str] = []

    bloco = len(spec["setores"]) * len(spec["colunas"])
    if bloco > MAX_BLOCO_CELULAS:
        problemas.append(
            f"{len(spec['setores'])} setores × {len(spec['colunas'])} colunas dão "
            f"{_mil(bloco)} células por aba, acima do limite de {_mil(MAX_BLOCO_CELULAS)}. "
            "Reduza setores ou colunas."
        )

    celulas = contar_celulas(spec)
    if celulas > MAX_CELULAS:
        problemas.append(
            f"O molde inteiro ficaria com {_mil(celulas)} células, acima do limite de "
            f"{_mil(MAX_CELULAS)}. Reduza a quantidade de abas."
        )
    return problemas


def validar_spec(bruto: dict) -> dict:
    """Normaliza e valida a spec vinda da interface.

    Devolve a spec canonica ou levanta `ErroDeMolde` com TODOS os problemas
    encontrados — quem esta montando o molde corrige tudo de uma vez.
    """
    if not isinstance(bruto, dict):
        raise ErroDeMolde(["Estrutura do molde inválida."])

    problemas: list[str] = []
    opcoes_bruto = bruto.get("opcoes") if isinstance(bruto.get("opcoes"), dict) else {}
    opcoes = {
        "coluna_total_evento": bool(opcoes_bruto.get("coluna_total_evento", True)),
        "linha_total_bloco": bool(opcoes_bruto.get("linha_total_bloco", True)),
        "linha_total_geral": bool(opcoes_bruto.get("linha_total_geral", True)),
        "rodape_apelido": bool(opcoes_bruto.get("rodape_apelido", True)),
    }

    titulo = _texto(bruto.get("titulo"), 300)
    subtitulo = _texto(bruto.get("subtitulo"), 300)
    _checar_texto("Título", titulo, problemas)
    _checar_texto("Subtítulo", subtitulo, problemas)

    spec = {
        "versao": SPEC_VERSAO,
        "titulo": titulo,
        "subtitulo": subtitulo,
        "abas": _validar_abas(bruto.get("abas"), problemas),
        "colunas": _validar_colunas(bruto.get("colunas"), problemas),
        "setores": _validar_setores(bruto.get("setores"), opcoes["linha_total_bloco"], problemas),
        "tipos": _validar_tipos(bruto.get("tipos"), problemas),
        "opcoes": opcoes,
    }

    if not problemas:
        problemas.extend(_problemas_de_tamanho(spec))

    if problemas:
        raise ErroDeMolde(_resumir(problemas))
    return spec


def _resumir(problemas: list[str]) -> list[str]:
    """Corta a lista de problemas num tamanho que alguem consegue ler."""
    if len(problemas) <= MAX_PROBLEMAS:
        return problemas
    restantes = len(problemas) - MAX_PROBLEMAS
    return problemas[:MAX_PROBLEMAS] + [f"… e mais {_mil(restantes)} ponto(s) do mesmo tipo."]


# ---------------------------------------------------------------------------
# Layout — fonte unica da verdade (xlsx e previa consomem a mesma sequencia)
# ---------------------------------------------------------------------------

def _cel(texto, papel: str, formula: bool = False, moeda: bool = False) -> dict:
    return {"texto": texto, "papel": papel, "formula": formula, "moeda": moeda}


def _linhas_layout(spec: dict) -> list[dict]:
    """Descreve a aba inteira, linha a linha, sem tocar em openpyxl.

    Cada linha: {"n": numero, "papel": str, "mesclar": bool, "celulas": [...]}.
    As formulas ja saem com as referencias reais porque a numeracao das linhas
    e conhecida no momento em que cada uma e emitida.
    """
    colunas = spec["colunas"]
    opcoes = spec["opcoes"]
    tem_total_evento = opcoes["coluna_total_evento"]
    n_rubricas = len(colunas)
    ultima_rubrica = get_column_letter(1 + n_rubricas)
    n_cols = 1 + n_rubricas + (1 if tem_total_evento else 0)

    linhas: list[dict] = []
    atual = 1

    def emitir(papel: str, celulas: list[dict], mesclar: bool = False) -> None:
        nonlocal atual
        linhas.append({"n": atual, "papel": papel, "mesclar": mesclar, "celulas": celulas})
        atual += 1

    def faixa(texto: str, papel: str) -> list[dict]:
        return [_cel(texto, papel)] + [_cel("", papel) for _ in range(n_cols - 1)]

    def cabecalho() -> list[dict]:
        celulas = [_cel(ROTULO_TIPO, "cabecalho")]
        celulas += [_cel(c, "cabecalho") for c in colunas]
        if tem_total_evento:
            celulas.append(_cel(ROTULO_TOTAL_EVENTO, "cabecalho"))
        return celulas

    emitir("titulo", faixa(spec["titulo"], "titulo"), mesclar=True)
    emitir("subtitulo", faixa(spec["subtitulo"], "subtitulo"), mesclar=True)
    emitir("secao", faixa(LINHA_SECAO, "secao"), mesclar=True)

    totais_de_bloco: list[int] = []
    linhas_de_valor: list[int] = []

    for setor in spec["setores"]:
        emitir("setor", faixa(setor["nome"], "setor"), mesclar=True)
        emitir("cabecalho", cabecalho())

        primeira_valor = atual
        for tipo in spec["tipos"]:
            celulas = [_cel(tipo, "rotulo")]
            celulas += [_cel(0, "valor", moeda=True) for _ in colunas]
            if tem_total_evento:
                celulas.append(
                    _cel(f"=SUM(B{atual}:{ultima_rubrica}{atual})", "soma", formula=True, moeda=True)
                )
            linhas_de_valor.append(atual)
            emitir("valor", celulas)
        ultima_valor = atual - 1

        if opcoes["linha_total_bloco"]:
            celulas = [_cel(ROTULO_TOTAL, "rotulo")]
            for i in range(n_rubricas):
                letra = get_column_letter(2 + i)
                celulas.append(
                    _cel(f"=SUM({letra}{primeira_valor}:{letra}{ultima_valor})",
                         "soma", formula=True, moeda=True)
                )
            if tem_total_evento:
                celulas.append(
                    _cel(f"=SUM(B{atual}:{ultima_rubrica}{atual})", "soma", formula=True, moeda=True)
                )
            totais_de_bloco.append(atual)
            emitir("total", celulas)

        if opcoes["rodape_apelido"]:
            apelido = setor["apelido"] or setor["nome"]
            celulas = [_cel("", "vazio")]
            celulas += [_cel(apelido, "rodape") for _ in colunas]
            if tem_total_evento:
                celulas.append(_cel("", "vazio"))
            emitir("rodape", celulas)

        emitir("vazio", [_cel("", "vazio") for _ in range(n_cols)])

    if opcoes["linha_total_geral"]:
        # O cabecalho repetido rotula as colunas do total geral, que fica longe
        # do ultimo bloco em moldes grandes.
        emitir("cabecalho", cabecalho())
        alvos = totais_de_bloco or linhas_de_valor
        celulas = [_cel(ROTULO_TOTAL_GERAL, "total_geral")]
        for i in range(n_rubricas):
            letra = get_column_letter(2 + i)
            refs = ",".join(f"{letra}{n}" for n in alvos)
            celulas.append(_cel(f"=SUM({refs})", "soma", formula=True, moeda=True))
        if tem_total_evento:
            celulas.append(
                _cel(f"=SUM(B{atual}:{ultima_rubrica}{atual})", "soma", formula=True, moeda=True)
            )
        emitir("total_geral", celulas)

    return linhas


# ---------------------------------------------------------------------------
# Geracao do .xlsx
# ---------------------------------------------------------------------------

@lru_cache(maxsize=None)
def _estilo(papel: str) -> tuple[Font, PatternFill | None, Alignment]:
    """Estilos por papel, construidos uma unica vez.

    openpyxl aceita o mesmo objeto de estilo em varias celulas. Sem este
    cache, um molde de 280 mil celulas alocava 280 mil trincas Font/Fill/
    Alignment — e levava quase um minuto para ser gerado.
    """
    branco_bold = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
    if papel == "titulo":
        return (Font(name="Calibri", size=16, bold=True, color="FFFFFFFF"),
                PatternFill("solid", fgColor=_C_ESCURO),
                Alignment(horizontal="center", vertical="center", wrap_text=True))
    if papel == "subtitulo":
        return (Font(name="Calibri", size=12, bold=True),
                PatternFill("solid", fgColor=_C_CLARO),
                Alignment(horizontal="center", vertical="center", wrap_text=True))
    if papel == "secao":
        return (branco_bold, PatternFill("solid", fgColor=_C_ESCURO),
                Alignment(horizontal="center", vertical="center", wrap_text=True))
    if papel == "setor":
        return (branco_bold, PatternFill("solid", fgColor=_C_MEDIO),
                Alignment(horizontal="left", vertical="center", wrap_text=True))
    if papel == "cabecalho":
        return (branco_bold, PatternFill("solid", fgColor=_C_ESCURO),
                Alignment(horizontal="center", vertical="center", wrap_text=True))
    if papel == "rotulo":
        return (Font(name="Calibri", size=10, bold=True),
                PatternFill("solid", fgColor=_C_ROTULO),
                Alignment(horizontal="left", vertical="center", wrap_text=True))
    if papel == "valor":
        return (Font(name="Calibri", size=10),
                PatternFill("solid", fgColor=_C_VALOR),
                Alignment(horizontal="right", vertical="center", wrap_text=True))
    if papel == "soma":
        return (Font(name="Calibri", size=10, bold=True),
                PatternFill("solid", fgColor=_C_TOTAL),
                Alignment(horizontal="right", vertical="center", wrap_text=True))
    if papel == "rodape":
        return (branco_bold, PatternFill("solid", fgColor=_C_MEDIO),
                Alignment(horizontal="center", vertical="center", wrap_text=True))
    if papel == "total_geral":
        return (branco_bold, PatternFill("solid", fgColor=_C_ESCURO),
                Alignment(horizontal="left", vertical="center", wrap_text=True))
    return (Font(name="Calibri", size=10), None,
            Alignment(horizontal="left", vertical="center"))


def _escrever_aba(ws, spec: dict, linhas: list[dict], com_estilo: bool = True) -> None:
    """Escreve a aba. Sem estilo, sai o mesmo conteudo — so mais rapido.

    O motor de preenchimento le apenas `.value` e a malha de mesclagens; fonte,
    cor e formato nao mudam uma virgula do que ele enxerga. Por isso a
    verificacao dispensa os estilos, e so o arquivo entregue os recebe.
    """
    n_cols = len(linhas[0]["celulas"])
    ultima_coluna = get_column_letter(n_cols)

    for linha in linhas:
        for indice, celula in enumerate(linha["celulas"], start=1):
            texto = celula["texto"]
            if texto == "" and not com_estilo:
                continue  # celula vazia sem estilo nao precisa nem existir
            alvo = ws.cell(row=linha["n"], column=indice)
            alvo.value = texto if texto != "" else None
            if not com_estilo:
                continue
            fonte, preenchido, alinhamento = _estilo(celula["papel"])
            alvo.font = fonte
            if preenchido is not None:
                alvo.fill = preenchido
            alvo.alignment = alinhamento
            if celula["moeda"]:
                alvo.number_format = _FORMATO_MOEDA

        if linha["mesclar"] and n_cols > 1:
            ws.merge_cells(f"A{linha['n']}:{ultima_coluna}{linha['n']}")

    if not com_estilo:
        return

    for linha in linhas:
        if linha["papel"] == "titulo":
            ws.row_dimensions[linha["n"]].height = 22
        elif linha["papel"] in ("secao", "cabecalho"):
            ws.row_dimensions[linha["n"]].height = 36

    ws.column_dimensions["A"].width = 22
    for indice in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(indice)].width = 15
    if spec["opcoes"]["coluna_total_evento"]:
        ws.column_dimensions[ultima_coluna].width = 19.5

    ws.page_setup.orientation = "landscape"
    ws.sheet_view.zoomScale = 115
    # Congela os banners, o 1o cabecalho e a coluna dos rotulos: em molde
    # longo e largo, ninguem perde de vista qual coluna e qual linha esta lendo.
    ws.freeze_panes = "B6"


def gerar_workbook(spec: dict, com_estilo: bool = True) -> Workbook:
    """Monta o .xlsx completo a partir de uma spec ja validada."""
    linhas = _linhas_layout(spec)
    wb = Workbook()
    wb.remove(wb.active)
    for nome in spec["abas"]:
        _escrever_aba(wb.create_sheet(title=nome), spec, linhas, com_estilo)
    preenchimento.preservar_formulas(wb)
    return wb


# ---------------------------------------------------------------------------
# Verificacao — o molde e lido de volta pelo motor de preenchimento
# ---------------------------------------------------------------------------

def verificar_workbook(wb, spec: dict) -> list[str]:
    """Le o workbook gerado com o codigo REAL do processamento.

    Se o motor nao enxergar exatamente os setores, colunas e tipos desenhados,
    a divergencia aparece aqui — antes de o molde virar o padrao da secretaria.
    """
    divergencias: list[str] = []
    esperados = [s["nome"] for s in spec["setores"]]

    for nome in spec["abas"]:
        if nome not in wb.sheetnames:
            divergencias.append(f"Aba “{nome}” não foi gerada.")
            continue
        ws = wb[nome]
        blocos = preenchimento.localizar_blocos_setores(ws)
        lidos = [b["setor"] for b in blocos]
        if lidos != esperados:
            divergencias.append(
                f"Aba “{nome}”: o motor leu {len(lidos)} setor(es) contra {len(esperados)} "
                f"desenhado(s). Diferença em “{_primeira_diferenca(lidos, esperados)}”."
            )
            continue

        colunas = preenchimento.listar_colunas_modelo(ws)
        if colunas != spec["colunas"]:
            divergencias.append(
                f"Aba “{nome}”: o motor leu {len(colunas)} coluna(s) contra "
                f"{len(spec['colunas'])} desenhada(s). "
                f"Diferença em “{_primeira_diferenca(colunas, spec['colunas'])}”."
            )
            continue

        for bloco in blocos:
            faltando = [t for t in spec["tipos"] if t not in bloco["linhas_tipo"]]
            if faltando:
                divergencias.append(
                    f"Aba “{nome}”, setor “{bloco['setor']}”: linha(s) {', '.join(faltando)} "
                    "não foram reconhecidas."
                )
            if spec["opcoes"]["linha_total_bloco"] and bloco["linha_total"] is None:
                divergencias.append(
                    f"Aba “{nome}”, setor “{bloco['setor']}”: linha TOTAL não foi reconhecida."
                )

    return divergencias


def _primeira_diferenca(lidos: list[str], esperados: list[str]) -> str:
    """Primeiro item onde as duas listas divergem — a pista util no erro."""
    for lido, esperado in zip(lidos, esperados):
        if lido != esperado:
            return f"{esperado} → {lido}"
    faltando = esperados[len(lidos):]
    if faltando:
        return f"{faltando[0]} → ausente"
    sobrando = lidos[len(esperados):]
    return f"a mais: {sobrando[0]}" if sobrando else "—"


def construir(bruto: dict) -> tuple[dict, Workbook, list[str]]:
    """Valida, gera e verifica o arquivo REAL. Para download e gravacao.

    Devolve (spec, workbook estilizado, divergencias).
    """
    spec = validar_spec(bruto)
    wb = gerar_workbook(spec)
    return spec, wb, verificar_workbook(wb, spec)


def conferir(bruto: dict) -> tuple[dict, list[str]]:
    """Valida e verifica sem produzir o arquivo entregue. Para a previa.

    Dispensa os estilos e confere uma aba so — as demais sao copias da mesma
    sequencia de linhas. Mantem a garantia (o motor relendo a estrutura) a um
    custo compativel com rodar a cada tecla digitada.
    """
    spec = validar_spec(bruto)
    amostra = dict(spec, abas=spec["abas"][:1])
    wb = gerar_workbook(amostra, com_estilo=False)
    return spec, verificar_workbook(wb, amostra)


# ---------------------------------------------------------------------------
# Previa para a tela
# ---------------------------------------------------------------------------

def grade(spec: dict, limite_setores: int = 2) -> dict:
    """Recorte da planilha para desenhar na tela, sem gerar arquivo.

    Usa `_linhas_layout` — o mesmo layout do .xlsx — e corta nos primeiros
    `limite_setores` blocos para nao despejar um molde inteiro no navegador.
    """
    ocultos = max(0, len(spec["setores"]) - limite_setores)
    recorte = spec if not ocultos else dict(spec, setores=spec["setores"][:limite_setores])
    bruto = _linhas_layout(recorte)

    # Com o molde cortado, o TOTAL GERAL somaria só os blocos visíveis — seria
    # uma prévia mentirosa. Fora ele e o cabeçalho que o rotula.
    if ocultos and bruto and bruto[-1]["papel"] == "total_geral":
        bruto = bruto[:-2]

    linhas = [
        {
            "papel": linha["papel"],
            "mesclar": linha["mesclar"],
            "celulas": [
                {"texto": _texto_previa(c), "papel": c["papel"], "formula": c["formula"]}
                for c in linha["celulas"]
            ],
        }
        for linha in bruto
    ]

    return {
        "linhas": linhas,
        "setores_ocultos": ocultos,
        "colunas": len(spec["colunas"]) + (2 if spec["opcoes"]["coluna_total_evento"] else 1),
    }


def _texto_previa(celula: dict) -> str:
    if celula["texto"] == 0 and not celula["formula"]:
        return "0,00"
    return str(celula["texto"])


def resumo(spec: dict) -> dict:
    """Numeros que a tela mostra sobre o molde desenhado."""
    return {
        "abas": len(spec["abas"]),
        "setores": len(spec["setores"]),
        "colunas": len(spec["colunas"]),
        "tipos": len(spec["tipos"]),
        "celulas_preenchiveis": (
            len(spec["abas"]) * len(spec["setores"]) * len(spec["colunas"]) * len(spec["tipos"])
        ),
    }


# ---------------------------------------------------------------------------
# Engenharia reversa: .xlsx existente -> spec editavel
# ---------------------------------------------------------------------------

def extrair_spec(caminho: str | Path) -> dict:
    """Le um molde .xlsx e devolve a spec correspondente.

    Serve para editar na tela um molde que ja existe, em vez de obrigar a
    redesenhar tudo. O que nao for reconhecido volta como padrao — a spec
    resultante e sempre valida ou levanta `ErroDeMolde` com o motivo.
    """
    wb = load_workbook(caminho, data_only=False)
    try:
        abas = [n for n in wb.sheetnames if "CONFER" not in normalizar_texto(n)]
        if not abas:
            raise ErroDeMolde(["A planilha não tem nenhuma aba de molde (só conferência)."])

        ws = wb[abas[0]]
        blocos = preenchimento.localizar_blocos_setores(ws)
        if not blocos:
            raise ErroDeMolde([
                "A planilha não apresenta blocos de setor no formato esperado "
                "(nome do setor seguido de uma linha “Tipo | rubricas…”)."
            ])

        colunas = preenchimento.listar_colunas_modelo(ws)

        tipos: list[str] = []
        for tipo in TIPOS_SUPORTADOS:
            if any(tipo in bloco["linhas_tipo"] for bloco in blocos):
                tipos.append(tipo)

        setores = [{"nome": b["setor"], "apelido": _apelido_do_bloco(ws, b)} for b in blocos]

        primeira_linha_bloco = blocos[0]["linha_setor"]
        titulo, subtitulo = _titulos_acima(ws, primeira_linha_bloco)

        return validar_spec({
            "titulo": titulo,
            "subtitulo": subtitulo,
            "abas": abas,
            "colunas": colunas,
            "setores": setores,
            "tipos": tipos or list(TIPOS_SUPORTADOS),
            "opcoes": {
                "coluna_total_evento": _tem_coluna_total(ws, blocos[0]),
                "linha_total_bloco": all(b["linha_total"] is not None for b in blocos),
                "linha_total_geral": _tem_total_geral(ws),
                "rodape_apelido": any(s["apelido"] for s in setores),
            },
        })
    finally:
        wb.close()


def _apelido_do_bloco(ws, bloco: dict) -> str:
    """Le o rodape do bloco (linha logo abaixo do TOTAL, coluna B)."""
    if not bloco["linhas_tipo"]:
        return ""
    base = bloco["linha_total"] or max(bloco["linhas_tipo"].values())
    valor = ws.cell(row=base + 1, column=2).value
    return _texto(valor, 60) if valor is not None else ""


def _tem_coluna_total(ws, bloco: dict) -> bool:
    for coluna in range(1, ws.max_column + 1):
        valor = ws.cell(row=bloco["linha_cabecalho"], column=coluna).value
        if valor is not None and normalizar_texto(str(valor)).startswith("TOTAL"):
            return True
    return False


def _tem_total_geral(ws) -> bool:
    alvo = normalizar_texto(ROTULO_TOTAL_GERAL)
    for linha in range(1, ws.max_row + 1):
        valor = ws.cell(row=linha, column=1).value
        if valor is not None and normalizar_texto(str(valor)) == alvo:
            return True
    return False


def _titulos_acima(ws, primeira_linha_bloco: int) -> tuple[str, str]:
    """Recupera titulo e subtitulo das linhas de banner acima do 1o bloco."""
    encontrados: list[str] = []
    secao = normalizar_texto(LINHA_SECAO)
    for linha in range(1, primeira_linha_bloco):
        valor = ws.cell(row=linha, column=1).value
        texto = _texto(valor, 300)
        if not texto or normalizar_texto(texto) == secao:
            continue
        encontrados.append(texto)
    titulo = encontrados[0] if encontrados else ""
    subtitulo = encontrados[1] if len(encontrados) > 1 else ""
    return titulo, subtitulo


# ---------------------------------------------------------------------------
# Persistencia da estrutura por perfil
# ---------------------------------------------------------------------------

def caminho_estrutura(slug: str) -> Path:
    return perfis.caminho_molde(slug).with_name("molde_estrutura.json")


def salvar_estrutura(slug: str, spec: dict) -> None:
    """Guarda a spec ao lado do molde: o desenho e reabrivel e auditavel.

    Escrita atomica: grava num temporario e so entao troca. Uma queda no meio
    da escrita deixaria um JSON pela metade — e o desenho de dezenas de setores
    voltaria como 'folha em branco' na proxima abertura.
    """
    caminho = caminho_estrutura(slug)
    temporario = caminho.with_suffix(".json.tmp")
    with temporario.open("w", encoding="utf-8") as fh:
        json.dump(spec, fh, ensure_ascii=False, indent=2)
    os.replace(temporario, caminho)


def carregar_estrutura(slug: str) -> dict | None:
    """Le a spec salva do perfil. None se nao houver ou se estiver corrompida."""
    caminho = caminho_estrutura(slug)
    if not caminho.exists():
        return None
    try:
        with caminho.open("r", encoding="utf-8") as fh:
            return validar_spec(json.load(fh))
    except (OSError, json.JSONDecodeError, ErroDeMolde) as exc:
        log.warning("molde_estrutura.json do perfil '%s' inválido: %s", slug, exc)
        return None


def spec_inicial(slug: str, nome_secretaria: str = "") -> tuple[dict, str]:
    """Spec que a tela abre e de onde ela veio ('estrutura' | 'molde' | 'branco')."""
    salva = carregar_estrutura(slug)
    if salva is not None:
        return salva, "estrutura"

    if perfis.existe_molde(slug):
        try:
            return extrair_spec(perfis.caminho_molde(slug)), "molde"
        except (ErroDeMolde, OSError, ValueError) as exc:
            log.warning("Não foi possível ler o molde do perfil '%s': %s", slug, exc)

    return spec_em_branco(nome_secretaria), "branco"
